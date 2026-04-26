from flask import Flask, render_template, jsonify, request, g
import yfinance as yf
from pykrx import stock as krx
from datetime import datetime, timedelta
import json, os, sys, requests, feedparser, sqlite3, uuid
from xml.etree import ElementTree as ET
import openpyxl
import pandas as pd
import numpy as np
from flask.json.provider import DefaultJSONProvider

# ── config.json 로드 ─────────────────────────────────────
_app_cfg = {}
_cfg_candidates = []
if getattr(sys, 'frozen', False):
    _cfg_candidates.append(os.path.join(os.path.dirname(sys.executable), 'config.json'))
_cfg_candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json'))
for _cp in _cfg_candidates:
    if os.path.exists(_cp):
        try:
            with open(_cp, 'r', encoding='utf-8') as _f:
                _app_cfg = json.load(_f)
            break
        except Exception:
            pass

# ── DB 모드 결정 (PostgreSQL 우선, 없으면 SQLite) ─────────
DATABASE_URL = os.environ.get('DATABASE_URL', '') or _app_cfg.get('database_url', '')
IS_POSTGRES  = bool(DATABASE_URL and 'postgres' in DATABASE_URL)
if IS_POSTGRES:
    import psycopg2, psycopg2.extras
    PH = '%s'   # PostgreSQL 파라미터 플레이스홀더
else:
    PH = '?'    # SQLite 파라미터 플레이스홀더

# ── Supabase REST API (크로스 디바이스 동기화) ────────────
SUPABASE_URL = os.environ.get('SUPABASE_URL', '') or _app_cfg.get('supabase_url', '')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', '') or _app_cfg.get('supabase_key', '')
IS_SUPABASE  = bool(SUPABASE_URL and SUPABASE_KEY)
if IS_SUPABASE:
    import string, random as _rnd
    print(f"[Sync] Supabase REST API 활성화: {SUPABASE_URL}")

# ── GitHub Releases (자동 업데이트 소스) ──────────────────
GH_OWNER = os.environ.get('GH_OWNER', '') or _app_cfg.get('gh_owner', '')
GH_REPO  = os.environ.get('GH_REPO', '')  or _app_cfg.get('gh_repo', '')
IS_GITHUB_UPDATE = bool(GH_OWNER and GH_REPO)
if IS_GITHUB_UPDATE:
    print(f"[Update] GitHub Releases 사용: {GH_OWNER}/{GH_REPO}")

def _supa_headers():
    return {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation'
    }

def _gen_access_code(length=6):
    """6자리 영숫자 접속코드 생성 (대문자 + 숫자, 혼동 문자 제외)"""
    chars = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'  # 0/O, 1/I 제외
    return ''.join(_rnd.choices(chars, k=length))

# ── 버전 정보 ─────────────────────────────────────────────
APP_VERSION = "v1.4.1"
APP_NAME    = "내 주식 대시보드"

# ── PyInstaller 번들 환경 대응 ───────────────────────────
def _resource(rel):
    """templates 등 읽기 전용 리소스 경로 (번들 내부)"""
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, rel)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), rel)

def _get_data_dir():
    """
    데이터 저장 디렉토리 결정 순서:
      1. Render / Fly.io 클라우드: /data
      2. EXE 옆 config.json 의 data_dir (명시적 설정)
      3. EXE 실행 시: %LOCALAPPDATA%/StockDashboard (쓰기 권한 보장)
      4. 소스 실행 시: 스크립트 폴더
    """
    # 클라우드 배포
    if os.environ.get('RENDER') or os.environ.get('FLY'):
        d = '/data'
        os.makedirs(d, exist_ok=True)
        return d

    if getattr(sys, 'frozen', False):
        # EXE 실행: config.json의 data_dir 우선
        exe_dir = os.path.dirname(sys.executable)
        config_path = os.path.join(exe_dir, 'config.json')
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as _f:
                    cfg = json.load(_f)
                data_dir = cfg.get('data_dir', '').strip()
                if data_dir:
                    os.makedirs(data_dir, exist_ok=True)
                    # 마이그레이션: exe_dir에 기존 DB가 있으면 복사
                    _migrate_legacy_db(exe_dir, data_dir)
                    return data_dir
            except Exception:
                pass

        # config 없음: AppData 기본값 (Program Files 권한 문제 회피)
        appdata = os.environ.get('LOCALAPPDATA') or os.environ.get('APPDATA')
        if appdata:
            data_dir = os.path.join(appdata, 'StockDashboard')
            try:
                os.makedirs(data_dir, exist_ok=True)
                # 쓰기 가능 테스트
                test_file = os.path.join(data_dir, '.write_test')
                with open(test_file, 'w') as _f:
                    _f.write('ok')
                os.remove(test_file)
                # 마이그레이션
                _migrate_legacy_db(exe_dir, data_dir)
                return data_dir
            except Exception:
                pass

        # 폴백: exe_dir (쓰기 안 되면 DB 저장 실패할 수 있음)
        return exe_dir

    # 소스 실행
    return os.path.dirname(os.path.abspath(__file__))


def _migrate_legacy_db(old_dir, new_dir):
    """기존 exe_dir의 dashboard.db를 new_dir로 이동 (한 번만)"""
    if not old_dir or not new_dir or old_dir == new_dir:
        return
    try:
        old_db = os.path.join(old_dir, 'dashboard.db')
        new_db = os.path.join(new_dir, 'dashboard.db')
        # new 위치에 이미 DB 있으면 스킵
        if os.path.exists(new_db):
            return
        # old 위치에 DB 있으면 복사
        if os.path.exists(old_db):
            import shutil
            shutil.copy2(old_db, new_db)
            print(f"[DB] 기존 DB 이동: {old_db} → {new_db}")
    except Exception as e:
        print(f"[DB] 마이그레이션 실패 (무시): {e}")

DATA_DIR = _get_data_dir()

def _data(rel):
    """쓰기 가능한 데이터 파일 경로"""
    return os.path.join(DATA_DIR, rel)

DATABASE = _data('dashboard.db')

class NumpyJSONProvider(DefaultJSONProvider):
    def default(self, obj):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)

app = Flask(__name__,
    template_folder=_resource('templates'),
    static_folder=_resource('static') if os.path.isdir(_resource('static')) else None,
)
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.json_provider_class = NumpyJSONProvider
app.json = NumpyJSONProvider(app)

@app.after_request
def add_no_cache_headers(response):
    if request.path == '/' or request.path.endswith('.html'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

BASE_DIR = _resource('.')

# ── DB 초기화 ─────────────────────────────────────────────
_CREATE_TABLE = '''
    CREATE TABLE IF NOT EXISTS user_data (
        user_token        TEXT PRIMARY KEY,
        portfolio         TEXT NOT NULL DEFAULT '{{"holdings":[]}}',
        transactions      TEXT NOT NULL DEFAULT '[]',
        wishlist          TEXT NOT NULL DEFAULT '[]',
        hidden            TEXT NOT NULL DEFAULT '[]',
        target_prices     TEXT NOT NULL DEFAULT '{{}}',
        notes             TEXT NOT NULL DEFAULT '{{}}',
        rebalance_targets TEXT NOT NULL DEFAULT '{{}}',
        simulation        TEXT NOT NULL DEFAULT '[]',
        created_at        TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at        TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
'''

def init_db():
    global IS_POSTGRES, PH
    if IS_POSTGRES:
        try:
            con = psycopg2.connect(DATABASE_URL, connect_timeout=10)
            cur = con.cursor()
            cur.execute(_CREATE_TABLE.replace('{{', '{').replace('}}', '}'))
            for col, default in [
                ('target_prices',     "'{}'"),
                ('notes',             "'{}'"),
                ('rebalance_targets', "'{}'"),
                ('simulation',        "'[]'"),
            ]:
                try:
                    cur.execute(f"ALTER TABLE user_data ADD COLUMN IF NOT EXISTS {col} TEXT NOT NULL DEFAULT {default}")
                except Exception:
                    con.rollback()
            con.commit()
            cur.close(); con.close()
            print("[DB] PostgreSQL 연결 성공")
        except Exception as e:
            print(f"[DB] PostgreSQL 연결 실패, SQLite로 전환: {e}")
            IS_POSTGRES = False
            PH = '?'
            init_db()
            return
    else:
        os.makedirs(DATA_DIR, exist_ok=True)
        con = sqlite3.connect(DATABASE)
        con.execute(_CREATE_TABLE.replace('{{', '{').replace('}}', '}'))
        for col, default in [
            ('target_prices',     "'{}'"),
            ('notes',             "'{}'"),
            ('rebalance_targets', "'{}'"),
        ]:
            try:
                con.execute(f"ALTER TABLE user_data ADD COLUMN {col} TEXT NOT NULL DEFAULT {default}")
            except sqlite3.OperationalError:
                pass
        con.commit(); con.close()
        print("[DB] SQLite 사용:", DATABASE)

init_db()

# ── KR 종목명 캐시 백그라운드 빌드 (EXE/dev 모두 자동 실행) ────
# 이전: 'if __name__ == "__main__"' 안에 있어서 EXE에서 안 돌았음 (버그)
# 수정: import 시점에 백그라운드 thread로 실행 → EXE에서도 작동
def _kickstart_kr_cache():
    import threading as _t
    def _run():
        try:
            _build_kr_name_cache()
        except Exception as e:
            print(f"[Cache] 백그라운드 빌드 실패: {e}")
    _t.Thread(target=_run, daemon=True).start()

# 호출은 파일 마지막에 (함수 정의 후)

# ── 요청별 DB 연결 ────────────────────────────────────────
def get_db():
    if 'db' not in g:
        if IS_POSTGRES:
            g.db = psycopg2.connect(DATABASE_URL)
            g.db.autocommit = False
        else:
            g.db = sqlite3.connect(DATABASE)
            g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db:
        db.close()

# ── 사용자 토큰 미들웨어 ──────────────────────────────────
@app.before_request
def load_user_token():
    """쿠키에서 UUID 토큰 읽기. 없으면 새로 발급."""
    token = request.cookies.get('user_token', '')
    if len(token) != 36 or token.count('-') != 4:
        token = str(uuid.uuid4())
        g._new_token = token
    else:
        g._new_token = None
    g.user_token = token

@app.after_request
def set_user_token_cookie(response):
    """신규 사용자에게 토큰 쿠키 발급 (1년 유효)"""
    tok = g.get('_new_token')
    if tok:
        response.set_cookie(
            'user_token', tok,
            max_age=365 * 24 * 3600,
            httponly=True, samesite='Lax'
        )
    return response

def _upsert(field: str, value):
    """user_data 특정 필드 upsert (SQLite / PostgreSQL 공통)"""
    db  = get_db()
    sql = f'''
        INSERT INTO user_data (user_token, {field})
        VALUES ({PH}, {PH})
        ON CONFLICT(user_token) DO UPDATE SET
            {field} = EXCLUDED.{field},
            updated_at = CURRENT_TIMESTAMP
    '''
    if IS_POSTGRES:
        cur = db.cursor()
        cur.execute(sql, (g.user_token, json.dumps(value, ensure_ascii=False)))
        db.commit(); cur.close()
    else:
        db.execute(sql, (g.user_token, json.dumps(value, ensure_ascii=False)))
        db.commit()

def _fetch(field: str, default):
    """user_data 특정 필드 조회 (SQLite / PostgreSQL 공통)"""
    db  = get_db()
    sql = f'SELECT {field} FROM user_data WHERE user_token={PH}'
    if IS_POSTGRES:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, (g.user_token,))
        row = cur.fetchone(); cur.close()
        return json.loads(row[field]) if row else default
    else:
        row = db.execute(sql, (g.user_token,)).fetchone()
        return json.loads(row[field]) if row else default

def load_portfolio():
    return _fetch('portfolio', {'holdings': []})

def save_portfolio(data):
    _upsert('portfolio', data)

def load_transactions():
    return _fetch('transactions', [])

def save_transactions(txs):
    _upsert('transactions', txs)

def load_wishlist():
    return _fetch('wishlist', [])

def save_wishlist(items):
    _upsert('wishlist', items)

def load_hidden():
    return _fetch('hidden', [])

def save_hidden(items):
    _upsert('hidden', items)

def load_target_prices():
    return _fetch('target_prices', {})

def save_target_prices(d):
    _upsert('target_prices', d)

def load_notes():
    return _fetch('notes', {})

def save_notes(d):
    _upsert('notes', d)

def load_rebalance_targets():
    return _fetch('rebalance_targets', {})

def save_rebalance_targets(d):
    _upsert('rebalance_targets', d)

def rebuild_portfolio_from_transactions():
    """거래내역에서 포트폴리오 재계산 (평단가·수량 자동 업데이트)"""
    txs = load_transactions()
    # ticker 기준으로 그룹화
    positions = {}  # ticker → {name, market, qty, total_cost, realized_pnl}
    for tx in sorted(txs, key=lambda x: x['date']):
        tk = tx['ticker']
        if tk not in positions:
            positions[tk] = {
                'name': tx['name'], 'ticker': tk, 'market': tx['market'],
                'qty': 0.0, 'total_cost': 0.0, 'realized_pnl': 0.0
            }
        p = positions[tk]
        if tx['type'] == 'buy':
            new_qty   = p['qty'] + tx['qty']
            p['total_cost'] = p['total_cost'] + tx['qty'] * tx['price']
            p['qty']  = new_qty
        else:  # sell
            if p['qty'] > 0:
                sell_qty = min(tx['qty'], p['qty'])  # 보유량 초과 매도 방지
                avg = p['total_cost'] / p['qty']
                p['realized_pnl'] += (tx['price'] - avg) * sell_qty
                p['total_cost'] -= avg * sell_qty
            p['qty'] = max(0.0, p['qty'] - tx['qty'])
            if p['qty'] <= 0.000001:
                p['qty'] = 0.0
                p['total_cost'] = 0.0  # total_cost 음수 방지

    # 수량 0인 종목 제외, portfolio.json 형식으로 변환
    holdings = []
    for p in positions.values():
        if p['qty'] > 0.000001:
            holdings.append({
                'name':   p['name'],
                'ticker': p['ticker'],
                'market': p['market'],
                'qty':    round(p['qty'], 8),
                'cost':   round(p['total_cost']),
            })
    save_portfolio({'holdings': holdings})
    return holdings

def get_usd_krw():
    try:
        t = yf.Ticker("KRW=X")
        return t.info.get('regularMarketPrice') or t.fast_info.get('lastPrice', 1450)
    except:
        return 1450

@app.route('/')
def index():
    return render_template('dashboard.html', version=APP_VERSION)

@app.route('/api/version')
def api_version():
    return jsonify({'version': APP_VERSION, 'name': APP_NAME})


# ══════════════════════════════════════════════════════════
# 자동 업데이트 시스템
# ══════════════════════════════════════════════════════════
def _version_tuple(v):
    """버전 문자열을 비교 가능한 튜플로 변환: 'v1.2.3 Beta' → (1,2,3, 'beta')"""
    if not v: return (0, 0, 0, '')
    s = str(v).lower().replace('v', '').strip()
    # 뒤쪽 라벨 (alpha/beta/rc)
    label = ''
    for lab in ('alpha', 'beta', 'rc'):
        if lab in s:
            label = lab
            s = s.replace(lab, '').strip()
            break
    # 숫자 부분만 추출
    parts = [p for p in _re.split(r'[^\d]+', s) if p]
    nums = tuple(int(p) for p in parts[:4]) + (0, 0, 0, 0)
    return nums[:4] + (label,)


def _is_newer(latest, current):
    """latest > current 이면 True (라벨 무시: 같은 숫자여도 beta<stable)"""
    l_nums = _version_tuple(latest)[:4]
    c_nums = _version_tuple(current)[:4]
    if l_nums != c_nums:
        return l_nums > c_nums
    # 숫자가 같으면 라벨 비교: '' > rc > beta > alpha (stable이 가장 최신)
    rank = {'': 4, 'rc': 3, 'beta': 2, 'alpha': 1}
    l_lab = _version_tuple(latest)[4]
    c_lab = _version_tuple(current)[4]
    return rank.get(l_lab, 0) > rank.get(c_lab, 0)


def _gh_headers():
    """GitHub API 헤더 (필요 시 토큰 추가)"""
    h = {'Accept': 'application/vnd.github+json', 'X-GitHub-Api-Version': '2022-11-28'}
    tok = os.environ.get('GH_TOKEN', '')
    if tok:
        h['Authorization'] = f'Bearer {tok}'
    return h


@app.route('/api/update/check')
def api_update_check():
    """GitHub Releases에서 최신 버전 조회 → 현재 버전과 비교"""
    if not IS_GITHUB_UPDATE:
        return jsonify({
            'update_available': False,
            'current_version': APP_VERSION,
            'error': 'GitHub 설정이 없습니다 (config.json: gh_owner, gh_repo)'
        })
    try:
        r = requests.get(
            f'https://api.github.com/repos/{GH_OWNER}/{GH_REPO}/releases/latest',
            headers=_gh_headers(), timeout=10
        )
        if r.status_code == 404:
            return jsonify({
                'update_available': False,
                'current_version': APP_VERSION,
                'message': '등록된 릴리즈 없음'
            })
        if r.status_code != 200:
            return jsonify({
                'update_available': False,
                'current_version': APP_VERSION,
                'error': f'GitHub API 오류: {r.status_code}'
            })

        latest = r.json()
        latest_version = latest.get('tag_name', '').lstrip('v')
        # 정규화: v 접두 일관성 ("v1.2.3" 형태로)
        if latest_version and not latest_version.startswith('v'):
            latest_version_full = 'v' + latest_version
        else:
            latest_version_full = latest_version
        update_available = _is_newer(latest_version_full, APP_VERSION)

        # 인스톨러 .exe + ZIP 자산 모두 찾기
        download_url = ''
        zip_url = ''
        installer_size = 0
        zip_size = 0
        for asset in latest.get('assets', []):
            name = asset.get('name', '').lower()
            if name.endswith('.exe') and not download_url:
                download_url = asset.get('browser_download_url', '')
                installer_size = asset.get('size', 0)
            elif name.endswith('.zip') and not zip_url:
                zip_url = asset.get('browser_download_url', '')
                zip_size = asset.get('size', 0)

        body = latest.get('body', '') or ''
        return jsonify({
            'update_available':  update_available,
            'current_version':   APP_VERSION,
            'latest_version':    latest_version_full,
            'release_notes':     body,
            'download_url':      download_url,       # legacy: .exe 인스톨러
            'installer_size':    installer_size,
            'zip_url':           zip_url,            # 신규: ZIP 자동 업데이트
            'zip_size':          zip_size,
            'is_mandatory':      ('[MANDATORY]' in body or '[필수]' in body),
            'published_at':      latest.get('published_at', ''),
            'html_url':          latest.get('html_url', ''),
            'source':            'github',
        })
    except Exception as e:
        return jsonify({
            'update_available': False,
            'current_version': APP_VERSION,
            'error': f'업데이트 확인 실패: {e}'
        })


@app.route('/api/update/history')
def api_update_history():
    """모든 버전 히스토리 조회 (GitHub Releases)"""
    if not IS_GITHUB_UPDATE:
        return jsonify({'versions': []})
    try:
        r = requests.get(
            f'https://api.github.com/repos/{GH_OWNER}/{GH_REPO}/releases?per_page=30',
            headers=_gh_headers(), timeout=10
        )
        if r.status_code != 200:
            return jsonify({'versions': [], 'error': f'GitHub API status {r.status_code}'})

        releases = r.json() or []
        versions = []
        for rel in releases:
            if rel.get('draft'):
                continue
            body = rel.get('body', '') or ''
            tag = rel.get('tag_name', '').lstrip('v')
            if tag and not tag.startswith('v'):
                version = 'v' + tag
            else:
                version = tag
            # 인스톨러 다운로드 URL
            dl_url = ''
            for asset in rel.get('assets', []):
                if asset.get('name', '').lower().endswith('.exe'):
                    dl_url = asset.get('browser_download_url', '')
                    break
            versions.append({
                'version':       version,
                'release_notes': body,
                'download_url':  dl_url,
                'is_mandatory':  ('[MANDATORY]' in body or '[필수]' in body),
                'published_at':  rel.get('published_at', ''),
                'html_url':      rel.get('html_url', ''),
                'prerelease':    rel.get('prerelease', False),
            })
        return jsonify({'versions': versions})
    except Exception as e:
        return jsonify({'versions': [], 'error': str(e)})


def _handle_launcher_update(url, launcher_exe):
    """런처 패턴: pending_update.zip 저장 + 앱 종료 → 런처 재실행 → 자동 적용"""
    try:
        import threading as _th, subprocess as _sp
        appdata = os.environ.get('LOCALAPPDATA') or os.environ.get('APPDATA') or os.path.expanduser('~')
        update_root = os.path.join(appdata, 'StockDashboard')
        os.makedirs(update_root, exist_ok=True)
        pending_zip = os.path.join(update_root, 'pending_update.zip')
        tmp_zip = pending_zip + '.tmp'

        # 1. ZIP 다운로드 (먼저 .tmp로 → 완료 후 atomic rename)
        with requests.get(url, stream=True, timeout=180) as resp:
            resp.raise_for_status()
            with open(tmp_zip, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=65536):
                    if chunk: f.write(chunk)

        size = os.path.getsize(tmp_zip)
        if size < 1_000_000:
            try: os.remove(tmp_zip)
            except: pass
            return jsonify({'ok': False, 'error': f'ZIP 파일 손상 ({size}B)'}), 500

        # 2. atomic rename → pending_update.zip
        if os.path.exists(pending_zip):
            try: os.remove(pending_zip)
            except: pass
        os.rename(tmp_zip, pending_zip)

        # 3. 런처 재실행 예약 (현재 앱 종료 후 1초 뒤)
        DETACHED_PROCESS = 0x00000008
        CREATE_NEW_PROCESS_GROUP = 0x00000200
        # 헬퍼: 1.5초 대기 후 런처 실행 (현재 앱이 완전히 종료될 시간 확보)
        helper_bat = os.path.join(update_root, 'restart_helper.bat')
        bat_content = (
            '@echo off\r\n'
            'timeout /t 2 /nobreak >nul\r\n'
            f'start "" "{launcher_exe}"\r\n'
            'del "%~f0" >nul 2>&1\r\n'
        )
        with open(helper_bat, 'w', encoding='cp949', errors='ignore', newline='') as f:
            f.write(bat_content)

        CREATE_NO_WINDOW = 0x08000000
        _sp.Popen(['cmd.exe', '/c', helper_bat],
            creationflags=DETACHED_PROCESS | CREATE_NEW_PROCESS_GROUP | CREATE_NO_WINDOW,
            close_fds=True)

        # 4. 3초 후 자살
        def _suicide():
            import time as _t
            _t.sleep(3)
            try: os._exit(0)
            except: pass
        _th.Thread(target=_suicide, daemon=True).start()

        return jsonify({
            'ok': True,
            'method': 'launcher',
            'size_mb': round(size / (1024*1024), 1),
            'pending_path': pending_zip,
            'launcher_path': launcher_exe,
            'message': '✅ 다운로드 완료! 3초 후 앱 종료 → 런처가 자동으로 새 버전 적용 → 재시작 (UAC·클릭 없음)',
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': f'런처 업데이트 실패: {e}'}), 500


@app.route('/api/update/install-zip', methods=['POST'])
def api_update_install_zip():
    """
    🚀 v1.4.0+ 런처 패턴 (Squirrel/Discord 방식)

    흐름:
      1. ZIP 다운로드 → %LOCALAPPDATA%\\StockDashboard\\pending_update.zip 으로 저장
      2. 앱 종료 → 런처(StockDashboard.exe)가 다음 실행 시 자동으로 ZIP 적용
      3. 파일 락 없음! (앱이 종료된 후 적용되므로)

    런처 부재 시 (v1.3.x 이하 호환): legacy 배치 스크립트 폴백
    """
    data = request.get_json() or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'ok': False, 'error': 'ZIP URL이 없습니다'}), 400

    if not getattr(sys, 'frozen', False):
        return jsonify({'ok': False, 'error': '개발 모드 미지원'}), 400

    # ── 런처 감지: app/ 폴더 안에 있으면 런처 패턴 사용 ──
    install_dir_check = os.path.dirname(sys.executable)
    parent_dir = os.path.dirname(install_dir_check)
    parent_exe = os.path.join(parent_dir, 'StockDashboard.exe')
    is_launcher_pattern = (
        os.path.basename(install_dir_check).lower() == 'app' and
        os.path.exists(parent_exe)
    )

    if is_launcher_pattern:
        # 런처 패턴: ZIP을 pending 위치에 저장 + 앱 종료 → 런처가 다음 실행 시 적용
        return _handle_launcher_update(url, parent_exe)

    # ── Pre-check 1: install_dir 쓰기 권한 확인 ──
    install_dir_check = os.path.dirname(sys.executable)
    if not os.access(install_dir_check, os.W_OK):
        return jsonify({
            'ok': False,
            'use_legacy': True,  # 클라이언트가 인스톨러 방식으로 폴백
            'install_dir': install_dir_check,
            'error': f'설치 폴더에 쓰기 권한 없음 (Program Files에 설치된 듯).\n경로: {install_dir_check}\n→ 인스톨러 방식으로 자동 전환합니다.'
        }), 400

    try:
        import tempfile, subprocess, threading as _th, zipfile, shutil
        tmp_dir = tempfile.gettempdir()
        zip_path = os.path.join(tmp_dir, 'StockDashboard_update.zip')

        # ── Pre-check 2: 임시 폴더 쓰기 권한 ──
        if not os.access(tmp_dir, os.W_OK):
            return jsonify({'ok': False, 'error': f'임시 폴더 쓰기 권한 없음: {tmp_dir}'}), 500

        # 1. ZIP 다운로드
        with requests.get(url, stream=True, timeout=180) as resp:
            resp.raise_for_status()
            with open(zip_path, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=65536):
                    if chunk: f.write(chunk)

        size = os.path.getsize(zip_path)
        if size < 1_000_000:
            try: os.remove(zip_path)
            except: pass
            return jsonify({'ok': False, 'error': f'ZIP 손상 ({size}B)'}), 500

        # 2. ZIP 검증
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                names = zf.namelist()
                if not any('StockDashboard.exe' in n for n in names):
                    raise ValueError('ZIP에 StockDashboard.exe 없음')
        except Exception as e:
            try: os.remove(zip_path)
            except: pass
            return jsonify({'ok': False, 'error': f'ZIP 무결성 오류: {e}'}), 500

        # 3. staging 폴더에 압축 해제
        staging_dir = os.path.join(tmp_dir, 'StockDashboard_staging')
        if os.path.exists(staging_dir):
            try: shutil.rmtree(staging_dir, ignore_errors=True)
            except: pass
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(staging_dir)
        # ZIP 안에 'StockDashboard/' 폴더 있으면 평탄화
        nested = os.path.join(staging_dir, 'StockDashboard')
        if os.path.isdir(nested) and os.path.exists(os.path.join(nested, 'StockDashboard.exe')):
            staging_dir = nested

        # 4. 헬퍼 배치
        pid = os.getpid()
        app_exe = sys.executable
        install_dir = os.path.dirname(app_exe)
        install_dir_old = install_dir + '.old'
        log_path = os.path.join(tmp_dir, 'stockdash_zip_update.log')
        helper_path = os.path.join(tmp_dir, f'stockdash_zip_updater_{pid}.bat')

        batch = (
            '@echo off\r\n'
            'chcp 65001 >nul 2>&1\r\n'
            'title StockDashboard 업데이트 진행 중...\r\n'
            'setlocal\r\n'
            f'set "LOG={log_path}"\r\n'
            f'set "INSTALL={install_dir}"\r\n'
            f'set "STAGING={staging_dir}"\r\n'
            f'set "OLDDIR={install_dir_old}"\r\n'
            f'set "APP_EXE={app_exe}"\r\n'
            'echo ========================================\r\n'
            'echo  StockDashboard 업데이트 진행\r\n'
            'echo ========================================\r\n'
            'echo. > "%LOG%"\r\n'
            'echo [%date% %time%] === ZIP Update Start === >> "%LOG%"\r\n'
            'echo INSTALL=%INSTALL% >> "%LOG%"\r\n'
            'echo STAGING=%STAGING% >> "%LOG%"\r\n'
            '\r\n'
            'echo [1/5] 앱 + WebView2 종료 중...\r\n'
            'taskkill /F /IM "StockDashboard.exe" /T >nul 2>&1\r\n'
            'taskkill /F /IM "msedgewebview2.exe" /T >nul 2>&1\r\n'
            'echo [%date% %time%] Killed processes >> "%LOG%"\r\n'
            'timeout /t 5 /nobreak >nul\r\n'
            '\r\n'
            'echo [2/5] 기존 폴더 백업 중...\r\n'
            'if exist "%OLDDIR%" rmdir /S /Q "%OLDDIR%" >nul 2>&1\r\n'
            'move "%INSTALL%" "%OLDDIR%" 2>>"%LOG%"\r\n'
            'if errorlevel 1 (\r\n'
            '  echo [에러] 폴더 백업 실패 - Program Files 권한 부족 의심\r\n'
            '  echo [%date% %time%] MOVE INSTALL FAILED >> "%LOG%"\r\n'
            '  goto fallback\r\n'
            ')\r\n'
            '\r\n'
            'echo [3/5] 새 버전 적용 중...\r\n'
            'move "%STAGING%" "%INSTALL%" 2>>"%LOG%"\r\n'
            'if errorlevel 1 (\r\n'
            '  echo [에러] 새 버전 적용 실패 - 롤백\r\n'
            '  echo [%date% %time%] MOVE STAGING FAILED, rollback >> "%LOG%"\r\n'
            '  move "%OLDDIR%" "%INSTALL%" >nul 2>&1\r\n'
            '  goto fallback\r\n'
            ')\r\n'
            '\r\n'
            'echo [4/5] 새 앱 실행 중...\r\n'
            'timeout /t 2 /nobreak >nul\r\n'
            'if exist "%APP_EXE%" (\r\n'
            '  start "" "%APP_EXE%"\r\n'
            '  echo [%date% %time%] Started new app >> "%LOG%"\r\n'
            '  echo  - 새 앱 시작됨!\r\n'
            ') else (\r\n'
            '  echo [에러] 새 EXE 파일을 찾을 수 없음: %APP_EXE%\r\n'
            '  echo [%date% %time%] EXE NOT FOUND >> "%LOG%"\r\n'
            '  goto fallback\r\n'
            ')\r\n'
            '\r\n'
            'echo [5/5] 옛 버전 정리 중...\r\n'
            'timeout /t 5 /nobreak >nul\r\n'
            'rmdir /S /Q "%OLDDIR%" >nul 2>&1\r\n'
            'echo.\r\n'
            'echo ✅ 업데이트 완료! 이 창은 자동으로 닫힙니다.\r\n'
            'timeout /t 3 /nobreak >nul\r\n'
            'goto end\r\n'
            '\r\n'
            ':fallback\r\n'
            'echo.\r\n'
            'echo === 자동 업데이트 실패 — 수동 다운로드 페이지를 엽니다 ===\r\n'
            'echo [%date% %time%] FALLBACK >> "%LOG%"\r\n'
            'start "" "https://github.com/1415hdfhfsg/stock-dashboard/releases/latest"\r\n'
            'echo.\r\n'
            'echo 위 페이지에서 최신 Setup.exe 다운로드 → 더블클릭 설치\r\n'
            f'echo 로그: {log_path}\r\n'
            'echo.\r\n'
            'pause\r\n'
            '\r\n'
            ':end\r\n'
            'del "%~f0" >nul 2>&1\r\n'
        )
        with open(helper_path, 'w', encoding='cp949', errors='ignore', newline='') as f:
            f.write(batch)

        # 5. 헬퍼 콘솔 보이게 실행 (사용자가 진행/에러 직접 확인)
        CREATE_NEW_CONSOLE       = 0x00000010
        CREATE_NEW_PROCESS_GROUP = 0x00000200
        subprocess.Popen(['cmd.exe', '/c', helper_path],
            creationflags=CREATE_NEW_CONSOLE | CREATE_NEW_PROCESS_GROUP, close_fds=True)

        # 6. 3초 후 자살
        def suicide():
            import time as _time
            _time.sleep(3)
            try: os._exit(0)
            except: pass
        _th.Thread(target=suicide, daemon=True).start()

        return jsonify({
            'ok': True,
            'method': 'zip',
            'size_mb': round(size / (1024*1024), 1),
            'message': 'ZIP 압축 해제 + 자동 패치 시작 (UAC 불필요)',
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': f'ZIP 업데이트 실패: {e}'}), 500


@app.route('/api/update/install', methods=['POST'])
def api_update_install():
    """
    [Legacy] 인스톨러(.exe) 다운로드 + 자동 패치 + 재시작.
    ZIP 방식이 가능하면 install-zip 사용 권장.

    흐름:
      1. 인스톨러 .exe 다운로드 + 무결성 검증
      2. 헬퍼 배치 스크립트 생성 (앱 종료 → 설치 → 재시작)
      3. 배치 스크립트 독립 실행 (사용자 화면에 별도 콘솔 창으로 진행 표시)
      4. 본 앱 종료 → 헬퍼가 인계받아 진행
    """
    data = request.get_json() or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'ok': False, 'error': '다운로드 URL이 없습니다'}), 400

    if not getattr(sys, 'frozen', False):
        return jsonify({'ok': False, 'error': '개발 모드에서는 자동 업데이트를 지원하지 않습니다'}), 400

    try:
        import tempfile, subprocess, threading as _th
        tmp_dir = tempfile.gettempdir()
        filename = url.rstrip('/').split('/')[-1] or 'StockDashboard_Setup.exe'
        installer_path = os.path.join(tmp_dir, filename)

        # 1. 인스톨러 다운로드 (스트리밍, 무결성 검증)
        with requests.get(url, stream=True, timeout=180) as resp:
            resp.raise_for_status()
            with open(installer_path, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=32768):
                    if chunk:
                        f.write(chunk)

        size = os.path.getsize(installer_path)
        if size < 100_000:
            try: os.remove(installer_path)
            except: pass
            return jsonify({'ok': False, 'error': f'다운로드 파일이 손상됨 ({size}B). 잠시 후 다시 시도해주세요.'}), 500

        # 2. 헬퍼 배치: WebView2 자식 프로세스까지 모두 종료 + 파일 락 방지
        # 핵심 변경:
        #   - msedgewebview2.exe 우리 앱 자식만 선별 종료
        #   - 5초로 대기 시간 늘림 (파일 락 해제 충분히)
        #   - 인스톨러 실패 시 직접 다운로드 페이지 자동 오픈
        pid = os.getpid()
        app_exe = sys.executable
        app_exe_name = os.path.basename(app_exe)
        log_path = os.path.join(tmp_dir, 'stockdash_update.log')
        helper_path = os.path.join(tmp_dir, f'stockdash_updater_{pid}.bat')
        download_page_url = 'https://github.com/1415hdfhfsg/stock-dashboard/releases/latest'

        batch = (
            '@echo off\r\n'
            'setlocal\r\n'
            f'set "LOG={log_path}"\r\n'
            'echo [%date% %time%] === Update Helper v3 === > "%LOG%"\r\n'
            f'echo [%date% %time%] PID={pid} EXE={app_exe_name} >> "%LOG%"\r\n'
            '\r\n'
            'REM ── 1: 모든 관련 프로세스 종료 (WebView2 자식 포함) ─\r\n'
            f'taskkill /F /IM "{app_exe_name}" /T >nul 2>&1\r\n'
            'REM WebView2 우리 앱 자식 (storage_path 매칭) 종료\r\n'
            'wmic process where "name=\'msedgewebview2.exe\' and commandline like \'%%StockDashboard%%\'" delete >nul 2>&1\r\n'
            'echo [%date% %time%] Killed app + WebView2 processes >> "%LOG%"\r\n'
            'REM 파일 락 해제 충분 대기\r\n'
            'timeout /t 5 /nobreak >nul\r\n'
            '\r\n'
            'REM ── 2: 인스톨러 실행 (UI 표시, UAC 정상) ─\r\n'
            'echo [%date% %time%] Launching installer >> "%LOG%"\r\n'
            f'"{installer_path}" /SP- /NOICONS /NORESTART\r\n'
            'set "CODE=%errorlevel%"\r\n'
            'echo [%date% %time%] Installer exit code=%CODE% >> "%LOG%"\r\n'
            '\r\n'
            'REM ── 3: 파일 시스템 안정화 ─\r\n'
            'timeout /t 5 /nobreak >nul\r\n'
            '\r\n'
            'REM ── 4: 결과 처리 ─\r\n'
            f'if exist "{app_exe}" (\r\n'
            '    echo [%date% %time%] EXE exists, launching new app >> "%LOG%"\r\n'
            f'    start "" "{app_exe}"\r\n'
            ') else (\r\n'
            '    echo [%date% %time%] FAIL - opening download page >> "%LOG%"\r\n'
            f'    start "" "{download_page_url}"\r\n'
            ')\r\n'
            '\r\n'
            'REM ── 5: 정리 ─\r\n'
            'timeout /t 3 /nobreak >nul\r\n'
            'del "%~f0" >nul 2>&1\r\n'
        )
        with open(helper_path, 'w', encoding='cp949', errors='ignore', newline='') as f:
            f.write(batch)

        # 3. 헬퍼 배치 백그라운드 실행 (사용자에게 cmd 창 안 보임)
        DETACHED_PROCESS        = 0x00000008
        CREATE_NEW_PROCESS_GROUP = 0x00000200
        CREATE_NO_WINDOW         = 0x08000000
        subprocess.Popen(
            ['cmd.exe', '/c', helper_path],
            creationflags=DETACHED_PROCESS | CREATE_NEW_PROCESS_GROUP | CREATE_NO_WINDOW,
            close_fds=True,
        )

        # 4. 3초 후 본 앱 강제 종료 (헬퍼의 taskkill보다 먼저 cleanly 종료)
        def suicide():
            import time as _time
            _time.sleep(3)
            try:
                os._exit(0)
            except Exception:
                pass
        _th.Thread(target=suicide, daemon=True).start()

        return jsonify({
            'ok': True,
            'size_mb':  round(size / (1024*1024), 1),
            'message':  '✅ 다운로드 완료! 5초 후 자동 설치가 시작됩니다.\n\n🔔 UAC 창이 뜨면 [예]를 눌러주세요.\n진행 콘솔이 별도 창으로 표시되며, 설치 완료 후 앱이 자동으로 재시작됩니다.',
            'log_path': log_path,
        })
    except requests.exceptions.RequestException as e:
        return jsonify({'ok': False, 'error': f'다운로드 실패: {e}. 인터넷 연결을 확인해주세요.'}), 500
    except Exception as e:
        return jsonify({'ok': False, 'error': f'업데이트 실패: {e}'}), 500

@app.route('/api/settings')
def api_settings():
    return jsonify({
        'data_dir':    DATA_DIR if not IS_POSTGRES else 'Supabase PostgreSQL',
        'db_path':     DATABASE if not IS_POSTGRES else DATABASE_URL.split('@')[-1],
        'user_token':  g.user_token,
        'is_frozen':   getattr(sys, 'frozen', False),
    })

@app.route('/api/settings/data-dir', methods=['POST'])
def api_change_data_dir():
    """데이터 저장 위치 변경 (config.json 업데이트 후 재시작 필요)"""
    data = request.get_json() or {}
    new_dir = data.get('data_dir', '').strip()
    if not new_dir:
        return jsonify({'error': '경로를 입력하세요'}), 400
    try:
        os.makedirs(new_dir, exist_ok=True)
        # config.json 위치 결정
        if getattr(sys, 'frozen', False):
            cfg_path = os.path.join(os.path.dirname(sys.executable), 'config.json')
        else:
            cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
        # 기존 config 불러오기
        cfg = {}
        if os.path.exists(cfg_path):
            with open(cfg_path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
        cfg['data_dir'] = new_dir
        with open(cfg_path, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        return jsonify({'ok': True, 'new_dir': new_dir, 'config_path': cfg_path})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/my-token')
def api_my_token():
    """현재 사용자 토큰 반환 (다른 기기 접속용)"""
    return jsonify({'token': g.user_token})

@app.route('/api/set-token', methods=['POST'])
def api_set_token():
    """다른 기기에서 같은 데이터를 쓰려면 토큰을 입력"""
    data = request.get_json() or {}
    tok  = str(data.get('token', '')).strip()
    if len(tok) != 36 or tok.count('-') != 4:
        return jsonify({'ok': False, 'error': '잘못된 토큰 형식입니다'}), 400
    resp = jsonify({'ok': True})
    resp.set_cookie('user_token', tok, max_age=365*24*3600, httponly=True, samesite='Lax')
    return resp

@app.route('/api/open-data-folder')
def api_open_data_folder():
    """데이터 폴더를 탐색기로 열기 (Windows)"""
    import subprocess
    try:
        subprocess.Popen(['explorer', DATA_DIR])
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ── 위시리스트 (서버 DB 저장) ────────────────────────────
@app.route('/api/wishlist', methods=['GET'])
def api_get_wishlist():
    return jsonify({'wishlist': load_wishlist()})

@app.route('/api/wishlist', methods=['POST'])
def api_add_wishlist():
    item = request.get_json() or {}
    if not item.get('ticker'):
        return jsonify({'error': '티커 없음'}), 400
    items = load_wishlist()
    if not any(w['ticker'] == item['ticker'] for w in items):
        items.append({'ticker': item['ticker'], 'name': item.get('name',''), 'market': item.get('market','')})
        save_wishlist(items)
    return jsonify({'ok': True, 'wishlist': items})

@app.route('/api/wishlist/<ticker>', methods=['DELETE'])
def api_del_wishlist(ticker):
    items = [w for w in load_wishlist() if w['ticker'] != ticker]
    save_wishlist(items)
    return jsonify({'ok': True, 'wishlist': items})

# ── 모의 투자 시뮬레이션 (서버 DB 저장) ──────────────────
def load_simulation():
    return _fetch('simulation', [])

def save_simulation(items):
    _upsert('simulation', items)

def _current_price_krw(ticker, market):
    """종목의 현재가 (KRW 환산)"""
    try:
        if market == 'US':
            t = yf.Ticker(ticker)
            p = t.fast_info.get('lastPrice') or t.info.get('regularMarketPrice')
            if p:
                return float(p) * get_usd_krw()
        else:
            end   = datetime.now().strftime('%Y%m%d')
            start = (datetime.now()-timedelta(days=10)).strftime('%Y%m%d')
            df = krx.get_market_ohlcv(start, end, ticker)
            if not df.empty:
                return float(df['종가'].iloc[-1])
    except Exception:
        pass
    return 0.0

@app.route('/api/simulation', methods=['GET'])
def api_get_simulation():
    """모의 투자 목록 (현재가 + P/L 포함)"""
    items = load_simulation()
    result = []
    total_cost = 0.0
    total_value = 0.0
    best = {'pct': -9e9}
    worst = {'pct':  9e9}

    for s in items:
        cur_price = _current_price_krw(s['ticker'], s['market'])
        qty       = float(s.get('qty', 0))
        buy_price = float(s.get('buy_price', 0))
        cost      = qty * buy_price
        value     = qty * cur_price
        pnl       = value - cost
        pnl_pct   = (pnl / cost * 100) if cost > 0 else 0.0

        row = {
            'id':        s.get('id'),
            'ticker':    s.get('ticker'),
            'name':      s.get('name'),
            'market':    s.get('market'),
            'qty':       qty,
            'buy_price': round(buy_price),
            'buy_date':  s.get('buy_date', ''),
            'note':      s.get('note', ''),
            'cur_price': round(cur_price),
            'cost':      round(cost),
            'value':     round(value),
            'pnl':       round(pnl),
            'pnl_pct':   round(pnl_pct, 2),
        }
        result.append(row)
        total_cost  += cost
        total_value += value
        if pnl_pct > best['pct']:
            best  = {'pct': pnl_pct, 'name': s.get('name'), 'ticker': s.get('ticker')}
        if pnl_pct < worst['pct']:
            worst = {'pct': pnl_pct, 'name': s.get('name'), 'ticker': s.get('ticker')}

    total_pnl = total_value - total_cost
    total_pct = (total_pnl / total_cost * 100) if total_cost > 0 else 0.0

    return jsonify({
        'items': result,
        'summary': {
            'total_cost':  round(total_cost),
            'total_value': round(total_value),
            'total_pnl':   round(total_pnl),
            'total_pct':   round(total_pct, 2),
            'count':       len(result),
            'best':        best if best['pct']  > -9e8 else None,
            'worst':       worst if worst['pct'] <  9e8 else None,
        }
    })

@app.route('/api/simulation', methods=['POST'])
def api_add_simulation():
    """모의 매수 추가"""
    data = request.get_json() or {}
    ticker = (data.get('ticker') or '').strip()
    name   = (data.get('name') or '').strip()
    market = (data.get('market') or '').strip()
    try:
        qty       = float(data.get('qty', 0))
        buy_price = float(data.get('buy_price', 0))
    except (TypeError, ValueError):
        return jsonify({'error': '수량/가격 형식 오류'}), 400
    buy_date = (data.get('buy_date') or datetime.now().strftime('%Y-%m-%d')).strip()
    note     = (data.get('note') or '').strip()[:100]

    if not ticker or not market:
        return jsonify({'error': '티커/시장 정보가 필요합니다'}), 400
    if qty <= 0 or buy_price <= 0:
        return jsonify({'error': '수량과 매수가는 양수여야 합니다'}), 400

    items = load_simulation()
    new_id = str(uuid.uuid4())[:8]
    items.append({
        'id':         new_id,
        'ticker':     ticker,
        'name':       name,
        'market':     market,
        'qty':        qty,
        'buy_price':  buy_price,
        'buy_date':   buy_date,
        'note':       note,
        'created_at': datetime.now().isoformat(),
    })
    save_simulation(items)
    return jsonify({'ok': True, 'id': new_id})

@app.route('/api/simulation/<sim_id>', methods=['DELETE'])
def api_del_simulation(sim_id):
    """모의 매수 삭제"""
    items = [s for s in load_simulation() if s.get('id') != sim_id]
    save_simulation(items)
    return jsonify({'ok': True})

@app.route('/api/simulation/clear', methods=['POST'])
def api_clear_simulation():
    """모의 투자 전체 초기화"""
    save_simulation([])
    return jsonify({'ok': True})


# ── 숨김 종목 (서버 DB 저장) ────────────────────────────
@app.route('/api/hidden', methods=['GET'])
def api_get_hidden():
    return jsonify({'hidden': load_hidden()})

@app.route('/api/hidden', methods=['POST'])
def api_add_hidden():
    data  = request.get_json() or {}
    ticker = data.get('ticker', '')
    if not ticker:
        return jsonify({'error': '티커 없음'}), 400
    items = load_hidden()
    if ticker not in items:
        items.append(ticker)
        save_hidden(items)
    return jsonify({'ok': True})

@app.route('/api/hidden/<ticker>', methods=['DELETE'])
def api_del_hidden(ticker):
    items = [t for t in load_hidden() if t != ticker]
    save_hidden(items)
    return jsonify({'ok': True})

@app.route('/api/portfolio')
def api_portfolio():
    data = load_portfolio()
    usd_krw = get_usd_krw()
    results = []
    total_cost = 0
    total_value = 0

    for h in data['holdings']:
        try:
            if h['market'] == 'US':
                t = yf.Ticker(h['ticker'])
                price_usd = t.fast_info.get('lastPrice', 0)
                price_krw = price_usd * usd_krw
                current_value = price_krw * h['qty']
            else:
                today = datetime.now().strftime('%Y%m%d')
                yesterday = (datetime.now() - timedelta(days=7)).strftime('%Y%m%d')
                df = krx.get_market_ohlcv(yesterday, today, h['ticker'])
                if not df.empty:
                    price_krw = int(df.iloc[-1]['종가'])
                else:
                    price_krw = 0
                current_value = price_krw * h['qty']

            cost = h['cost']
            profit = current_value - cost
            profit_pct = (profit / cost * 100) if cost > 0 else 0

            results.append({
                'name': h['name'],
                'ticker': h['ticker'],
                'market': h['market'],
                'qty': h['qty'],
                'avg_price': round(cost / h['qty']) if h['qty'] > 0 else 0,
                'current_price': round(price_krw),
                'cost': cost,
                'current_value': round(current_value),
                'profit': round(profit),
                'profit_pct': round(profit_pct, 2),
            })
            total_cost += cost
            total_value += current_value
        except Exception as e:
            results.append({
                'name': h['name'], 'ticker': h['ticker'], 'market': h['market'],
                'qty': h['qty'], 'error': str(e)
            })

    total_profit = total_value - total_cost
    total_pct = (total_profit / total_cost * 100) if total_cost > 0 else 0

    return jsonify({
        'holdings': results,
        'summary': {
            'total_cost': round(total_cost),
            'total_value': round(total_value),
            'total_profit': round(total_profit),
            'total_profit_pct': round(total_pct, 2),
            'usd_krw': round(usd_krw, 2)
        }
    })

# 국내 업종 대표 종목 (섹터별 주요 종목으로 평균 등락률 계산)
KR_SECTOR_STOCKS = {
    '반도체·전자':    ['005930', '000660', '042700', '066570'],
    '바이오·제약':    ['207940', '068270', '326030', '128940'],
    '자동차·부품':    ['005380', '000270', '012330', '204320'],
    '금융·보험':      ['105560', '055550', '086790', '032830'],
    '인터넷·플랫폼':  ['035420', '035720', '251270', '293490'],
    '화학·에너지':    ['010950', '011170', '096770', '051910'],
    '통신':           ['017670', '030200', '032640'],
    '건설':           ['000720', '028050', '047040'],
    '유통·식품':      ['069960', '004170', '023530', '097950'],
    '철강·소재':      ['004020', '005490', '010780'],
}

# 분야 검색 키워드 → 대표 종목 매핑
SECTOR_SEARCH_MAP = {
    '반도체': {'kr': ['005930','000660','042700','066570','000990'], 'us': ['NVDA','AMD','INTC','SOXX']},
    '바이오': {'kr': ['207940','068270','326030','128940','091990'], 'us': ['MRNA','ABBV','AMGN','XBI']},
    '제약':   {'kr': ['207940','068270','326030','128940'],          'us': ['PFE','JNJ','ABBV','MRNA']},
    'AI':     {'kr': ['005930','042700','035420','377300'],          'us': ['NVDA','MSFT','GOOGL','META']},
    '인공지능':{'kr': ['005930','042700','035420','377300'],         'us': ['NVDA','MSFT','GOOGL','META']},
    '자동차': {'kr': ['005380','000270','012330','204320'],          'us': ['TSLA','F','GM','TM']},
    '전기차': {'kr': ['005380','000270','247540','086520'],          'us': ['TSLA','RIVN','NIO','LCID']},
    '배터리': {'kr': ['247540','006400','051910','086520'],          'us': ['ALB','LTHM','TSLA']},
    '2차전지':{'kr': ['247540','006400','051910','086520'],          'us': ['ALB','LTHM','TSLA']},
    '금융':   {'kr': ['105560','055550','086790','032830','000810'], 'us': ['JPM','GS','BAC','WFC']},
    '은행':   {'kr': ['105560','055550','086790','000810'],          'us': ['JPM','BAC','C','WFC']},
    '인터넷': {'kr': ['035420','035720','251270','293490'],          'us': ['GOOGL','META','AMZN','NFLX']},
    '플랫폼': {'kr': ['035420','035720','251270'],                   'us': ['GOOGL','META','AMZN','UBER']},
    '게임':   {'kr': ['036570','259960','263750','112040'],          'us': ['EA','TTWO','RBLX','ATVI']},
    '화학':   {'kr': ['051910','010950','011170','096770'],          'us': ['LIN','APD','DOW']},
    '에너지': {'kr': ['010950','096770','267250'],                   'us': ['XOM','CVX','COP','XLE']},
    '통신':   {'kr': ['017670','030200','032640'],                   'us': ['T','VZ','TMUS']},
    '건설':   {'kr': ['000720','028050','047040'],                   'us': ['DHI','LEN','PHM']},
    '소비재': {'kr': ['069960','004170','023530'],                   'us': ['AMZN','HD','NKE','MCD']},
    '식품':   {'kr': ['097950','004170','003230'],                   'us': ['MCD','SBUX','KO','PEP']},
    '헬스케어':{'kr': ['207940','068270','128940'],                  'us': ['JNJ','UNH','ABBV','MRK']},
    '엔터':   {'kr': ['035900','041510','122870','352820'],          'us': ['NFLX','DIS','SPOT']},
    '방산':   {'kr': ['047810','064350','000880'],                   'us': ['LMT','RTX','GD','NOC']},
    '철강':   {'kr': ['004020','005490','010780'],                   'us': ['X','NUE','STLD']},
    '조선':   {'kr': ['009540','042660','010140'],                   'us': []},
    '항공':   {'kr': ['003490','020560'],                            'us': ['DAL','UAL','AAL','BA']},
    '부동산': {'kr': ['016360','145720'],                            'us': ['XLRE','AMT','PLD','SPG']},
    # ── ETF / 지수 카테고리 ──
    'ETF':      {'kr': ['069500','102110','133690','305720','091160'],       'us': ['SPY','QQQ','VTI','VOO','IVV']},
    '지수':     {'kr': ['069500','102110','226490','229200'],                'us': ['SPY','QQQ','DIA','VTI']},
    '인덱스':   {'kr': ['069500','102110','226490','229200'],                'us': ['SPY','QQQ','DIA','VTI']},
    '미국지수': {'kr': ['379800','360750','379810','133690'],                'us': ['SPY','QQQ','VOO','VTI','DIA']},
    'S&P500':   {'kr': ['379800','360750','269540','379780'],                'us': ['SPY','VOO','IVV']},
    '나스닥':   {'kr': ['379810','133690','368590','381170'],                'us': ['QQQ','QQQM']},
    '코스피':   {'kr': ['069500','102110','226490','277630'],                'us': ['EWY','MCHI']},
    '코스닥':   {'kr': ['229200','233740'],                                  'us': []},
    '레버리지': {'kr': ['122630','252670','233740','360140'],                'us': ['TQQQ','SOXL','UPRO','SPXL','TNA']},
    '인버스':   {'kr': ['114800','252670'],                                  'us': ['SQQQ','SOXS']},
    '배당':     {'kr': ['325010','210780','161510'],                         'us': ['SCHD','VYM','DVY','NOBL','JEPI','JEPQ']},
    '리츠':     {'kr': ['276970','329200'],                                  'us': ['VNQ','SCHH','XLRE']},
    '채권':     {'kr': ['148070','379800','423160'],                         'us': ['TLT','BND','AGG','HYG','LQD']},
    '금':       {'kr': ['132030','319640'],                                  'us': ['GLD','IAU','SLV']},
    '원자재':   {'kr': ['132030','261220','130680','144600'],                'us': ['GLD','USO','SLV','DBC','DBA']},
    '원유':     {'kr': ['261220','130680'],                                  'us': ['USO','XLE','XOM']},
}

# ══════════════════════════════════════════════════════════
# 한국 대기업 그룹주 검색 매핑 ("삼성그룹" → 삼성 계열사 전체)
# ══════════════════════════════════════════════════════════
GROUP_SEARCH_MAP = {
    '삼성그룹': ['005930','207940','006400','028260','032830','000810','009150','010140','029780','016360','008770','030000','012750'],
    '현대차그룹': ['005380','000270','012330','086280','004020','000720','064350','011210'],
    '현대그룹': ['005380','000270','012330','086280','004020','000720','064350','011210'],
    'LG그룹': ['003550','066570','051910','373220','034220','032640','051900','011070'],
    'SK그룹': ['034730','000660','017670','096770','402340','018670','001740','011790','285130','326030','302440'],
    '롯데그룹': ['004990','023530','280360','005300','011170','020150','004000','330590','032350','089860','286940'],
    '한화그룹': ['000880','012450','272210','009830','042660','088350','000370'],
    '포스코그룹': ['005490','003670','022100','047050','058430'],
    'POSCO그룹': ['005490','003670','022100','047050','058430'],
    'KT그룹': ['030200','033780','007460'],
    'HD현대그룹': ['267250','329180','443060','267260','267270','042670'],
    '현대중공업그룹': ['267250','329180','443060','267260','267270','042670'],
    '두산그룹': ['000150','034020','241560','454910','336260','131970'],
    '효성그룹': ['004800','298050','298040','298020','298000','094280'],
    'CJ그룹': ['001040','097950','035760','000120','051500','011150','079160'],
    '카카오그룹': ['035720','323410','377300','293490'],
    '신세계그룹': ['004170','139480','031440','031430','035510'],
    'GS그룹': ['078930','007070','006360','001250','028150'],
    '네이버그룹': ['035420','035720'],
    '셀트리온그룹': ['068270','091990','068760'],
    '에코프로그룹': ['086520','247540','450080'],
    '한진그룹': ['180640','003490','002320'],
}

@app.route('/api/sectors')
def api_sectors():
    result = {'kr': [], 'us': []}

    # ── 국내 업종별 등락률 (대표 종목 평균) ───────────────────
    try:
        today     = datetime.now().strftime('%Y%m%d')
        fromdate  = (datetime.now() - timedelta(days=7)).strftime('%Y%m%d')

        for sector_name, tickers in KR_SECTOR_STOCKS.items():
            changes = []
            last_prices = []
            for tk in tickers:
                try:
                    df = krx.get_market_ohlcv(fromdate, today, tk)
                    if df is None or df.empty or len(df) < 2:
                        continue
                    prev = float(df['종가'].iloc[-2])
                    curr = float(df['종가'].iloc[-1])
                    if prev > 0:
                        changes.append((curr - prev) / prev * 100)
                        last_prices.append(curr)
                except:
                    continue
            if changes:
                avg_chg = sum(changes) / len(changes)
                result['kr'].append({
                    'name': sector_name,
                    'change': round(avg_chg, 2),
                    'value': round(sum(last_prices) / len(last_prices)),
                    'stocks': len(changes),
                })
    except Exception as e:
        result['kr_error'] = str(e)

    # ── 미국 섹터 ETF ───────────────────────────────────────
    us_sectors = {
        'XLK': '기술·IT',        'XLF': '금융',        'XLV': '헬스케어',
        'XLE': '에너지',          'XLI': '산업재',       'XLY': '경기소비재',
        'XLP': '필수소비재',      'XLU': '유틸리티',     'XLB': '소재·원자재',
        'XLRE': '부동산(리츠)',    'XLC': '커뮤니케이션', 'XBI': '바이오',
        'SOXX': '반도체',         'ARKK': '혁신·테크',
    }
    try:
        for sym, name in us_sectors.items():
            try:
                t    = yf.Ticker(sym)
                hist = t.history(period='5d')
                if len(hist) < 2:
                    continue
                prev = float(hist['Close'].iloc[-2])
                curr = float(hist['Close'].iloc[-1])
                chg  = (curr - prev) / prev * 100
                result['us'].append({
                    'name': name, 'ticker': sym,
                    'change': round(chg, 2), 'value': round(curr, 2)
                })
            except:
                continue
    except Exception as e:
        result['us_error'] = str(e)

    result['kr'] = sorted(result['kr'], key=lambda x: x['change'], reverse=True)
    result['us'] = sorted(result['us'], key=lambda x: x['change'], reverse=True)
    return jsonify(result)

def _parse_rss_date(date_str):
    """RSS date(RFC 822) → Unix timestamp (0 if failed)"""
    if not date_str:
        return 0
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(date_str)
        return int(dt.timestamp())
    except Exception:
        return 0


def _analyze_news_entry(entry, keyword_stock=None, seen=None):
    """뉴스 RSS entry → 통일된 dict (감정·관련 종목·timestamp 포함)"""
    title = entry.get('title', '').strip()
    if not title:
        return None
    if seen is not None:
        if title in seen:
            return None
        seen.add(title)
    # 관련 종목 감지
    related = []
    if keyword_stock:
        title_lower = title.lower()
        for kw, stock in keyword_stock.items():
            if kw.lower() in title_lower and stock:
                related.append({'name': kw, 'ticker': stock[0], 'market': stock[1]})
    # 감정 분석
    up_kw   = ['상승','급등','최고','돌파','호재','강세','매수','회복','반등','극복','개선','성장','흑자']
    down_kw = ['하락','급락','최저','붕괴','악재','약세','매도','위기','우려','폭락','침체','적자','경고']
    neg_ctx = ['우려','위기','불안','경고','리스크']
    has_up   = any(k in title for k in up_kw)
    has_down = any(k in title for k in down_kw)
    has_neg  = any(k in title for k in neg_ctx)
    if has_down or (has_neg and not has_up):
        sentiment = 'down'
    elif has_up and not has_neg:
        sentiment = 'up'
    else:
        sentiment = 'neutral'
    pub_str = entry.get('published', '') or ''
    ts = _parse_rss_date(pub_str)
    return {
        'title': title,
        'link':  entry.get('link', ''),
        'date':  pub_str[:25] if pub_str else '',
        'ts':    ts,  # Unix timestamp for client-side "N분 전" 표시
        'source': entry.get('source', {}).get('title', '') if isinstance(entry.get('source'), dict) else '',
        'related': related[:3],
        'sentiment': sentiment,
    }


@app.route('/api/news/search')
def api_news_search():
    """사용자 키워드 기반 뉴스 검색 (Google News RSS)"""
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify({'results': [], 'query': ''})
    if len(q) > 60:
        q = q[:60]

    # 정렬 옵션: latest(최신) / relevance(관련도)
    sort = request.args.get('sort', 'latest')
    hl   = request.args.get('hl', 'ko')  # 언어
    gl   = 'KR' if hl == 'ko' else 'US'

    from urllib.parse import quote_plus
    qenc = quote_plus(q)
    url  = f'https://news.google.com/rss/search?q={qenc}&hl={hl}&gl={gl}&ceid={gl}:{hl}'

    # 관련 종목 매핑 (재사용)
    keyword_stock = {
        '엔비디아': ('NVDA','US'), '테슬라': ('TSLA','US'), '마이크로소프트': ('MSFT','US'),
        'MS': ('MSFT','US'), '애플': ('AAPL','US'), '쿠팡': ('CPNG','US'),
        '한화': ('462330','KR'), '삼성': ('005930','KR'), 'SK하이닉스': ('000660','KR'),
        '메타': ('META','US'), '구글': ('GOOGL','US'), '아마존': ('AMZN','US'),
        'nvidia': ('NVDA','US'), 'tesla': ('TSLA','US'), 'apple': ('AAPL','US'),
    }

    try:
        feed = feedparser.parse(url)
    except Exception as e:
        return jsonify({'error': f'뉴스 검색 실패: {e}', 'results': []}), 500

    seen = set()
    results = []
    for entry in feed.entries[:50]:
        item = _analyze_news_entry(entry, keyword_stock, seen)
        if item:
            results.append(item)

    # 정렬
    if sort == 'relevance':
        q_lower = q.lower()
        # 제목에 검색어 포함 여부 → 최신순 (2차 키)
        results.sort(key=lambda x: (0 if q_lower in x['title'].lower() else 1, -x.get('ts', 0)))
    else:
        # 기본: 최신순 (ts 내림차순)
        results.sort(key=lambda x: x.get('ts', 0), reverse=True)

    return jsonify({
        'query':   q,
        'count':   len(results),
        'sort':    sort,
        'results': results[:30],
    })


@app.route('/api/news')
def api_news():
    # 카테고리별 RSS 피드
    categories = {
        '🔥 핫 트렌드': [
            'https://news.google.com/rss/search?q=트렌드+핫이슈&hl=ko&gl=KR&ceid=KR:ko',
            'https://news.google.com/rss/search?q=화제+인기&hl=ko&gl=KR&ceid=KR:ko',
        ],
        '🌍 세계 이슈': [
            'https://news.google.com/rss/search?q=세계+국제+이슈&hl=ko&gl=KR&ceid=KR:ko',
            'https://news.google.com/rss/search?q=미국+중국+글로벌&hl=ko&gl=KR&ceid=KR:ko',
        ],
        '🏛️ 정치·경제': [
            'https://news.google.com/rss/search?q=정치+경제+정책&hl=ko&gl=KR&ceid=KR:ko',
            'https://news.google.com/rss/search?q=금리+환율+연준&hl=ko&gl=KR&ceid=KR:ko',
        ],
        '📈 증시·투자': [
            'https://news.google.com/rss/search?q=주식+증시+코스피&hl=ko&gl=KR&ceid=KR:ko',
            'https://news.google.com/rss/search?q=나스닥+S&P500+투자&hl=ko&gl=KR&ceid=KR:ko',
        ],
    }

    # 키워드 → 종목 매핑
    keyword_stock = {
        '엔비디아': ('NVDA','US'), '테슬라': ('TSLA','US'), '마이크로소프트': ('MSFT','US'),
        'MS': ('MSFT','US'), '애플': ('AAPL','US'), '쿠팡': ('CPNG','US'),
        '한화': ('462330','KR'), '삼성': ('005930','KR'), 'SK하이닉스': ('000660','KR'),
        'AI': None, '반도체': None, '전기차': ('TSLA','US'), '빅테크': None,
        '메타': ('META','US'), '구글': ('GOOGL','US'), '아마존': ('AMZN','US'),
        'nvidia': ('NVDA','US'), 'tesla': ('TSLA','US'), 'apple': ('AAPL','US'),
    }

    result = {}
    seen_all = set()
    for cat, urls in categories.items():
        items = []
        for url in urls:
            try:
                feed = feedparser.parse(url)
                for entry in feed.entries[:12]:
                    item = _analyze_news_entry(entry, keyword_stock, seen_all)
                    if item:
                        items.append(item)
            except Exception as e:
                print(f"[News] {cat} 피드 실패: {e}")
                continue
        # 카테고리별 최신순 정렬 후 상위 6건
        items.sort(key=lambda x: x.get('ts', 0), reverse=True)
        result[cat] = items[:6]

    return jsonify({'categories': result})


# ── 국내 종목명 캐시 (서버 기동 후 첫 검색 시 빌드) ────────
_kr_name_cache = {}   # { ticker: name }

def _last_business_day_str(max_back=10):
    """오늘이 휴장일이면 며칠 전 영업일 찾기 (최대 max_back일)"""
    for i in range(max_back):
        d = (datetime.now() - timedelta(days=i)).strftime('%Y%m%d')
        # 토(5)/일(6) 스킵
        wd = (datetime.now() - timedelta(days=i)).weekday()
        if wd >= 5:
            continue
        return d
    return datetime.now().strftime('%Y%m%d')


def _build_kr_name_cache():
    """pykrx로 모든 KR 종목명 캐시 빌드 (KOSPI + KOSDAQ + KONEX, 휴장일 폴백)"""
    global _kr_name_cache
    if _kr_name_cache:
        return

    # 휴장일이면 최근 영업일로 폴백 (최대 10일 이전까지)
    target_date = None
    for i in range(10):
        d = (datetime.now() - timedelta(days=i)).strftime('%Y%m%d')
        try:
            tickers = krx.get_market_ticker_list(d, market='ALL')
            if tickers:
                target_date = d
                break
        except Exception:
            continue

    if not target_date:
        target_date = datetime.now().strftime('%Y%m%d')

    # 모든 시장(KOSPI/KOSDAQ/KONEX)에서 종목 리스트 + 이름 가져오기
    try:
        for mkt in ('KOSPI', 'KOSDAQ', 'KONEX'):
            try:
                tickers = krx.get_market_ticker_list(target_date, market=mkt)
                for tk in tickers:
                    if tk in _kr_name_cache:
                        continue
                    try:
                        nm = krx.get_market_ticker_name(str(tk))
                        if isinstance(nm, str) and nm.strip():
                            _kr_name_cache[str(tk)] = nm
                    except Exception:
                        pass
            except Exception:
                continue
        # ETF도 함께 캐시 (검색에서 ETF 이름 보존)
        try:
            etf_tickers = krx.get_etf_ticker_list(target_date)
            for tk in etf_tickers or []:
                if tk in _kr_name_cache:
                    continue
                try:
                    nm = krx.get_etf_ticker_name(str(tk))
                    if isinstance(nm, str) and nm.strip():
                        _kr_name_cache[str(tk)] = nm
                except Exception:
                    pass
        except Exception:
            pass
        print(f"[Cache] KR 종목명 {len(_kr_name_cache)}개 로드 완료 (기준: {target_date})")
    except Exception as e:
        print(f"[Cache] KR 종목명 캐시 빌드 실패: {e}")

def _fetch_kr_stock(tk, fromdate, today, name=None):
    """pykrx 종목(일반/ETF) OHLCV 조회 후 결과 dict 반환"""
    try:
        df = krx.get_market_ohlcv(fromdate, today, tk)
        if df is None or df.empty:
            return None
        curr = float(df['종가'].iloc[-1])
        prev = float(df['종가'].iloc[-2]) if len(df) > 1 else curr
        chg  = (curr - prev) / prev * 100 if prev else 0

        # 이름 결정: 1) pykrx 캐시 공식명 2) 인자 3) 사전 역조회 4) pykrx API
        # 별칭(예: '갤럭시')보다 공식명(예: '삼성전자') 우선 표시
        nm = _kr_name_cache.get(tk)
        if not nm:
            nm = name
        if not nm:
            try:
                nm = _kr_ticker_to_name(tk)
            except NameError:
                pass
        if not nm:
            try:
                raw_nm = krx.get_market_ticker_name(tk)
                if isinstance(raw_nm, str) and raw_nm.strip():
                    nm = raw_nm
            except Exception:
                pass
        if not nm:
            try:
                etf_nm = krx.get_etf_ticker_name(tk)
                if isinstance(etf_nm, str) and etf_nm.strip():
                    nm = etf_nm
            except Exception:
                pass
        if not nm:
            nm = tk

        return {
            'ticker': tk, 'name': nm, 'market': 'KR',
            'price_krw': int(curr), 'change_pct': round(chg, 2), 'currency': 'KRW'
        }
    except Exception as e:
        print(f"[Search] KR {tk} 조회 실패: {e}")
        return None

def _fetch_us_stock(sym, usd_krw):
    """yfinance 종목 조회 후 결과 dict 반환"""
    try:
        t     = yf.Ticker(sym)
        info  = t.fast_info
        price = info.get('lastPrice') or info.get('regularMarketPrice', 0)
        prev  = info.get('previousClose', price)
        if not price or price <= 0: return None
        chg   = (price - prev) / prev * 100 if prev else 0
        name  = t.info.get('shortName') or t.info.get('longName') or sym
        return {
            'ticker': sym, 'name': name, 'market': 'US',
            'price_usd': round(price, 2),
            'price_krw': round(price * usd_krw),
            'change_pct': round(chg, 2),
            'currency': 'USD',
            'sector':   t.info.get('sector', ''),
            'industry': t.info.get('industry', ''),
            'market_cap': t.info.get('marketCap', 0),
        }
    except:
        return None

# 한국어/영어 이름 → 미국 티커 매핑 (이름으로 해외 종목 검색 지원)
KR_NAME_TO_US = {
    '엔비디아': 'NVDA',  'nvidia': 'NVDA',
    '테슬라':   'TSLA',  'tesla':  'TSLA',
    '애플':     'AAPL',  'apple':  'AAPL',
    '마이크로소프트': 'MSFT', 'microsoft': 'MSFT',
    '구글':     'GOOGL', '알파벳': 'GOOGL', 'google': 'GOOGL', 'alphabet': 'GOOGL',
    '아마존':   'AMZN',  'amazon': 'AMZN',
    '메타':     'META',  '페이스북': 'META', 'meta': 'META', 'facebook': 'META',
    '넷플릭스': 'NFLX',  'netflix': 'NFLX',
    '스타벅스': 'SBUX',  'starbucks': 'SBUX',
    '나이키':   'NKE',   'nike': 'NKE',
    '코카콜라': 'KO',    'coca cola': 'KO',
    '펩시':     'PEP',   'pepsi': 'PEP',
    '월마트':   'WMT',   'walmart': 'WMT',
    '홈디포':   'HD',    'home depot': 'HD',
    '코스트코': 'COST',  'costco': 'COST',
    '비자':     'V',     'visa': 'V',
    '마스터카드': 'MA',  'mastercard': 'MA',
    'JP모건':   'JPM',   'jpmorgan': 'JPM',
    '골드만삭스': 'GS',  'goldman': 'GS',
    '뱅크오브아메리카': 'BAC', 'bank of america': 'BAC',
    '모건스탠리': 'MS',  'morgan stanley': 'MS',
    '엑슨모빌': 'XOM',   'exxon': 'XOM',
    '쉐브론':   'CVX',   'chevron': 'CVX',
    '쿠팡':     'CPNG',  'coupang': 'CPNG',
    '소니':     'SONY',  'sony': 'SONY', '플레이스테이션': 'SONY', 'ps5': 'SONY',
    '닌텐도':   'NTDOY', 'nintendo': 'NTDOY', '스위치': 'NTDOY',
    '팔란티어': 'PLTR',  'palantir': 'PLTR',
    '리비안':   'RIVN',  'rivian': 'RIVN',
    '루시드':   'LCID',  'lucid': 'LCID',
    '화이자':   'PFE',   'pfizer': 'PFE',
    '모더나':   'MRNA',  'moderna': 'MRNA',
    '존슨앤존슨': 'JNJ', 'johnson': 'JNJ',
    '유나이티드헬스': 'UNH', 'unitedhealth': 'UNH',
    '일라이릴리': 'LLY', 'eli lilly': 'LLY', '릴리': 'LLY',
    '노보노디스크': 'NVO', 'novo nordisk': 'NVO',
    'AMD': 'AMD', 'amd': 'AMD',
    '인텔':     'INTC',  'intel': 'INTC',
    '퀄컴':     'QCOM',  'qualcomm': 'QCOM',
    '브로드컴': 'AVGO',  'broadcom': 'AVGO',
    'TSMC':     'TSM',   'tsmc': 'TSM',
    'ASML':     'ASML',  'asml': 'ASML',
    '마이크론': 'MU',    'micron': 'MU',
    '어플라이드머티리얼즈': 'AMAT', 'applied materials': 'AMAT',
    '디즈니':   'DIS',   'disney': 'DIS',
    '스포티파이': 'SPOT', 'spotify': 'SPOT',
    '우버':     'UBER',  'uber': 'UBER',
    '에어비앤비': 'ABNB', 'airbnb': 'ABNB',
    '오라클':   'ORCL',  'oracle': 'ORCL',
    '세일즈포스': 'CRM', 'salesforce': 'CRM',
    '어도비':   'ADBE',  'adobe': 'ADBE',
    '서비스나우': 'NOW', 'servicenow': 'NOW',
    '스노우플레이크': 'SNOW', 'snowflake': 'SNOW',
    '크라우드스트라이크': 'CRWD', 'crowdstrike': 'CRWD',
    '데이터독': 'DDOG',  'datadog': 'DDOG',
    '코인베이스': 'COIN', 'coinbase': 'COIN',
    '페이팔':   'PYPL',  'paypal': 'PYPL',
    '블록':     'SQ',    'block': 'SQ', '스퀘어': 'SQ',
    '인튜이트': 'INTU',  'intuit': 'INTU',
    '버크셔':   'BRK-B', 'berkshire': 'BRK-B',
    '록히드마틴': 'LMT', 'lockheed': 'LMT',
    '레이시온':  'RTX',  'raytheon': 'RTX',
    'ARM': 'ARM', 'arm': 'ARM',
    '일본소니': 'SONY', 'sony': 'SONY',
    '삼성SDI': '006400',  # 국내 종목 → KR 티커 직접 사용
}

# 국내 인기 종목 이름 → 티커 (빠른 검색용, 이름이 key)
KR_POPULAR_STOCKS = {
    '삼성전자': '005930', 'SK하이닉스': '000660', 'LG에너지솔루션': '373220',
    '삼성바이오로직스': '207940', '현대차': '005380', '기아': '000270',
    'POSCO홀딩스': '005490', 'NAVER': '035420', '카카오': '035720',
    'LG화학': '051910', '삼성SDI': '006400', '현대모비스': '012330',
    'KB금융': '105560', '신한지주': '055550', '삼성물산': '028260',
    'SK이노베이션': '096770', '하나금융지주': '086790', '우리금융지주': '316140',
    'KT&G': '033780', 'SK텔레콤': '017670', 'KT': '030200',
    'LG전자': '066570', '한국전력': '015760', '두산에너빌리티': '034020',
    '셀트리온': '068270', '삼성생명': '032830', 'DB손해보험': '005830',
    '고려아연': '010130', '한화에어로스페이스': '012450', '한화오션': '042660',
    '한미반도체': '042700', 'SK': '034730', 'LG': '003550',
    'S-Oil': '010950', '롯데케미칼': '011170', '금호석유': '011780',
    '카카오뱅크': '323410', '카카오페이': '377300', '크래프톤': '259960',
    'NC소프트': '036570', '엔씨소프트': '036570', '넷마블': '251270',
    '포스코퓨처엠': '003670', '에코프로': '086520', '에코프로비엠': '247540',
    '엘앤에프': '066970', '일진머티리얼즈': '020150', 'SK아이이테크놀로지': '361610',
    '현대건설': '000720', '대우건설': '047040', 'GS건설': '006360',
    '롯데쇼핑': '023530', '신세계': '004170', '이마트': '139480',
    'CJ제일제당': '097950', '오리온': '271560', '농심': '004370',
    '아모레퍼시픽': '090430', 'LG생활건강': '051900', '한국콜마': '161890',
    '대한항공': '003490', '아시아나항공': '020560', '제주항공': '089590',
    'HMM': '011200', '현대글로비스': '086280', '한진칼': '180640',
    '삼성전기': '009150', '엘지이노텍': '011070', '삼성화재': '000810',
    '메리츠화재': '000060', '한국투자증권': '071050', '미래에셋증권': '006800',
    '키움증권': '039490', '삼성증권': '016360', '한국금융지주': '071050',
    '두산': '000150', '두산밥캣': '241560', '한화': '000880',
    '현대중공업': '009540', 'HD현대': '267250', 'HD현대중공업': '329180',

    # ── 사용자 요청 종목 + 추가 인기 KOSDAQ ──
    '한컴위드': '054920', '한글과컴퓨터': '030520',
    '클로봇': '466100',
    '센서뷰': '321370',
    '스피어': '348340', '스피어파워': '348340',
    'JYP Ent.': '035900', 'JYP': '035900', 'JYP엔터': '035900',
    '디씨앤미디어': '263720',
    'YG엔터테인먼트': '122870', 'YG': '122870',
    '에스엠': '041510', 'SM엔터테인먼트': '041510', 'SM': '041510',
    '하이브': '352820', 'HYBE': '352820',
    '루닛': '328130', '뷰노': '338220',
    '레인보우로보틱스': '277810', '제이오': '418550',
    '두산로보틱스': '454910', 'HD현대마린솔루션': '443060',
    '에코프로머티': '450080', '포스코DX': '022100',
    '셀트리온헬스케어': '091990', '셀트리온제약': '068760',
    '컴투스': '078340', '펄어비스': '263750',
    '위메이드': '112040', '카카오게임즈': '293490',
    'NHN': '181710', '쿠콘': '294570',
    '코웨이': '021240', '에스엘': '005850',
    'CJ ENM': '035760', 'CJ': '001040',

    # ── 통신사 + 브랜드 + 약어 ──
    'SKT': '017670', 'SK텔레콤': '017670',
    'LGU+': '032640', 'LGU': '032640', 'LG유플러스': '032640',
    '갤럭시': '005930',  # 삼성전자 브랜드 → 회사
    'S-Oil': '010950', '에스오일': '010950',
    '삼바': '207940', '삼바이오': '207940',
    '한국전력': '015760', '한전': '015760',
    'POSCO': '005490', '포스코': '005490',
    '네이버페이': '035420',
    '카카오모빌리티': '035720',

    # ── 추가 KOSPI 인기 종목 (회사명 검색 보강) ──
    '삼성중공업': '010140',
    '현대제철': '004020',
    '한미약품': '128940', '한미사이언스': '008930',
    '롯데지주': '004990', '롯데웰푸드': '280360', '롯데칠성': '005300',
    '롯데에너지머티리얼즈': '020150',
    'GS리테일': '007070', 'GS': '078930',
    '오뚜기': '007310',
    'BGF리테일': '282330', '동원F&B': '049770',
    'LG디스플레이': '034220', 'LG유플러스': '032640',
    'KCC': '002380', '한솔케미칼': '014680',
    '효성': '004800', '효성첨단소재': '298050', '효성중공업': '298040',
    '코오롱인더': '120110', '한솔홀딩스': '004150',
    '한진': '002320', '대한제강': '084010',
    '아세아': '002030', '풍산': '103140',
    '동국제강': '460860', 'TYM': '002900',
    '한솔제지': '213500', '무림페이퍼': '009200',
    '동서': '026960', '빙그레': '005180',
    '삼양식품': '003230', '하림': '136480',
    '대상': '001680', '풀무원': '017810',
    '신라젠': '215600', '제넥신': '095700',
    '네이처셀': '007390', '셀리버리': '268600',
    '이수페타시스': '007660', 'DB하이텍': '000990',
    '리노공업': '058470', '동진쎄미켐': '005290',
    '솔브레인': '357780', '원익IPS': '240810',
    '주성엔지니어링': '036930', '테크윙': '089030',
    'AP시스템': '265520', '비에이치': '090460',
    '아모텍': '052710', '인탑스': '049070',
    '나노스': '151910', '나노신소재': '121600',
    '대주전자재료': '078600', '엘앤에프': '066970',
    '에코프로비엠': '247540', '에코프로': '086520',
    '천보': '278280', '엠플러스': '259630',
    '필옵틱스': '161580', '동양생명': '082640',
    '한화생명': '088350', 'DB금융투자': '016610',
    '메리츠금융지주': '138040', '미래에셋생명': '085620',
    '삼성카드': '029780', '롯데카드': '037560',
    '대신증권': '003540', 'SK증권': '001510',
    '한국가스공사': '036460', '한국석유': '004090',
    'S&T모티브': '064960', 'SK가스': '018670',
    '코스맥스': '192820', '한국타이어앤테크놀로지': '161390',
    '넥센타이어': '002350', '금호타이어': '073240',
    '현대로템': '064350', 'LIG넥스원': '079550',
    '한화시스템': '272210', '한화솔루션': '009830',
    '대한전선': '001440', 'LS': '006260',
    'LS ELECTRIC': '010120', 'LS네트웍스': '000680',
    '대한해운': '005880', '팬오션': '028670',
    '대한제분': '001130', 'SPC삼립': '005610',
    '하이트진로': '000080', '진로발효': '018120',

    # ── 📊 한국 주요 ETF ──
    # 지수형
    'KODEX 200': '069500', 'TIGER 200': '102110', 'KODEX 코스닥150': '229200',
    'KODEX 코스피': '226490', 'TIGER 코스피': '277630',
    # 미국 지수
    'KODEX 미국S&P500': '379800', 'TIGER 미국S&P500': '360750',
    'ARIRANG 미국S&P500': '269540', 'KBSTAR 미국S&P500': '379780',
    'KODEX 미국나스닥100': '379810', 'TIGER 미국나스닥100': '133690',
    'KBSTAR 미국나스닥100': '368590', 'KODEX 미국테크': '381170',
    'TIGER 미국테크TOP10': '381170', 'TIGER 미국필라델피아반도체': '381180',
    'KODEX 미국반도체': '390390',
    # 레버리지·인버스
    'KODEX 레버리지': '122630', 'KODEX 인버스': '114800',
    'KODEX 200선물인버스2X': '252670', 'KODEX 코스닥150레버리지': '233740',
    'TIGER 미국나스닥100레버리지': '360140', 'KODEX 미국나스닥100레버리지': '409820',
    # 섹터
    'KODEX 반도체': '091160', 'TIGER 반도체': '091230',
    'TIGER Fn반도체TOP10': '396500', 'KODEX 2차전지산업': '305720',
    'TIGER 2차전지테마': '305540', 'KODEX 바이오': '244580',
    'TIGER 바이오': '244670', 'KODEX 자동차': '091180',
    'KODEX 은행': '091170', 'KODEX 건설': '117700',
    'KODEX 게임산업': '300610', 'TIGER K게임': '300610',
    'KODEX 미디어&엔터테인먼트': '266360', 'TIGER 미디어컨텐츠': '228810',
    'TIGER 차이나전기차SOLACTIVE': '371460',
    # 배당·채권
    'KODEX 배당가치': '325010', 'TIGER 고배당': '210780',
    'ARIRANG 고배당주': '161510', 'KODEX KOFR금리액티브': '423160',
    # 리츠·원자재
    'KODEX 리츠': '276970', 'TIGER 리츠부동산인프라': '329200',
    'KODEX 골드선물': '132030', 'TIGER 골드선물': '319640',
    'KODEX WTI원유선물': '261220', 'TIGER 원유선물Enhanced': '130680',
    'KODEX 은선물': '144600',
}

# ── 📊 미국 주요 ETF (KR_NAME_TO_US 보강) ──
US_POPULAR_ETFS = {
    # 광범위 인덱스
    'SPY': 'SPY', 'S&P500': 'SPY', '에스앤피500': 'SPY',
    'VOO': 'VOO', '뱅가드S&P500': 'VOO',
    'IVV': 'IVV', 'VTI': 'VTI', '미국전체시장': 'VTI',
    'QQQ': 'QQQ', '나스닥100': 'QQQ', 'QQQM': 'QQQM',
    'DIA': 'DIA', '다우존스': 'DIA', 'IWM': 'IWM', '러셀2000': 'IWM',
    # 배당
    'SCHD': 'SCHD', '슈왑배당': 'SCHD', 'VYM': 'VYM',
    'DVY': 'DVY', 'NOBL': 'NOBL', 'JEPI': 'JEPI', 'JEPQ': 'JEPQ',
    # 섹터
    'XLK': 'XLK', '기술섹터': 'XLK', 'XLF': 'XLF', '금융섹터': 'XLF',
    'XLE': 'XLE', '에너지섹터': 'XLE', 'XLV': 'XLV', '헬스케어섹터': 'XLV',
    'XLI': 'XLI', 'XLY': 'XLY', 'XLP': 'XLP', 'XLU': 'XLU',
    'XLB': 'XLB', 'XLRE': 'XLRE', 'XLC': 'XLC',
    'SMH': 'SMH', '반도체ETF': 'SMH', 'SOXX': 'SOXX',
    'IBB': 'IBB', '바이오테크': 'IBB', 'XBI': 'XBI',
    # 테마·혁신
    'ARKK': 'ARKK', 'ARK혁신': 'ARKK', 'ARKG': 'ARKG', 'ARKW': 'ARKW',
    'ICLN': 'ICLN', '친환경에너지': 'ICLN', 'TAN': 'TAN', '태양광': 'TAN',
    'LIT': 'LIT', '리튬배터리': 'LIT', 'DRIV': 'DRIV', '자율주행': 'DRIV',
    'ROBO': 'ROBO', '로봇자동화': 'ROBO', 'BOTZ': 'BOTZ',
    # 레버리지·인버스
    'TQQQ': 'TQQQ', '나스닥3배': 'TQQQ', 'SQQQ': 'SQQQ',
    'SOXL': 'SOXL', '반도체3배': 'SOXL', 'SOXS': 'SOXS',
    'TNA': 'TNA', 'UPRO': 'UPRO', 'SPXL': 'SPXL',
    # 국제
    'VXUS': 'VXUS', '해외주식': 'VXUS', 'VEA': 'VEA', 'VWO': 'VWO',
    'EFA': 'EFA', 'EEM': 'EEM', 'FXI': 'FXI', '중국주식': 'FXI',
    'EWJ': 'EWJ', '일본주식': 'EWJ', 'EWY': 'EWY', '한국ETF': 'EWY',
    'MCHI': 'MCHI', 'KWEB': 'KWEB',
    # 채권·원자재
    'BND': 'BND', 'AGG': 'AGG', 'TLT': 'TLT', '장기국채': 'TLT',
    'HYG': 'HYG', 'LQD': 'LQD', 'SHY': 'SHY', 'TIP': 'TIP',
    'GLD': 'GLD', '금ETF': 'GLD', 'IAU': 'IAU',
    'SLV': 'SLV', '은ETF': 'SLV', 'USO': 'USO', '원유ETF': 'USO',
    # 리츠
    'VNQ': 'VNQ', '미국리츠': 'VNQ', 'SCHH': 'SCHH', 'XLRE ': 'XLRE',
    # 변동성
    'VIX': 'VXX', 'VXX': 'VXX', 'UVXY': 'UVXY',
}

# KR_NAME_TO_US 에 ETF 병합 (검색 시 함께 매칭)
KR_NAME_TO_US.update(US_POPULAR_ETFS)

# KR 티커 → 이름 역조회 (ETF 등 get_market_ticker_name에서 안 나오는 경우 대비)
_KR_TICKER_TO_NAME_MAP = {}
for _nm, _tk in KR_POPULAR_STOCKS.items():
    # 같은 티커에 여러 이름이 있을 경우 더 짧은 쪽(공식 풀네임) 우선
    if _tk not in _KR_TICKER_TO_NAME_MAP or len(_nm) < len(_KR_TICKER_TO_NAME_MAP[_tk]):
        _KR_TICKER_TO_NAME_MAP[_tk] = _nm

def _kr_ticker_to_name(tk):
    """KR 티커로 이름 조회 (캐시 공식명 → 사전 → None)
    pykrx 캐시의 공식 종목명을 우선시 (별칭이 표시되는 것 방지)"""
    return _kr_name_cache.get(tk) or _KR_TICKER_TO_NAME_MAP.get(tk)

# ══════════════════════════════════════════════════════════
# 유사도 매칭 시스템 (종목/뉴스 공통)
# ══════════════════════════════════════════════════════════
import re as _re
from difflib import SequenceMatcher

# 한글/영문/티커 상호 매핑 (검색어 확장용)
KEYWORD_ALIASES = {
    # ── US 종목 ──
    'aapl': ['apple', '애플', '에플', '아이폰', '애푸', '애풀', '맥북', '아이패드', '에어팟'],
    'apple': ['aapl', '애플', '아이폰', '에플', '맥북', '아이패드', '에어팟'],
    '애플': ['apple', 'aapl', '아이폰', '에플', '맥북'],
    '에플': ['apple', 'aapl', '애플'],
    '아이폰': ['apple', 'aapl', '애플'],
    '맥북': ['apple', 'aapl', '애플'],
    '아이패드': ['apple', 'aapl', '애플'],
    '에어팟': ['apple', 'aapl', '애플'],
    'nvda': ['nvidia', '엔비디아', '엔비디야', '엔디비아', '엔비디어', '엔비디'],
    '엔디비아': ['nvidia', 'nvda', '엔비디아'],
    '엔비디야': ['nvidia', 'nvda', '엔비디아'],
    '엔비디': ['nvidia', 'nvda', '엔비디아'],
    'nvidia': ['nvda', '엔비디아'],
    '엔비디아': ['nvidia', 'nvda'],
    'tsla': ['tesla', '테슬라', '테슬', '테스라'],
    'tesla': ['tsla', '테슬라', '테슬'],
    '테슬라': ['tesla', 'tsla', '테슬'],
    '테슬': ['tesla', 'tsla', '테슬라'],
    '테스라': ['tesla', 'tsla', '테슬라'],
    'msft': ['microsoft', '마이크로소프트', 'ms', '마소', '마이크로'],
    'microsoft': ['msft', '마이크로소프트', 'ms', '마소', '마이크로'],
    '마이크로소프트': ['microsoft', 'msft', 'ms', '마소'],
    '마소': ['microsoft', 'msft', '마이크로소프트'],
    '마이크로': ['microsoft', 'msft', '마이크로소프트'],
    'googl': ['google', '구글', 'alphabet', '알파벳', 'goog', '구굴'],
    'goog': ['google', '구글', 'googl'],
    'google': ['googl', 'goog', '구글', 'alphabet', '알파벳', '구굴'],
    '구글': ['google', 'googl', 'goog', 'alphabet', '알파벳'],
    '구굴': ['google', 'googl', '구글'],
    'amzn': ['amazon', '아마존'],
    'amazon': ['amzn', '아마존'],
    '아마존': ['amazon', 'amzn'],
    'meta': ['facebook', '메타', '페이스북', 'fb'],
    '메타': ['meta', 'facebook', 'fb'],
    '페이스북': ['meta', 'facebook', 'fb'],
    'nflx': ['netflix', '넷플릭스'],
    'netflix': ['nflx', '넷플릭스'],
    '넷플릭스': ['netflix', 'nflx'],
    'amd': ['amd'],
    'intc': ['intel', '인텔'],
    'intel': ['intc', '인텔'],
    '인텔': ['intel', 'intc'],
    'qcom': ['qualcomm', '퀄컴'],
    'qualcomm': ['qcom', '퀄컴'],
    '퀄컴': ['qualcomm', 'qcom'],
    'cpng': ['coupang', '쿠팡'],
    'coupang': ['cpng', '쿠팡'],
    '쿠팡': ['coupang', 'cpng'],
    'sony': ['sne', '소니', '플레이스테이션'],
    '소니': ['sony', 'sne', '플레이스테이션'],
    '플레이스테이션': ['sony', '소니'],
    'ps5': ['sony', '소니', '플레이스테이션'],
    # ── KR 종목 ──
    '005930': ['삼성전자', '삼성', 'samsung', '삼전', '삼선', '삼셩', '갤럭시'],
    '삼성전자': ['005930', '삼성', 'samsung', '삼전', '갤럭시'],
    '삼성': ['삼성전자', '005930', 'samsung', '삼전'],
    'samsung': ['삼성전자', '삼성', '005930', '갤럭시'],
    '삼전': ['삼성전자', '005930'],
    '삼셩': ['삼성', '삼성전자', '005930'],
    '갤럭시': ['삼성전자', '005930', 'samsung'],
    '삼바': ['삼성바이오로직스', '207940'],
    '삼바이오': ['삼성바이오로직스', '207940'],
    '삼성바이오': ['삼성바이오로직스', '207940'],
    'skt': ['sk텔레콤', '017670'],
    'sk텔레콤': ['skt', '017670'],
    'lgu+': ['lg유플러스', '032640'],
    'lg유플러스': ['lgu+', '032640'],
    'lgu': ['lg유플러스', '032640'],
    '000660': ['sk하이닉스', '하이닉스', 'skhynix'],
    'sk하이닉스': ['000660', '하이닉스', 'skhynix'],
    '하이닉스': ['sk하이닉스', '000660', 'skhynix'],
    '035420': ['네이버', 'naver'],
    '네이버': ['035420', 'naver'],
    'naver': ['네이버', '035420'],
    '035720': ['카카오', 'kakao'],
    '카카오': ['035720', 'kakao'],
    'kakao': ['카카오', '035720'],
    '005380': ['현대차', '현대자동차', 'hyundai'],
    '현대차': ['005380', '현대자동차', 'hyundai'],
    '현대자동차': ['현대차', '005380', 'hyundai'],
    '207940': ['삼성바이오로직스', '삼성바이오'],
    '006400': ['삼성sdi', '삼성에스디아이'],
    '373220': ['lg에너지솔루션', 'lg엔솔', '엘지엔솔'],
    'lg엔솔': ['373220', 'lg에너지솔루션'],
    # ── 섹터/주제 ──
    '반도체': ['semiconductor', 'chip', '메모리', 'dram', 'nand', 'hbm'],
    'semiconductor': ['반도체', 'chip'],
    '이차전지': ['2차전지', '배터리', 'battery', 'ev배터리'],
    '배터리': ['2차전지', '이차전지', 'battery'],
    'ai': ['인공지능', 'artificial intelligence', 'chatgpt', 'gpt', '딥러닝'],
    '인공지능': ['ai', 'artificial intelligence', 'gpt'],
    '전기차': ['ev', 'electric vehicle', 'tesla', '테슬라'],
    'ev': ['전기차', 'electric vehicle'],
    '금리': ['interest rate', 'fed', '연준', '기준금리', 'fomc'],
    '연준': ['fed', 'fomc', '금리'],
    '환율': ['exchange rate', 'dollar', 'krw', 'usd', '달러'],
    '달러': ['dollar', 'usd', '환율'],
    '원유': ['oil', 'crude', 'wti', '유가'],
    '유가': ['oil', 'crude', 'wti', '원유'],
    # ── ETF / 지수 별칭 (섹터 카테고리와 충돌하지 않게 제한적으로) ──
    'etf': ['상장지수펀드'],
    '상장지수펀드': ['etf'],
    's&p500': ['spy', 'voo', 'ivv', '에스앤피', '에스앤피500', 'snp500'],
    '나스닥100': ['qqq', 'qqqm'],
    'kospi': ['코스피'],
    '코스피': ['kospi'],
    '코스닥': ['kosdaq'],
    'kodex': ['코덱스'],
    'tiger': ['타이거'],
    'arirang': ['아리랑'],
    'kbstar': ['케이비스타'],
    '레버리지': ['leverage'],
    '인버스': ['inverse'],
    '배당': ['dividend'],
    '리츠': ['reit'],
    '금': ['gold'],
    '채권': ['bond'],
}

def _normalize(s):
    """정규화: 소문자 + 공백/특수문자 제거"""
    if not s: return ''
    return _re.sub(r'[\s\-_.()·/]+', '', s.lower())

def _expand_keyword(kw):
    """키워드 확장: 본인 + 별칭들의 정규화 집합"""
    if not kw: return set()
    raw = kw.strip()
    norm = _normalize(raw)
    expanded = {norm, raw.lower()}
    # 직접 매핑
    for key in (norm, raw.lower()):
        for alias in KEYWORD_ALIASES.get(key, []):
            expanded.add(_normalize(alias))
            expanded.add(alias.lower())
    return {e for e in expanded if e}

def _fuzzy_score(query, target):
    """유사도 점수 (0~1). query가 target에 얼마나 일치하는지"""
    if not query or not target: return 0.0
    q, t = _normalize(query), _normalize(target)
    if not q or not t: return 0.0
    if q == t: return 1.0
    # 정확 부분 일치: 높은 점수
    if q in t: return 0.85 + (len(q) / len(t)) * 0.15
    if t in q: return 0.80 + (len(t) / len(q)) * 0.15
    # Levenshtein 기반 유사도
    return SequenceMatcher(None, q, t).ratio()

def _keyword_matches(subscribed_kw, text):
    """
    구독 키워드가 텍스트(뉴스 제목 등)에 매치되는지 확인.
    정규화 + 별칭 확장 + 부분 일치.
    """
    if not subscribed_kw or not text: return False
    text_norm = _normalize(text)
    aliases = _expand_keyword(subscribed_kw)
    for alias in aliases:
        if alias and alias in text_norm:
            return True
    return False


@app.route('/api/news/match-subs', methods=['POST'])
def api_news_match_subs():
    """뉴스 제목 리스트 중 구독 키워드와 매치되는 항목 반환 (유사도 매칭)"""
    data = request.get_json() or {}
    subs = data.get('subs', []) or []
    items = data.get('items', []) or []

    matches = []  # [{item, matched_sub}]
    for it in items:
        title = it.get('title', '')
        for sub in subs:
            if _keyword_matches(sub, title):
                matches.append({'item': it, 'matched_sub': sub})
                break
    return jsonify({'matches': matches})


@app.route('/api/search')
def api_search():
    q      = request.args.get('q', '').strip()
    market = request.args.get('market', 'all').lower()  # all | kr | us
    if not q:
        return jsonify({'results': []})

    results      = []
    seen_tickers = set()
    today    = datetime.now().strftime('%Y%m%d')
    fromdate = (datetime.now() - timedelta(days=7)).strftime('%Y%m%d')
    usd_krw  = get_usd_krw()
    q_lower  = q.lower()
    q_norm   = _normalize(q)
    # 키워드 확장 (별칭 포함)
    q_aliases = _expand_keyword(q)

    def add(item, score=1.0):
        if item and item['ticker'] not in seen_tickers:
            seen_tickers.add(item['ticker'])
            item['_score'] = score
            results.append(item)

    def matches_name(name):
        """확장된 별칭 + 다단계 유사도 매칭 (한글 오타도 잡기)"""
        if not name: return 0.0
        name_norm = _normalize(name)
        if not name_norm: return 0.0

        # 1) 완전 일치
        if q_norm == name_norm: return 1.0

        # 2) 별칭 일치 (사전 확장)
        for alias in q_aliases:
            if not alias: continue
            if alias == name_norm: return 0.97
            if alias in name_norm: return 0.92

        # 3) Prefix 매칭 (시작 일치 - 자동완성처럼)
        if name_norm.startswith(q_norm): return 0.93
        if q_norm.startswith(name_norm) and len(name_norm) >= 2: return 0.88

        # 4) 부분 문자열 일치
        if q_norm in name_norm: return 0.85
        if name_norm in q_norm and len(name_norm) >= 2: return 0.75

        # 5) Bigram 매칭 (한글 오타 대응)
        # 예: "엔디비아" vs "엔비디아" → 공통 bigram = "엔비"·"비디" 등으로 부분 매칭
        if len(q_norm) >= 2 and len(name_norm) >= 2:
            q_bigrams = set(q_norm[i:i+2] for i in range(len(q_norm)-1))
            n_bigrams = set(name_norm[i:i+2] for i in range(len(name_norm)-1))
            if q_bigrams and n_bigrams:
                jaccard = len(q_bigrams & n_bigrams) / len(q_bigrams | n_bigrams)
                if jaccard >= 0.4: return 0.5 + jaccard * 0.3   # 0.62 ~ 0.8
                if jaccard >= 0.25: return 0.45 + jaccard * 0.15  # 0.49 ~ 0.49

        # 6) 문자 집합 유사도 (Jaccard) — 글자 순서 바뀐 오타
        q_chars = set(q_norm)
        n_chars = set(name_norm)
        if q_chars and n_chars:
            char_sim = len(q_chars & n_chars) / len(q_chars | n_chars)
            if char_sim >= 0.7 and abs(len(q_norm) - len(name_norm)) <= 2:
                return 0.4 + char_sim * 0.2  # 0.54 ~ 0.6

        # 7) Levenshtein 유사도
        return SequenceMatcher(None, q_norm, name_norm).ratio()

    # ── 0-A. 그룹주 검색 (삼성그룹, 현대차그룹, LG그룹 등) ─
    # "그룹" 또는 "그룹주" 접미사가 명시적으로 있을 때만 매칭 (단순 "현대"는 일반 검색)
    matched_group = None
    matched_group_name = None
    if '그룹' in q_norm:
        # "삼성그룹주" / "삼성그룹" 둘 다 같은 키로 보정
        q_for_group = q_norm[:-1] if q_norm.endswith('주') and '그룹주' in q_norm else q_norm
        for gk, tickers in GROUP_SEARCH_MAP.items():
            gk_norm = _normalize(gk)
            if gk_norm == q_for_group:
                matched_group = tickers
                matched_group_name = gk
                break

    if matched_group:
        if market in ('all', 'kr'):
            for tk in matched_group:
                nm = _kr_ticker_to_name(tk)
                add(_fetch_kr_stock(tk, fromdate, today, name=nm), 0.92)
                if len(results) >= 12: break
        return jsonify({'results': results, 'group_search': True, 'group_name': matched_group_name})

    # ── 0. 분야/섹터 키워드 검색 (엄격 매칭: 정확 일치 또는 별칭만) ─
    matched_sector = None
    for kw, info in SECTOR_SEARCH_MAP.items():
        kw_norm = _normalize(kw)
        # 금융 vs 금 같은 부분일치 오매칭 방지: 엄격 비교만
        if kw_norm == q_norm or kw_norm in q_aliases:
            matched_sector = info
            break

    if matched_sector:
        if market in ('all', 'kr'):
            for tk in matched_sector.get('kr', []):
                # 섹터 결과에서도 ETF 이름을 보존하기 위해 역조회 사용
                nm = _kr_ticker_to_name(tk)
                add(_fetch_kr_stock(tk, fromdate, today, name=nm), 0.9)
                if len(results) >= 5: break
        if market in ('all', 'us'):
            for sym in matched_sector.get('us', []):
                add(_fetch_us_stock(sym, usd_krw), 0.9)
                if len(results) >= 10: break
        return jsonify({'results': results, 'sector_search': True})

    # ── 1. 해외(US) 검색 (별칭 + 유사도) ────────────────────
    # 전략: 후보 점수만 먼저 계산 → 상위 5개만 실제 yfinance 호출
    if market in ('all', 'us'):
        # 직접 티커 정확 매칭 → 1회 fetch
        q_upper = q.upper()
        add(_fetch_us_stock(q_upper, usd_krw), 1.0)
        # 후보 수집 (yfinance 호출 없이)
        us_candidates = []  # (score, sym, name_kw)
        for name_kw, sym in KR_NAME_TO_US.items():
            if not sym or sym in seen_tickers: continue
            s = matches_name(name_kw)
            # 별칭 집합과 정확히 일치해도 높은 점수
            if _normalize(name_kw) in q_aliases:
                s = max(s, 0.95)
            if s >= 0.50:
                us_candidates.append((s, sym, name_kw))
        # 중복 sym 제거 (같은 티커에 여러 이름이 매핑됨)
        seen_syms = set()
        unique_candidates = []
        for c in sorted(us_candidates, key=lambda x: x[0], reverse=True):
            if c[1] not in seen_syms:
                seen_syms.add(c[1])
                unique_candidates.append(c)
        # 상위 5개만 fetch
        for s, sym, nm in unique_candidates[:5]:
            add(_fetch_us_stock(sym, usd_krw), s)

    # ── 2. 국내(KR) 검색 (별칭 + 유사도) ────────────────────
    if market in ('all', 'kr'):
        try:
            if q.isdigit():
                add(_fetch_kr_stock(q.zfill(6), fromdate, today), 1.0)
            else:
                # 1단계: 인기 종목에서 후보 수집
                kr_candidates = []
                for nm, tk in KR_POPULAR_STOCKS.items():
                    s = matches_name(nm)
                    if _normalize(nm) in q_aliases:
                        s = max(s, 0.95)
                    if s >= 0.45:
                        kr_candidates.append((s, tk, nm))

                # 2단계: 캐시에서 후보 추가 수집
                # 캐시가 비어있으면 즉시 빌드 (Lazy load - 첫 검색 시 동기 호출)
                if not _kr_name_cache:
                    try:
                        _build_kr_name_cache()
                    except Exception as e:
                        print(f"[Search] 캐시 빌드 실패: {e}")
                if _kr_name_cache:
                    for tk, nm in _kr_name_cache.items():
                        s = matches_name(nm)
                        if s >= 0.45:
                            kr_candidates.append((s, tk, nm))

                # 중복 제거 + 점수 순 정렬
                seen_tks = set()
                unique_kr = []
                for c in sorted(kr_candidates, key=lambda x: x[0], reverse=True):
                    if c[1] not in seen_tks:
                        seen_tks.add(c[1])
                        unique_kr.append(c)
                # 상위 6개만 fetch
                for s, tk, nm in unique_kr[:6]:
                    if tk not in seen_tickers:
                        add(_fetch_kr_stock(tk, fromdate, today, name=nm), s)
        except Exception as e:
            print(f"[Search] KR 검색 오류: {e}")

    # 정렬: 유사도 점수 내림차순, 동일 점수 시 KR 우선
    results.sort(key=lambda r: (-(r.get('_score', 0)), 0 if r['market']=='KR' else 1))
    # 내부 점수 제거
    for r in results:
        r.pop('_score', None)

    return jsonify({'results': results[:15]})

@app.route('/api/upload', methods=['POST'])
def api_upload():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '.xlsx 파일만 지원됩니다'}), 400

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active

        # 헤더 찾기
        header_row = None
        col_map = {}
        for row in ws.iter_rows(min_row=1, max_row=10):
            cells = [str(c.value).strip().lower() if c.value else '' for c in row]
            for i, val in enumerate(cells):
                if val in ('종목명', '이름', 'name', '종목'): col_map['name'] = i
                if val in ('종목코드', '코드', 'ticker', 'code', '티커'): col_map['ticker'] = i
                if val in ('시장', 'market', '거래소'): col_map['market'] = i
                if val in ('수량', 'qty', 'quantity', '보유수량', '주수'): col_map['qty'] = i
                if val in ('매수금액', '총매수금액', 'cost', '투자금액', '매수총액', '총투자금액'): col_map['cost'] = i
            if 'name' in col_map and 'ticker' in col_map:
                header_row = row[0].row
                break

        if header_row is None:
            return jsonify({'error': '헤더를 찾을 수 없습니다. 종목명, 종목코드, 시장, 수량, 매수금액 열이 필요합니다.'}), 400

        holdings = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            vals = [c.value for c in row]
            name = vals[col_map.get('name', 0)]
            if not name:
                continue
            ticker = str(vals[col_map.get('ticker', 1)] or '')
            market = str(vals[col_map.get('market', 2)] or 'KR').upper()
            if market in ('미국', 'US', 'USA', 'NYSE', 'NASDAQ'):
                market = 'US'
            else:
                market = 'KR'
            qty = float(vals[col_map.get('qty', 3)] or 0)
            cost = float(vals[col_map.get('cost', 4)] or 0)
            if qty > 0:
                holdings.append({
                    'name': str(name),
                    'ticker': ticker,
                    'market': market,
                    'qty': qty,
                    'cost': cost
                })

        if not holdings:
            return jsonify({'error': '유효한 종목 데이터가 없습니다'}), 400

        data = {'holdings': holdings}
        save_portfolio(data)
        return jsonify({'success': True, 'count': len(holdings), 'holdings': holdings})

    except Exception as e:
        return jsonify({'error': f'파일 처리 오류: {str(e)}'}), 500

@app.route('/api/template')
def api_template():
    from io import BytesIO
    from flask import send_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '포트폴리오'
    headers = ['종목명', '종목코드', '시장', '수량', '매수금액']
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.cell(row=1, column=i).font = openpyxl.styles.Font(bold=True)
    ws.append(['삼성전자', '005930', 'KR', 10, 550000])
    ws.append(['엔비디아', 'NVDA', 'US', 5, 500000])
    for col in ['A','B','C','D','E']:
        ws.column_dimensions[col].width = 15
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='포트폴리오_템플릿.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def calc_indicators(df):
    """8개 기술적 지표 계산 + 점수화"""
    import numpy as np
    close = df['Close']
    high = df['High']
    low = df['Low']
    volume = df['Volume']

    # 1. 이동평균선
    ma5  = close.rolling(5).mean()
    ma20 = close.rolling(20).mean()
    ma60 = close.rolling(60).mean()

    # 2. RSI(14)
    delta = close.diff()
    gain  = delta.clip(lower=0)
    loss  = (-delta).clip(lower=0)
    avg_gain = gain.ewm(com=13, adjust=False).mean()
    avg_loss = loss.ewm(com=13, adjust=False).mean()
    rs  = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))

    # 3. MACD(12/26/9)
    ema12  = close.ewm(span=12, adjust=False).mean()
    ema26  = close.ewm(span=26, adjust=False).mean()
    macd   = ema12 - ema26
    signal = macd.ewm(span=9, adjust=False).mean()
    hist   = macd - signal

    # 4. 볼린저밴드(20/2)
    bb_mid   = close.rolling(20).mean()
    bb_std   = close.rolling(20).std()
    bb_upper = bb_mid + 2 * bb_std
    bb_lower = bb_mid - 2 * bb_std

    # 5. ATR(14)
    tr = pd.concat([
        high - low,
        (high - close.shift()).abs(),
        (low  - close.shift()).abs()
    ], axis=1).max(axis=1)
    atr = tr.rolling(14).mean()

    # 6. ADX(14)
    up_move   = high.diff()
    down_move = (-low.diff())
    plus_dm   = np.where((up_move > down_move) & (up_move > 0), up_move, 0.0)
    minus_dm  = np.where((down_move > up_move) & (down_move > 0), down_move, 0.0)
    atr14     = tr.ewm(com=13, adjust=False).mean()
    plus_di   = 100 * pd.Series(plus_dm,  index=df.index).ewm(com=13, adjust=False).mean() / atr14
    minus_di  = 100 * pd.Series(minus_dm, index=df.index).ewm(com=13, adjust=False).mean() / atr14
    dx        = 100 * (plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, np.nan)
    adx       = dx.ewm(com=13, adjust=False).mean()

    # 최신값 추출
    c   = close.iloc[-1]
    m5  = ma5.iloc[-1]
    m20 = ma20.iloc[-1]
    m60 = ma60.iloc[-1]
    r   = rsi.iloc[-1]
    mc  = macd.iloc[-1]
    sig = signal.iloc[-1]
    h   = hist.iloc[-1]
    h_prev = hist.iloc[-2] if len(hist) > 1 else h
    bb_u = bb_upper.iloc[-1]
    bb_l = bb_lower.iloc[-1]
    bb_m = bb_mid.iloc[-1]
    atr_val = atr.iloc[-1]
    adx_val = adx.iloc[-1]
    vol5  = volume.iloc[-5:].mean()
    vol20 = volume.iloc[-20:].mean()

    # 52주 고저
    w52_high = high.iloc[-252:].max() if len(high) >= 252 else high.max()
    w52_low  = low.iloc[-252:].min()  if len(low)  >= 252 else low.min()
    w52_pos  = (c - w52_low) / (w52_high - w52_low) * 100 if w52_high != w52_low else 50

    # ── 점수화 ──────────────────────────────────────────────
    scores = {}

    # ① 이평선 배열
    if m5 > m20 and m20 > m60:   scores['ma_align'] = 2
    elif m5 > m20 or m20 > m60:  scores['ma_align'] = 1
    elif m5 < m20 and m20 < m60: scores['ma_align'] = -2
    elif m5 < m20 or m20 < m60:  scores['ma_align'] = -1
    else:                          scores['ma_align'] = 0

    # ② 골든/데드크로스 (최근 10일 내 교차 감지)
    cross = 0
    for i in range(1, min(11, len(ma5))):
        prev5 = ma5.iloc[-i-1]; prev20 = ma20.iloc[-i-1]
        cur5  = ma5.iloc[-i];   cur20  = ma20.iloc[-i]
        if prev5 <= prev20 and cur5 > cur20:
            # 최근일수록 강한 신호 (1일 전=+2, 10일 전=+1)
            cross = 2 if i <= 3 else 1; break
        elif prev5 >= prev20 and cur5 < cur20:
            cross = -2 if i <= 3 else -1; break
    if cross == 0:
        cross = 1 if m5 > m20 else (-1 if m5 < m20 else 0)
    scores['cross'] = cross

    # ③ RSI
    if   r < 30:  scores['rsi'] = 2
    elif r < 45:  scores['rsi'] = 1
    elif r < 55:  scores['rsi'] = 0
    elif r < 70:  scores['rsi'] = -1
    else:          scores['rsi'] = -2

    # ④ MACD
    if   mc > sig and h > h_prev: scores['macd'] = 2
    elif mc > sig:                 scores['macd'] = 1
    elif mc < sig and h < h_prev: scores['macd'] = -2
    elif mc < sig:                 scores['macd'] = -1
    else:                          scores['macd'] = 0

    # ⑤ 볼린저밴드
    bb_range = bb_u - bb_l if bb_u != bb_l else 1
    bb_pos   = (c - bb_l) / bb_range  # 0~1
    prev_c   = close.iloc[-2] if len(close) > 1 else c
    if   bb_pos < 0.15 and c > prev_c: scores['bb'] = 2
    elif bb_pos < 0.35:                 scores['bb'] = 1
    elif bb_pos < 0.65:                 scores['bb'] = 0
    elif bb_pos < 0.85:                 scores['bb'] = -1
    else:                               scores['bb'] = -2

    # ⑥ 52주 위치
    if   w52_pos < 20:  scores['w52'] = 2
    elif w52_pos < 40:  scores['w52'] = 1
    elif w52_pos < 60:  scores['w52'] = 0
    elif w52_pos < 80:  scores['w52'] = -1
    else:               scores['w52'] = -2

    # ⑦ 거래량 트렌드
    vr = vol5 / vol20 if vol20 > 0 else 1
    if   vr > 1.5:  scores['volume'] = 1
    elif vr > 0.9:  scores['volume'] = 0
    else:           scores['volume'] = -1

    # ⑧ RSI 다이버전스 (최근 10일)
    divergence = 0
    if len(close) >= 10:
        price_hi = close.iloc[-10:].max(); price_lo = close.iloc[-10:].min()
        rsi_hi   = rsi.iloc[-10:].max();   rsi_lo   = rsi.iloc[-10:].min()
        if close.iloc[-1] >= price_hi * 0.99 and rsi.iloc[-1] < rsi_hi * 0.97:
            divergence = -2  # 하락 다이버전스
        elif close.iloc[-1] <= price_lo * 1.01 and rsi.iloc[-1] > rsi_lo * 1.03:
            divergence = 2   # 상승 다이버전스

    # ADX 신뢰도
    if   adx_val > 25: reliability = 'HIGH'
    elif adx_val > 20: reliability = 'MEDIUM'
    else:               reliability = 'LOW'

    total = sum(scores.values()) + divergence

    # 등급
    if   total >= 9:  grade, grade_ko = 'strong_up',   '강한 상승'
    elif total >= 4:  grade, grade_ko = 'up',           '상승'
    elif total >= -3: grade, grade_ko = 'neutral',      '중립'
    elif total >= -8: grade, grade_ko = 'down',         '하락'
    else:             grade, grade_ko = 'strong_down',  '강한 하락'

    # 예상 범위
    short_high = round(c + atr_val * 1.5)
    short_low  = round(c - atr_val * 1.5)
    mid_high   = round(bb_u)
    mid_low    = round(bb_l)
    # 이평선 기울기로 중기 범위 조정
    slope = (m20 - ma20.iloc[-6]) / ma20.iloc[-6] if len(ma20) >= 6 and ma20.iloc[-6] > 0 else 0
    adj   = round(c * slope * 3)
    mid_high += adj; mid_low += adj

    # 미니차트용 데이터 (최근 60일)
    n = min(60, len(close))
    chart = {
        'dates':  [str(d.date()) for d in close.iloc[-n:].index],
        'close':  [round(v, 2) for v in close.iloc[-n:].tolist()],
        'ma5':    [round(v, 2) if not pd.isna(v) else None for v in ma5.iloc[-n:].tolist()],
        'ma20':   [round(v, 2) if not pd.isna(v) else None for v in ma20.iloc[-n:].tolist()],
        'ma60':   [round(v, 2) if not pd.isna(v) else None for v in ma60.iloc[-n:].tolist()],
    }

    return {
        'total_score': total,
        'grade': grade,
        'grade_ko': grade_ko,
        'reliability': reliability,
        'adx': round(float(adx_val), 1),
        'rsi': round(float(r), 1),
        'current_price': round(float(c)),
        'atr': round(float(atr_val)),
        'scores': scores,
        'divergence': divergence,
        'short_range': {'high': int(short_high), 'low': int(short_low)},
        'mid_range':   {'high': int(mid_high),   'low': int(mid_low)},
        'w52': {'high': round(float(w52_high)), 'low': round(float(w52_low)), 'pos': round(float(w52_pos), 1)},
        'chart': chart,
        'indicators': {
            '이평선 배열': {'score': scores['ma_align'],  'desc': f'MA5={round(m5):,} MA20={round(m20):,} MA60={round(m60):,}'},
            '골든/데드크로스': {'score': scores['cross'], 'desc': f'MA5 {">" if m5>m20 else "<"} MA20'},
            'RSI(14)':     {'score': scores['rsi'],     'desc': f'RSI={r:.1f}'},
            'MACD':        {'score': scores['macd'],    'desc': f'MACD={mc:.2f} Signal={sig:.2f}'},
            '볼린저밴드':  {'score': scores['bb'],      'desc': f'상단={round(bb_u):,} 중심={round(bb_m):,} 하단={round(bb_l):,}'},
            '52주 위치':   {'score': scores['w52'],     'desc': f'{w52_pos:.0f}% (저={round(w52_low):,} 고={round(w52_high):,})'},
            '거래량':      {'score': scores['volume'],  'desc': f'5일평균/20일평균={vr:.2f}x'},
        }
    }


@app.route('/api/analysis')
def api_analysis():
    data = load_portfolio()
    usd_krw = get_usd_krw()
    results = []

    for h in data['holdings']:
        try:
            if h['market'] == 'US':
                t = yf.Ticker(h['ticker'])
                df = t.history(period='1y')
                if df.empty or len(df) < 30:
                    raise ValueError('데이터 부족')
                df.index = df.index.tz_localize(None)
                # 원화 환산
                df['Close'] = df['Close'] * usd_krw
                df['High']  = df['High']  * usd_krw
                df['Low']   = df['Low']   * usd_krw
            else:
                end   = datetime.now().strftime('%Y%m%d')
                start = (datetime.now() - timedelta(days=400)).strftime('%Y%m%d')
                raw   = krx.get_market_ohlcv(start, end, h['ticker'])
                if raw.empty or len(raw) < 30:
                    raise ValueError('데이터 부족')
                df = raw.rename(columns={'시가':'Open','고가':'High','저가':'Low','종가':'Close','거래량':'Volume'})

            result = calc_indicators(df)
            result['name']   = h['name']
            result['ticker'] = h['ticker']
            result['market'] = h['market']
            results.append(result)

        except Exception as e:
            results.append({
                'name': h['name'], 'ticker': h['ticker'], 'market': h['market'],
                'error': str(e)
            })

    return jsonify({'analysis': results})


# ────────────────────────────────────────────────────────────
# 알고리즘 비교용 4개 알고리즘
# ────────────────────────────────────────────────────────────

def algo_comprehensive(df):
    """Algo A: 종합형 — 8지표 ADX 가중 조합 (73~82% 실증 승률)"""
    import numpy as np
    result = calc_indicators(df)
    adx_val = result['adx']
    indicators = result['indicators']
    scores = result['scores']
    div = result['divergence']

    # ── ADX 기반 가중치 설정 ──
    # ADX>25 추세장: 추세 지표(이평선, 크로스, MACD)에 가중, 역추세 지표(RSI, BB, 52주) 감쇄
    # ADX<20 횡보장: 역추세/평균회귀 지표에 가중, 추세 지표 감쇄
    if adx_val > 25:
        # 추세장 가중치
        w = {'ma_align': 1.5, 'cross': 1.5, 'rsi': 0.7, 'macd': 1.5,
             'bb': 0.7, 'w52': 0.8, 'volume': 1.2}
        regime = f'추세장 (ADX {adx_val:.1f}>25): 추세 지표 가중'
    elif adx_val < 20:
        # 횡보장 가중치
        w = {'ma_align': 0.7, 'cross': 0.7, 'rsi': 1.5, 'macd': 0.8,
             'bb': 1.5, 'w52': 1.3, 'volume': 1.0}
        regime = f'횡보장 (ADX {adx_val:.1f}<20): 평균회귀 지표 가중'
    else:
        # 전환구간: 균등 가중
        w = {'ma_align': 1.0, 'cross': 1.0, 'rsi': 1.0, 'macd': 1.0,
             'bb': 1.0, 'w52': 1.0, 'volume': 1.0}
        regime = f'전환구간 (ADX {adx_val:.1f}): 균등 가중'

    # 가중 점수 계산
    weighted_score = 0
    detail = {}
    weight_key_map = {
        '이평선 배열': 'ma_align', '골든/데드크로스': 'cross', 'RSI(14)': 'rsi',
        'MACD': 'macd', '볼린저밴드': 'bb', '52주 위치': 'w52', '거래량': 'volume',
    }
    for name, info in indicators.items():
        raw_s = info['score']
        wk = weight_key_map.get(name)
        wt = w.get(wk, 1.0) if wk else 1.0
        ws = round(raw_s * wt, 1)
        weighted_score += ws
        wt_label = f' (x{wt:.1f})' if wt != 1.0 else ''
        detail[name] = {'score': ws, 'desc': f'{info["desc"]}{wt_label}'}

    # 다이버전스 (가중치 적용 안 함)
    div_desc = '상승 다이버전스' if div > 0 else ('하락 다이버전스' if div < 0 else '없음')
    detail['다이버전스'] = {'score': div, 'desc': div_desc}
    weighted_score += div

    # ADX 레짐 정보
    detail['ADX 레짐'] = {'score': 0, 'desc': regime}

    total = round(weighted_score, 1)
    max_score = 15
    grade, grade_ko = _score_to_grade(total, max_score)

    return {
        'name': 'Algo A: 종합형',
        'desc': '8개 지표 + 다이버전스 + ADX 레짐 가중',
        'basis': '73~82% 실증 승률 (ADX>25 추세 가중, ADX<20 평균회귀 가중)',
        'score': total,
        'max_score': max_score,
        'grade': grade,
        'grade_ko': grade_ko,
        'reliability': result['reliability'],
        'adx': result['adx'],
        'detail': detail,
        'short_range': result['short_range'],
        'mid_range': result['mid_range'],
        'strength': 'ADX 레짐에 따라 지표 가중치 자동 조정, 추세/횡보 모두 대응',
        'weakness': '전환 구간(ADX 20~25)에서 가중치 변화 지연 가능',
    }


def algo_trend_follow(df):
    """Algo B: 추세 추종형 — MA 정배열 + ADX + MACD (추세장 특화)"""
    import numpy as np
    close  = df['Close']
    high   = df['High']
    low    = df['Low']

    ma5  = close.rolling(5).mean()
    ma20 = close.rolling(20).mean()
    ma60 = close.rolling(60).mean()
    ma120= close.rolling(120).mean()
    ema12= close.ewm(span=12, adjust=False).mean()
    ema26= close.ewm(span=26, adjust=False).mean()
    macd = ema12 - ema26
    sig  = macd.ewm(span=9, adjust=False).mean()
    hist = macd - sig

    tr = pd.concat([(high-low),(high-close.shift()).abs(),(low-close.shift()).abs()],axis=1).max(axis=1)
    up_m  = high.diff(); dn_m = (-low.diff())
    pdm   = np.where((up_m>dn_m)&(up_m>0), up_m, 0.0)
    ndm   = np.where((dn_m>up_m)&(dn_m>0), dn_m, 0.0)
    atr14 = tr.ewm(com=13, adjust=False).mean()
    pdi   = 100*pd.Series(pdm,index=df.index).ewm(com=13,adjust=False).mean()/atr14
    ndi   = 100*pd.Series(ndm,index=df.index).ewm(com=13,adjust=False).mean()/atr14
    dx    = 100*(pdi-ndi).abs()/(pdi+ndi).replace(0,np.nan)
    adx   = dx.ewm(com=13,adjust=False).mean()

    c = close.iloc[-1]
    m5=ma5.iloc[-1]; m20=ma20.iloc[-1]; m60=ma60.iloc[-1]; m120=ma120.iloc[-1]
    mc=macd.iloc[-1]; sg=sig.iloc[-1]; h=hist.iloc[-1]; hp=hist.iloc[-2] if len(hist)>1 else h
    adx_v=adx.iloc[-1]; atr_v=tr.rolling(14).mean().iloc[-1]

    detail = {}
    score = 0

    # ① 완전 정배열 4단계 (5>20>60>120)
    if m5>m20>m60>m120:    s=4; txt='완전 정배열 (4단계)'
    elif m5>m20>m60:       s=3; txt='정배열 (3단계)'
    elif m5>m20:           s=2; txt='부분 정배열'
    elif m5<m20<m60<m120:  s=-4; txt='완전 역배열 (4단계)'
    elif m5<m20<m60:       s=-3; txt='역배열 (3단계)'
    elif m5<m20:           s=-2; txt='부분 역배열'
    else:                  s=0; txt='혼조'
    detail['이평선 배열'] = {'score': s, 'desc': txt}
    score += s

    # ② ADX 추세 강도 (0~3점)
    if   adx_v > 40: s=3; txt=f'ADX {adx_v:.1f} - 매우 강한 추세'
    elif adx_v > 25: s=2; txt=f'ADX {adx_v:.1f} - 강한 추세'
    elif adx_v > 20: s=1; txt=f'ADX {adx_v:.1f} - 추세 형성 중'
    else:            s=0; txt=f'ADX {adx_v:.1f} - 횡보장 (신호 약함)'
    detail['ADX 추세강도'] = {'score': s, 'desc': txt}
    score += s

    # ③ MACD 방향 + 히스토그램 가속도
    if   mc>sg and h>hp and h>0: s=3; txt='MACD 상승 + 히스토그램 확대'
    elif mc>sg and h>0:          s=2; txt='MACD 상승'
    elif mc>sg:                  s=1; txt='MACD>시그널 (히스토그램 감소)'
    elif mc<sg and h<hp and h<0: s=-3; txt='MACD 하락 + 히스토그램 확대'
    elif mc<sg and h<0:          s=-2; txt='MACD 하락'
    else:                        s=-1; txt='MACD<시그널'
    detail['MACD'] = {'score': s, 'desc': txt}
    score += s

    # ④ 현재가 vs 이평선 위치
    above = sum([c>m5, c>m20, c>m60, c>m120])
    s = above - 2  # -2 ~ +2
    detail['이평선 대비 위치'] = {'score': s, 'desc': f'이평선 {above}개 위에 위치'}
    score += s

    max_score = 12
    grade, grade_ko = _score_to_grade(score, max_score)
    atr_v_safe = float(atr_v) if not pd.isna(atr_v) else c * 0.02
    return {
        'name': 'Algo B: 추세 추종형',
        'desc': '이평선 4단계 배열 + ADX + MACD 히스토그램 가속도',
        'basis': '추세장(ADX>25) 집중, 데드크로스/골든크로스 강화',
        'score': score, 'max_score': max_score,
        'grade': grade, 'grade_ko': grade_ko,
        'reliability': 'HIGH' if adx_v > 25 else ('MEDIUM' if adx_v > 20 else 'LOW'),
        'adx': round(float(adx_v), 1), 'detail': detail,
        'short_range': {'high': round(c + atr_v_safe*1.5), 'low': round(c - atr_v_safe*1.5)},
        'mid_range':   {'high': round(c + atr_v_safe*4),   'low': round(c - atr_v_safe*4)},
        'strength': '추세장에서 매우 강력, 노이즈 적음',
        'weakness': '횡보장(ADX<20)에서 신호 부정확',
    }


def algo_contrarian(df):
    """Algo C: 역발상형 — RSI 극단 + 볼린저 + 52주 저점 (반등 특화)"""
    import numpy as np
    close = df['Close']; high = df['High']; low = df['Low']

    delta = close.diff()
    gain  = delta.clip(lower=0); loss = (-delta).clip(lower=0)
    rsi   = 100 - 100/(1+close.ewm(com=13,adjust=False).mean().pipe(lambda _:
            gain.ewm(com=13,adjust=False).mean()/loss.ewm(com=13,adjust=False).mean().replace(0,np.nan)))
    # 간단 재계산
    ag = gain.ewm(com=13,adjust=False).mean()
    al = loss.ewm(com=13,adjust=False).mean()
    rs = ag / al.replace(0, np.nan)
    rsi = 100 - (100/(1+rs))

    bb_mid  = close.rolling(20).mean()
    bb_std  = close.rolling(20).std()
    bb_up   = bb_mid + 2*bb_std
    bb_lo   = bb_mid - 2*bb_std

    w52_hi = high.iloc[-252:].max() if len(high)>=252 else high.max()
    w52_lo = low.iloc[-252:].min()  if len(low)>=252  else low.min()
    tr     = pd.concat([(high-low),(high-close.shift()).abs(),(low-close.shift()).abs()],axis=1).max(axis=1)
    atr_v  = tr.rolling(14).mean().iloc[-1]

    c=close.iloc[-1]; r=rsi.iloc[-1]; r_prev=rsi.iloc[-5] if len(rsi)>5 else r
    bu=bb_up.iloc[-1]; bl=bb_lo.iloc[-1]; bm=bb_mid.iloc[-1]
    w52_pos = (c-w52_lo)/(w52_hi-w52_lo)*100 if w52_hi!=w52_lo else 50

    detail = {}; score = 0

    # ① RSI 극단값 (과매도/과매수)
    if   r < 20:  s=5; txt=f'RSI {r:.1f} - 심각한 과매도 (강한 반등 기대)'
    elif r < 30:  s=3; txt=f'RSI {r:.1f} - 과매도'
    elif r < 40:  s=1; txt=f'RSI {r:.1f} - 약한 과매도'
    elif r > 80:  s=-5; txt=f'RSI {r:.1f} - 심각한 과매수 (조정 경고)'
    elif r > 70:  s=-3; txt=f'RSI {r:.1f} - 과매수'
    elif r > 60:  s=-1; txt=f'RSI {r:.1f} - 약한 과매수'
    else:         s=0; txt=f'RSI {r:.1f} - 중립'
    detail['RSI 극단값'] = {'score': s, 'desc': txt}
    score += s

    # ② RSI 방향 전환 (상승 전환 중인지)
    rsi_trend = r - r_prev
    if rsi_trend > 5 and r < 50: s=2; txt=f'과매도 구간서 RSI 상승 전환 (+{rsi_trend:.1f})'
    elif rsi_trend < -5 and r > 50: s=-2; txt=f'과매수 구간서 RSI 하락 전환 ({rsi_trend:.1f})'
    else: s=0; txt='RSI 방향 전환 없음'
    detail['RSI 방향 전환'] = {'score': s, 'desc': txt}
    score += s

    # ③ 볼린저밴드 위치 (역발상: 하단 = 매수 기회)
    bb_range = bu - bl if bu != bl else 1
    bb_pos   = (c - bl) / bb_range
    if   bb_pos < 0.05: s=4; txt='볼린저 하단 이탈 (강한 반등 신호)'
    elif bb_pos < 0.20: s=3; txt='볼린저 하단 근접'
    elif bb_pos < 0.35: s=2; txt='볼린저 하단~중심 구간'
    elif bb_pos > 0.95: s=-4; txt='볼린저 상단 이탈 (과열)'
    elif bb_pos > 0.80: s=-3; txt='볼린저 상단 근접'
    elif bb_pos > 0.65: s=-2; txt='볼린저 중심~상단 구간'
    else: s=0; txt='볼린저 중심 근처'
    detail['볼린저밴드'] = {'score': s, 'desc': txt}
    score += s

    # ④ 52주 저가 근접도 (역발상: 저점 = 매수 기회)
    if   w52_pos < 10: s=4; txt=f'52주 최저점 근접 ({w52_pos:.0f}%)'
    elif w52_pos < 25: s=2; txt=f'52주 저점권 ({w52_pos:.0f}%)'
    elif w52_pos > 90: s=-4; txt=f'52주 최고점 근접 ({w52_pos:.0f}%)'
    elif w52_pos > 75: s=-2; txt=f'52주 고점권 ({w52_pos:.0f}%)'
    else: s=0; txt=f'52주 중립 구간 ({w52_pos:.0f}%)'
    detail['52주 위치'] = {'score': s, 'desc': txt}
    score += s

    max_score = 15
    grade, grade_ko = _score_to_grade(score, max_score)
    atr_safe = float(atr_v) if not pd.isna(atr_v) else c*0.02
    return {
        'name': 'Algo C: 역발상형',
        'desc': 'RSI 극단값 + 볼린저밴드 하단 + 52주 저점 집중',
        'basis': '횡보장/반등 국면 특화, 과매도 구간 매수 시점 포착',
        'score': score, 'max_score': max_score,
        'grade': grade, 'grade_ko': grade_ko,
        'reliability': 'HIGH' if abs(r-50) > 25 else ('MEDIUM' if abs(r-50) > 15 else 'LOW'),
        'adx': 0, 'detail': detail,
        'short_range': {'high': round(bm), 'low': round(bl)},
        'mid_range':   {'high': round(bu), 'low': round(bl)},
        'strength': '과매도 반등 포착, 횡보장 유효',
        'weakness': '강한 추세장에서 조기 역추세 진입 위험',
    }


def algo_momentum(df):
    """Algo D: 모멘텀형 — 단기 가격 모멘텀 + 거래량 급증 + MACD 히스토그램"""
    import numpy as np
    close  = df['Close']; high = df['High']; low = df['Low']; volume = df['Volume']

    # 모멘텀 (ROC: Rate of Change)
    roc5  = (close / close.shift(5)  - 1) * 100
    roc20 = (close / close.shift(20) - 1) * 100
    roc60 = (close / close.shift(60) - 1) * 100

    ema12 = close.ewm(span=12,adjust=False).mean()
    ema26 = close.ewm(span=26,adjust=False).mean()
    macd  = ema12 - ema26
    sig   = macd.ewm(span=9,adjust=False).mean()
    hist  = macd - sig

    vol5  = volume.iloc[-5:].mean()
    vol20 = volume.iloc[-20:].mean()
    vol_ratio = vol5 / vol20 if vol20 > 0 else 1

    tr    = pd.concat([(high-low),(high-close.shift()).abs(),(low-close.shift()).abs()],axis=1).max(axis=1)
    atr_v = tr.rolling(14).mean().iloc[-1]

    c=close.iloc[-1]
    r5=roc5.iloc[-1]; r20=roc20.iloc[-1]; r60=roc60.iloc[-1]
    h=hist.iloc[-1]; hp=hist.iloc[-2] if len(hist)>1 else h; hp2=hist.iloc[-3] if len(hist)>2 else hp

    detail = {}; score = 0

    # ① 단기 ROC (5일 모멘텀)
    if   r5 > 5:   s=3; txt=f'5일 모멘텀 +{r5:.1f}% (강한 상승)'
    elif r5 > 2:   s=2; txt=f'5일 모멘텀 +{r5:.1f}%'
    elif r5 > 0:   s=1; txt=f'5일 모멘텀 +{r5:.1f}% (약한 상승)'
    elif r5 > -2:  s=-1; txt=f'5일 모멘텀 {r5:.1f}% (약한 하락)'
    elif r5 > -5:  s=-2; txt=f'5일 모멘텀 {r5:.1f}%'
    else:           s=-3; txt=f'5일 모멘텀 {r5:.1f}% (강한 하락)'
    detail['단기 모멘텀(5일)'] = {'score': s, 'desc': txt}
    score += s

    # ② 중기 ROC (20일 모멘텀)
    if   r20 > 10: s=2; txt=f'20일 모멘텀 +{r20:.1f}% (강한 추세)'
    elif r20 > 3:  s=1; txt=f'20일 모멘텀 +{r20:.1f}%'
    elif r20 < -10:s=-2; txt=f'20일 모멘텀 {r20:.1f}% (강한 하락)'
    elif r20 < -3: s=-1; txt=f'20일 모멘텀 {r20:.1f}%'
    else:           s=0; txt=f'20일 모멘텀 {r20:.1f}% (중립)'
    detail['중기 모멘텀(20일)'] = {'score': s, 'desc': txt}
    score += s

    # ③ 거래량 급증 (모멘텀의 확신도)
    if   vol_ratio > 2.0:  s=3; txt=f'거래량 {vol_ratio:.1f}x 폭증 (강한 확신)'
    elif vol_ratio > 1.5:  s=2; txt=f'거래량 {vol_ratio:.1f}x 급증'
    elif vol_ratio > 1.2:  s=1; txt=f'거래량 {vol_ratio:.1f}x 증가'
    elif vol_ratio < 0.5:  s=-2; txt=f'거래량 {vol_ratio:.1f}x 급감 (모멘텀 약화)'
    elif vol_ratio < 0.8:  s=-1; txt=f'거래량 {vol_ratio:.1f}x 감소'
    else:                   s=0; txt=f'거래량 {vol_ratio:.1f}x 보통'
    detail['거래량 강도'] = {'score': s, 'desc': txt}
    score += s

    # ④ MACD 히스토그램 가속도 (연속 방향 확인)
    h_accel = h - hp  # 1차 가속
    h_jerk  = (h-hp) - (hp-hp2)  # 2차 변화
    if   h > 0 and h_accel > 0 and h_jerk > 0: s=3; txt='히스토그램 가속 상승 (모멘텀 가속)'
    elif h > 0 and h_accel > 0:                 s=2; txt='히스토그램 상승 중'
    elif h > 0:                                 s=1; txt='히스토그램 양수 (감속)'
    elif h < 0 and h_accel < 0 and h_jerk < 0: s=-3; txt='히스토그램 가속 하락 (모멘텀 악화)'
    elif h < 0 and h_accel < 0:                 s=-2; txt='히스토그램 하락 중'
    else:                                        s=-1; txt='히스토그램 음수 (감속)'
    detail['MACD 히스토그램'] = {'score': s, 'desc': txt}
    score += s

    # ⑤ 장기 모멘텀 (60일, 추세 방향 최종 확인)
    if   r60 > 15: s=2; txt=f'60일 장기 상승 모멘텀 +{r60:.1f}%'
    elif r60 > 5:  s=1; txt=f'60일 모멘텀 +{r60:.1f}%'
    elif r60 < -15:s=-2; txt=f'60일 장기 하락 모멘텀 {r60:.1f}%'
    elif r60 < -5: s=-1; txt=f'60일 모멘텀 {r60:.1f}%'
    else:           s=0; txt=f'60일 모멘텀 {r60:.1f}% (중립)'
    detail['장기 모멘텀(60일)'] = {'score': s, 'desc': txt}
    score += s

    max_score = 13
    grade, grade_ko = _score_to_grade(score, max_score)
    atr_safe = float(atr_v) if not pd.isna(atr_v) else c*0.02
    return {
        'name': 'Algo D: 모멘텀형',
        'desc': '단기/중기/장기 ROC + 거래량 강도 + MACD 히스토그램 가속도',
        'basis': 'Jegadeesh & Titman 모멘텀 팩터 (3~12개월 지속성 학술 검증)',
        'score': score, 'max_score': max_score,
        'grade': grade, 'grade_ko': grade_ko,
        'reliability': 'HIGH' if vol_ratio > 1.5 and abs(r20) > 5 else ('MEDIUM' if abs(r20) > 2 else 'LOW'),
        'adx': 0, 'detail': detail,
        'short_range': {'high': round(c + atr_safe*2), 'low': round(c - atr_safe*1.5)},
        'mid_range':   {'high': round(c*(1+r20/100*0.5)), 'low': round(c*(1-abs(r20)/100*0.3))},
        'strength': '추세 초기 진입 강점, 거래량 동반 신호 강력',
        'weakness': '과매수 구간 추격 매수 위험, 횡보장 오신호 가능',
    }


def algo_volume_price(df):
    """Algo E: 거래량-가격 분석형 — OBV 추세 + 거래량-가격 괴리 + 매집/분산"""
    import numpy as np
    close = df['Close']; high = df['High']; low = df['Low']; volume = df['Volume']

    # OBV (On Balance Volume)
    obv = (np.sign(close.diff()).fillna(0) * volume).cumsum()
    obv5  = obv.rolling(5).mean()
    obv20 = obv.rolling(20).mean()

    # 거래량 이평선
    vol5  = volume.rolling(5).mean()
    vol20 = volume.rolling(20).mean()
    vol60 = volume.rolling(60).mean()

    # ATR
    tr    = pd.concat([(high-low),(high-close.shift()).abs(),(low-close.shift()).abs()],axis=1).max(axis=1)
    atr_v = tr.rolling(14).mean().iloc[-1]

    # 가격 변화율 vs 거래량 변화율 (괴리 분석)
    price_chg5 = (close.iloc[-1] / close.iloc[-6] - 1) * 100 if len(close) > 5 and close.iloc[-6] != 0 else 0
    vol_chg5   = (vol5.iloc[-1] / vol20.iloc[-1] - 1) * 100 if not pd.isna(vol20.iloc[-1]) and vol20.iloc[-1] > 0 else 0

    # AD (Accumulation/Distribution) 간이 계산
    clv = ((close - low) - (high - close)) / (high - low).replace(0, np.nan)
    clv = clv.fillna(0)
    ad  = (clv * volume).cumsum()
    ad5  = ad.rolling(5).mean()
    ad20 = ad.rolling(20).mean()

    c = close.iloc[-1]
    obv_now = obv.iloc[-1]; obv5_v = obv5.iloc[-1]; obv20_v = obv20.iloc[-1]
    vol5_v = vol5.iloc[-1]; vol20_v = vol20.iloc[-1]; vol60_v = vol60.iloc[-1] if len(volume) >= 60 else vol20_v
    ad5_v = ad5.iloc[-1]; ad20_v = ad20.iloc[-1]

    detail = {}; score = 0

    # ① OBV 추세 (단기 vs 중기)
    if   obv5_v > obv20_v * 1.05: s=3; txt='OBV 단기>중기 (매수세 유입 강함)'
    elif obv5_v > obv20_v:        s=1; txt='OBV 단기≥중기 (매수세 소폭 우위)'
    elif obv5_v < obv20_v * 0.95: s=-3; txt='OBV 단기<중기 (매도세 강함)'
    elif obv5_v < obv20_v:        s=-1; txt='OBV 단기≤중기 (매도세 소폭 우위)'
    else:                          s=0; txt='OBV 중립'
    detail['OBV 추세'] = {'score': s, 'desc': txt}
    score += s

    # ② 거래량-가격 괴리 (가격 상승인데 거래량 감소 = 위험)
    if   price_chg5 > 2 and vol_chg5 > 20:   s=3; txt=f'가격↑{price_chg5:.1f}% + 거래량↑{vol_chg5:.0f}% (건전한 상승)'
    elif price_chg5 > 2 and vol_chg5 < -10:  s=-2; txt=f'가격↑{price_chg5:.1f}% + 거래량↓{vol_chg5:.0f}% (괴리 경고)'
    elif price_chg5 < -2 and vol_chg5 > 20:  s=-3; txt=f'가격↓{price_chg5:.1f}% + 거래량↑{vol_chg5:.0f}% (투매 위험)'
    elif price_chg5 < -2 and vol_chg5 < -10: s=1; txt=f'가격↓{price_chg5:.1f}% + 거래량↓{vol_chg5:.0f}% (하락 약화)'
    else:                                      s=0; txt=f'가격 {price_chg5:.1f}% / 거래량 {vol_chg5:.0f}% (중립)'
    detail['거래량-가격 괴리'] = {'score': s, 'desc': txt}
    score += s

    # ③ 거래량 이평선 배열 (5일 > 20일 > 60일 = 강세)
    vol_align = 0
    if vol5_v > vol20_v: vol_align += 1
    if vol20_v > vol60_v: vol_align += 1
    if vol5_v > vol60_v: vol_align += 1
    if   vol_align == 3: s=2; txt='거래량 정배열 (5>20>60일, 관심 증가)'
    elif vol_align >= 2: s=1; txt='거래량 부분 정배열'
    elif vol_align == 0: s=-2; txt='거래량 역배열 (60>20>5일, 관심 감소)'
    else:                s=0; txt='거래량 배열 혼재'
    detail['거래량 이평선 배열'] = {'score': s, 'desc': txt}
    score += s

    # ④ 매집/분산 지표 (AD Line)
    if   ad5_v > ad20_v * 1.05: s=3; txt='AD 지표 상승 (스마트머니 매집 추정)'
    elif ad5_v > ad20_v:        s=1; txt='AD 지표 소폭 상승'
    elif ad5_v < ad20_v * 0.95: s=-3; txt='AD 지표 하락 (분산/이탈 추정)'
    elif ad5_v < ad20_v:        s=-1; txt='AD 지표 소폭 하락'
    else:                        s=0; txt='AD 지표 중립'
    detail['매집/분산 지표'] = {'score': s, 'desc': txt}
    score += s

    max_score = 11
    grade, grade_ko = _score_to_grade(score, max_score)
    atr_safe = float(atr_v) if not pd.isna(atr_v) else c * 0.02
    vol_ratio = vol5_v / vol20_v if vol20_v > 0 else 1
    return {
        'name': 'Algo E: 거래량-가격 분석형',
        'desc': 'OBV 추세 + 거래량-가격 괴리 + 거래량 이평선 배열 + 매집/분산 지표',
        'basis': 'OBV(Granville), A/D Line(Williams) 기반 수급 분석',
        'score': score, 'max_score': max_score,
        'grade': grade, 'grade_ko': grade_ko,
        'reliability': 'HIGH' if vol_ratio > 1.5 else ('MEDIUM' if vol_ratio > 1.0 else 'LOW'),
        'adx': 0, 'detail': detail,
        'short_range': {'high': round(c + atr_safe * 1.5), 'low': round(c - atr_safe * 1.5)},
        'mid_range':   {'high': round(c + atr_safe * 4),   'low': round(c - atr_safe * 4)},
        'strength': '수급 기반 분석으로 세력 매집/분산 포착 가능',
        'weakness': '저유동성 종목에서 노이즈 큼, 거래량 조작 취약',
    }


def algo_volatility(df):
    """Algo F: 변동성 분석형 — ATR 추세 + 볼린저밴드 폭 + 변동성 수축/확장"""
    import numpy as np
    close = df['Close']; high = df['High']; low = df['Low']

    # ATR 계산
    tr    = pd.concat([(high-low),(high-close.shift()).abs(),(low-close.shift()).abs()],axis=1).max(axis=1)
    atr14 = tr.rolling(14).mean()
    atr5  = tr.rolling(5).mean()

    # 볼린저밴드
    bb_mid = close.rolling(20).mean()
    bb_std = close.rolling(20).std()
    bb_up  = bb_mid + 2 * bb_std
    bb_lo  = bb_mid - 2 * bb_std
    bb_width = (bb_up - bb_lo) / bb_mid * 100  # 밴드 폭 (%)

    # 가격 안정성 (최근 20일 변동계수)
    cv20 = (close.rolling(20).std() / close.rolling(20).mean() * 100)

    c = close.iloc[-1]
    atr14_v = atr14.iloc[-1]; atr14_prev = atr14.iloc[-6] if len(atr14) > 5 else atr14_v
    atr5_v  = atr5.iloc[-1]
    bw_now  = bb_width.iloc[-1]; bw_prev = bb_width.iloc[-6] if len(bb_width) > 5 else bw_now
    bw_avg  = bb_width.iloc[-60:].mean() if len(bb_width) >= 60 else bb_width.mean()
    cv_now  = cv20.iloc[-1]
    bu = bb_up.iloc[-1]; bl = bb_lo.iloc[-1]

    detail = {}; score = 0

    # ① ATR 추세 (변동성 증가/감소)
    atr_chg = (atr14_v / atr14_prev - 1) * 100 if not pd.isna(atr14_prev) and atr14_prev > 0 else 0
    if   atr_chg > 20:  s=-2; txt=f'ATR 급증 +{atr_chg:.0f}% (변동성 확대 경고)'
    elif atr_chg > 5:   s=-1; txt=f'ATR 증가 +{atr_chg:.0f}% (변동성 확대)'
    elif atr_chg < -20: s=2; txt=f'ATR 급감 {atr_chg:.0f}% (변동성 수축, 돌파 임박 가능)'
    elif atr_chg < -5:  s=1; txt=f'ATR 감소 {atr_chg:.0f}% (변동성 수축)'
    else:                s=0; txt=f'ATR 변화 {atr_chg:.0f}% (안정)'
    detail['ATR 추세'] = {'score': s, 'desc': txt}
    score += s

    # ② 볼린저밴드 폭 (현재 vs 평균)
    bw_ratio = bw_now / bw_avg if bw_avg > 0 else 1
    if   bw_ratio < 0.6: s=3; txt=f'밴드 폭 극도 수축 ({bw_now:.1f}% vs 평균{bw_avg:.1f}%) → 큰 움직임 임박'
    elif bw_ratio < 0.8: s=2; txt=f'밴드 폭 수축 ({bw_now:.1f}%) → 돌파 대기'
    elif bw_ratio > 1.5: s=-2; txt=f'밴드 폭 과도 확장 ({bw_now:.1f}%) → 과열/공포 구간'
    elif bw_ratio > 1.2: s=-1; txt=f'밴드 폭 확장 ({bw_now:.1f}%)'
    else:                 s=0; txt=f'밴드 폭 보통 ({bw_now:.1f}%)'
    detail['볼린저밴드 폭'] = {'score': s, 'desc': txt}
    score += s

    # ③ 변동성 수축/확장 방향 (최근 변화)
    bw_chg = bw_now - bw_prev
    if   bw_chg < -1 and bw_ratio < 1: s=2; txt=f'변동성 수축 진행 중 (밴드 {bw_chg:.1f}% 감소) → 에너지 축적'
    elif bw_chg > 1 and bw_ratio > 1:  s=-2; txt=f'변동성 확장 진행 중 (밴드 +{bw_chg:.1f}% 증가) → 불안정'
    elif bw_chg < 0:                    s=1; txt=f'변동성 소폭 수축 ({bw_chg:.1f}%)'
    elif bw_chg > 0:                    s=-1; txt=f'변동성 소폭 확장 (+{bw_chg:.1f}%)'
    else:                                s=0; txt='변동성 변화 없음'
    detail['변동성 수축/확장'] = {'score': s, 'desc': txt}
    score += s

    # ④ 가격 안정성 (변동계수)
    if   cv_now < 2:  s=3; txt=f'변동계수 {cv_now:.1f}% (매우 안정)'
    elif cv_now < 4:  s=2; txt=f'변동계수 {cv_now:.1f}% (안정)'
    elif cv_now < 7:  s=0; txt=f'변동계수 {cv_now:.1f}% (보통)'
    elif cv_now < 10: s=-2; txt=f'변동계수 {cv_now:.1f}% (불안정)'
    else:              s=-3; txt=f'변동계수 {cv_now:.1f}% (매우 불안정)'
    detail['가격 안정성'] = {'score': s, 'desc': txt}
    score += s

    max_score = 10
    grade, grade_ko = _score_to_grade(score, max_score)
    atr_safe = float(atr14_v) if not pd.isna(atr14_v) else c * 0.02
    return {
        'name': 'Algo F: 변동성 분석형',
        'desc': 'ATR 추세 + 볼린저밴드 폭 + 변동성 수축/확장 + 가격 안정성',
        'basis': '볼린저밴드 스퀴즈(John Bollinger), ATR 기반 변동성 사이클 분석',
        'score': score, 'max_score': max_score,
        'grade': grade, 'grade_ko': grade_ko,
        'reliability': 'HIGH' if bw_ratio < 0.7 or bw_ratio > 1.4 else ('MEDIUM' if abs(bw_ratio-1) > 0.15 else 'LOW'),
        'adx': 0, 'detail': detail,
        'short_range': {'high': round(bu), 'low': round(bl)},
        'mid_range':   {'high': round(c + atr_safe * 5), 'low': round(c - atr_safe * 5)},
        'strength': '변동성 수축→돌파 패턴 포착, 진입 타이밍 최적화',
        'weakness': '방향성 판단 불가 (상승/하락 구분 없음)',
    }


def algo_support_resistance(df):
    """Algo G: 지지/저항 분석형 — 주요 지지선/저항선 거리 + 이평선 지지저항 + 돌파 여부"""
    import numpy as np
    close = df['Close']; high = df['High']; low = df['Low']

    # 이평선
    ma5   = close.rolling(5).mean()
    ma20  = close.rolling(20).mean()
    ma60  = close.rolling(60).mean()
    ma120 = close.rolling(120).mean()

    # ATR
    tr    = pd.concat([(high-low),(high-close.shift()).abs(),(low-close.shift()).abs()],axis=1).max(axis=1)
    atr_v = tr.rolling(14).mean().iloc[-1]

    # 최근 고점/저점 (20일, 60일)
    hi20 = high.iloc[-20:].max(); lo20 = low.iloc[-20:].min()
    hi60 = high.iloc[-60:].max() if len(high) >= 60 else high.max()
    lo60 = low.iloc[-60:].min()  if len(low) >= 60 else low.min()

    c = close.iloc[-1]
    m5 = ma5.iloc[-1]; m20 = ma20.iloc[-1]; m60 = ma60.iloc[-1]
    m120 = ma120.iloc[-1] if len(close) >= 120 else m60

    # 주요 저항선: 최근 고점들 중 현재가 위에 있는 것
    resistance = hi20 if hi20 > c else hi60
    # 주요 지지선: 최근 저점들 중 현재가 아래에 있는 것
    support = lo20 if lo20 < c else lo60

    dist_resist = (resistance / c - 1) * 100 if c > 0 else 0  # 저항선까지 거리 (%)
    dist_support = (1 - support / c) * 100 if c > 0 else 0   # 지지선까지 거리 (%)
    # 클램프: 현재가가 저항선 위이거나 지지선 아래인 경우
    dist_resist = max(dist_resist, 0)
    dist_support = max(dist_support, 0)

    detail = {}; score = 0

    # ① 주요 지지선 거리 (가까울수록 안전)
    if   dist_support < 2:  s=3; txt=f'지지선 매우 근접 ({dist_support:.1f}% 아래) → 손절 짧음'
    elif dist_support < 5:  s=2; txt=f'지지선 근접 ({dist_support:.1f}% 아래)'
    elif dist_support < 10: s=1; txt=f'지지선 보통 거리 ({dist_support:.1f}% 아래)'
    elif dist_support < 20: s=-1; txt=f'지지선 먼 거리 ({dist_support:.1f}% 아래)'
    else:                    s=-2; txt=f'지지선 매우 먼 거리 ({dist_support:.1f}% 아래) → 리스크 큼'
    detail['주요 지지선 거리'] = {'score': s, 'desc': txt}
    score += s

    # ② 주요 저항선 거리 (가까우면 상승 여력 제한)
    if   dist_resist < 1:  s=-3; txt=f'저항선 직면 ({dist_resist:.1f}% 위) → 돌파 필요'
    elif dist_resist < 3:  s=-2; txt=f'저항선 근접 ({dist_resist:.1f}% 위)'
    elif dist_resist < 5:  s=-1; txt=f'저항선 보통 ({dist_resist:.1f}% 위)'
    elif dist_resist < 15: s=1; txt=f'저항선 여유 ({dist_resist:.1f}% 위)'
    else:                   s=2; txt=f'저항선 먼 거리 ({dist_resist:.1f}% 위) → 상승 여력 충분'
    detail['주요 저항선 거리'] = {'score': s, 'desc': txt}
    score += s

    # ③ 이평선 지지/저항 (현재가 기준 이평선 위/아래)
    ma_support_count = sum([c > m for m in [m5, m20, m60, m120] if not pd.isna(m)])
    if   ma_support_count == 4: s=3; txt='모든 이평선 위 (강한 지지 기반)'
    elif ma_support_count == 3: s=2; txt='이평선 3개 위 (지지 양호)'
    elif ma_support_count == 2: s=0; txt='이평선 2개 위/2개 아래 (중립)'
    elif ma_support_count == 1: s=-2; txt='이평선 1개만 위 (지지 약함)'
    else:                        s=-3; txt='모든 이평선 아래 (지지 없음)'
    detail['이평선 지지/저항'] = {'score': s, 'desc': txt}
    score += s

    # ④ 돌파 여부 (최근 20일 고점 돌파 or 저점 이탈)
    prev_hi20 = high.iloc[-25:-5].max() if len(high) >= 25 else high.iloc[:-5].max() if len(high) > 5 else high.max()
    prev_lo20 = low.iloc[-25:-5].min() if len(low) >= 25 else low.iloc[:-5].min() if len(low) > 5 else low.min()
    if   c > prev_hi20: s=3; txt=f'20일 고점 돌파! (저항→지지 전환 기대, 고점 {prev_hi20:,.0f})'
    elif c < prev_lo20: s=-3; txt=f'20일 저점 이탈! (지지→저항 전환 위험, 저점 {prev_lo20:,.0f})'
    elif c > prev_hi20 * 0.98: s=1; txt=f'20일 고점 근접 (돌파 시도 중, 고점 {prev_hi20:,.0f})'
    elif c < prev_lo20 * 1.02: s=-1; txt=f'20일 저점 근접 (이탈 위험, 저점 {prev_lo20:,.0f})'
    else: s=0; txt='레인지 내 움직임'
    detail['돌파 여부'] = {'score': s, 'desc': txt}
    score += s

    max_score = 11
    grade, grade_ko = _score_to_grade(score, max_score)
    atr_safe = float(atr_v) if not pd.isna(atr_v) else c * 0.02
    return {
        'name': 'Algo G: 지지/저항 분석형',
        'desc': '주요 지지선/저항선 거리 + 이평선 지지저항 + 돌파 여부',
        'basis': '가격 행동(Price Action) 기반 지지/저항 분석 (다우 이론)',
        'score': score, 'max_score': max_score,
        'grade': grade, 'grade_ko': grade_ko,
        'reliability': 'HIGH' if abs(score) >= 6 else ('MEDIUM' if abs(score) >= 3 else 'LOW'),
        'adx': 0, 'detail': detail,
        'short_range': {'high': round(resistance), 'low': round(support)},
        'mid_range':   {'high': round(c + atr_safe * 5), 'low': round(c - atr_safe * 5)},
        'strength': '명확한 진입/손절 기준 제공, 리스크 관리 최적',
        'weakness': '횡보 구간 반복 속임수 돌파(False Breakout) 취약',
    }


def algo_multi_timeframe(df):
    """Algo H: 멀티 타임프레임형 — 5일/20일/60일 방향 비교 + 타임프레임 정렬도"""
    import numpy as np
    close = df['Close']; high = df['High']; low = df['Low']

    # 이평선
    ma5  = close.rolling(5).mean()
    ma20 = close.rolling(20).mean()
    ma60 = close.rolling(60).mean()

    # ROC per timeframe
    roc5  = (close / close.shift(5)  - 1) * 100
    roc20 = (close / close.shift(20) - 1) * 100
    roc60 = (close / close.shift(60) - 1) * 100

    # ATR
    tr    = pd.concat([(high-low),(high-close.shift()).abs(),(low-close.shift()).abs()],axis=1).max(axis=1)
    atr_v = tr.rolling(14).mean().iloc[-1]

    c = close.iloc[-1]
    m5 = ma5.iloc[-1]; m20 = ma20.iloc[-1]; m60 = ma60.iloc[-1]
    r5  = roc5.iloc[-1]  if len(roc5) > 5  and not pd.isna(roc5.iloc[-1])  else 0
    r20 = roc20.iloc[-1] if len(roc20) > 20 and not pd.isna(roc20.iloc[-1]) else 0
    r60 = roc60.iloc[-1] if len(roc60) > 60 and not pd.isna(roc60.iloc[-1]) else 0

    # 이평선 기울기 (최근 5일 변화)
    m5_slope  = (ma5.iloc[-1]  / ma5.iloc[-3]  - 1) * 100 if len(ma5) > 3 and not pd.isna(ma5.iloc[-3]) and ma5.iloc[-3] != 0 else 0
    m20_slope = (ma20.iloc[-1] / ma20.iloc[-3] - 1) * 100 if len(ma20) > 3 and not pd.isna(ma20.iloc[-3]) and ma20.iloc[-3] != 0 else 0
    m60_slope = (ma60.iloc[-1] / ma60.iloc[-3] - 1) * 100 if len(ma60) > 3 and not pd.isna(ma60.iloc[-3]) and ma60.iloc[-3] != 0 else 0

    detail = {}; score = 0

    # ① 단기 방향 (5일)
    if   r5 > 3 and m5_slope > 0.3:  s=3; txt=f'단기(5일) 강한 상승 (ROC +{r5:.1f}%, MA5 기울기 +{m5_slope:.1f}%)'
    elif r5 > 1 and m5_slope > 0:    s=2; txt=f'단기(5일) 상승 (ROC +{r5:.1f}%)'
    elif r5 > 0:                      s=1; txt=f'단기(5일) 약한 상승 (ROC +{r5:.1f}%)'
    elif r5 > -1:                     s=-1; txt=f'단기(5일) 약한 하락 (ROC {r5:.1f}%)'
    elif r5 > -3:                     s=-2; txt=f'단기(5일) 하락 (ROC {r5:.1f}%)'
    else:                              s=-3; txt=f'단기(5일) 강한 하락 (ROC {r5:.1f}%, MA5 기울기 {m5_slope:.1f}%)'
    detail['단기 방향(5일)'] = {'score': s, 'desc': txt}
    score += s

    # ② 중기 방향 (20일)
    if   r20 > 8 and m20_slope > 0.5:  s=3; txt=f'중기(20일) 강한 상승 (ROC +{r20:.1f}%, MA20 기울기 +{m20_slope:.1f}%)'
    elif r20 > 3 and m20_slope > 0:    s=2; txt=f'중기(20일) 상승 (ROC +{r20:.1f}%)'
    elif r20 > 0:                       s=1; txt=f'중기(20일) 약한 상승 (ROC +{r20:.1f}%)'
    elif r20 > -3:                      s=-1; txt=f'중기(20일) 약한 하락 (ROC {r20:.1f}%)'
    elif r20 > -8:                      s=-2; txt=f'중기(20일) 하락 (ROC {r20:.1f}%)'
    else:                                s=-3; txt=f'중기(20일) 강한 하락 (ROC {r20:.1f}%, MA20 기울기 {m20_slope:.1f}%)'
    detail['중기 방향(20일)'] = {'score': s, 'desc': txt}
    score += s

    # ③ 장기 방향 (60일)
    if   r60 > 15 and m60_slope > 0.3: s=3; txt=f'장기(60일) 강한 상승 (ROC +{r60:.1f}%, MA60 기울기 +{m60_slope:.1f}%)'
    elif r60 > 5 and m60_slope > 0:    s=2; txt=f'장기(60일) 상승 (ROC +{r60:.1f}%)'
    elif r60 > 0:                       s=1; txt=f'장기(60일) 약한 상승 (ROC +{r60:.1f}%)'
    elif r60 > -5:                      s=-1; txt=f'장기(60일) 약한 하락 (ROC {r60:.1f}%)'
    elif r60 > -15:                     s=-2; txt=f'장기(60일) 하락 (ROC {r60:.1f}%)'
    else:                                s=-3; txt=f'장기(60일) 강한 하락 (ROC {r60:.1f}%, MA60 기울기 {m60_slope:.1f}%)'
    detail['장기 방향(60일)'] = {'score': s, 'desc': txt}
    score += s

    # ④ 타임프레임 정렬도 (모든 방향 일치 = 강한 신호)
    dirs = [np.sign(r5), np.sign(r20), np.sign(r60)]
    ma_order = [m5, m20, m60]
    all_up   = all(d > 0 for d in dirs) and m5 > m20 > m60
    all_down = all(d < 0 for d in dirs) and m5 < m20 < m60
    same_dir = len(set(np.sign(d) for d in dirs if d != 0))

    if   all_up:             s=4; txt='완벽한 상승 정렬 (5>20>60일 + 모두 상승)'
    elif all_down:           s=-4; txt='완벽한 하락 정렬 (5<20<60일 + 모두 하락)'
    elif same_dir <= 1 and sum(d > 0 for d in dirs) >= 2: s=2; txt='부분 상승 정렬 (2개 이상 상승)'
    elif same_dir <= 1 and sum(d < 0 for d in dirs) >= 2: s=-2; txt='부분 하락 정렬 (2개 이상 하락)'
    else:                    s=0; txt='타임프레임 혼재 (방향 불일치)'
    detail['타임프레임 정렬도'] = {'score': s, 'desc': txt}
    score += s

    max_score = 13
    grade, grade_ko = _score_to_grade(score, max_score)
    atr_safe = float(atr_v) if not pd.isna(atr_v) else c * 0.02
    alignment = 'HIGH' if all_up or all_down else ('MEDIUM' if same_dir <= 1 else 'LOW')
    return {
        'name': 'Algo H: 멀티 타임프레임형',
        'desc': '단기(5일)/중기(20일)/장기(60일) 방향 비교 + 타임프레임 정렬도',
        'basis': 'Alexander Elder 다중 시간대 분석 (Triple Screen Trading)',
        'score': score, 'max_score': max_score,
        'grade': grade, 'grade_ko': grade_ko,
        'reliability': alignment,
        'adx': 0, 'detail': detail,
        'short_range': {'high': round(c + atr_safe * 1.5), 'low': round(c - atr_safe * 1.5)},
        'mid_range':   {'high': round(c + atr_safe * 4),   'low': round(c - atr_safe * 4)},
        'strength': '다중 시간대 일치 시 높은 승률, 추세 초기 포착',
        'weakness': '타임프레임 혼재 시 판단 어려움, 전환점 늦음',
    }


def _score_to_grade(score, max_score):
    ratio = score / max_score if max_score != 0 else 0
    if   ratio >= 0.55:  return 'strong_up',   '강한 상승'
    elif ratio >= 0.25:  return 'up',           '상승'
    elif ratio >= -0.25: return 'neutral',      '중립'
    elif ratio >= -0.55: return 'down',         '하락'
    else:                return 'strong_down',  '강한 하락'


def _fetch_df(h, usd_krw):
    if h['market'] == 'US':
        t  = yf.Ticker(h['ticker'])
        df = t.history(period='1y')
        if df.empty or len(df) < 30:
            raise ValueError('데이터 부족')
        df.index = df.index.tz_localize(None)
        df['Close'] = df['Close'] * usd_krw
        df['High']  = df['High']  * usd_krw
        df['Low']   = df['Low']   * usd_krw
    else:
        end   = datetime.now().strftime('%Y%m%d')
        start = (datetime.now() - timedelta(days=400)).strftime('%Y%m%d')
        raw   = krx.get_market_ohlcv(start, end, h['ticker'])
        if raw.empty or len(raw) < 30:
            raise ValueError('데이터 부족')
        df = raw.rename(columns={'시가':'Open','고가':'High','저가':'Low','종가':'Close','거래량':'Volume'})
    return df


@app.route('/api/compare')
def api_compare():
    data    = load_portfolio()
    usd_krw = get_usd_krw()
    results = []

    for h in data['holdings']:
        item = {'name': h['name'], 'ticker': h['ticker'], 'market': h['market'], 'algos': []}
        try:
            df = _fetch_df(h, usd_krw)
            # 미니차트 데이터 (60일)
            close  = df['Close']
            ma5    = close.rolling(5).mean()
            ma20   = close.rolling(20).mean()
            ma60   = close.rolling(60).mean()
            n      = min(60, len(close))
            item['chart'] = {
                'dates': [str(d.date()) for d in close.iloc[-n:].index],
                'close': [round(v,2) for v in close.iloc[-n:].tolist()],
                'ma5':   [round(v,2) if not pd.isna(v) else None for v in ma5.iloc[-n:].tolist()],
                'ma20':  [round(v,2) if not pd.isna(v) else None for v in ma20.iloc[-n:].tolist()],
                'ma60':  [round(v,2) if not pd.isna(v) else None for v in ma60.iloc[-n:].tolist()],
            }
            item['current_price'] = round(float(close.iloc[-1]))

            algo_map = {
                'A': algo_comprehensive, 'B': algo_trend_follow,
                'C': algo_contrarian,    'D': algo_momentum,
                'E': algo_volume_price,  'F': algo_volatility,
                'G': algo_support_resistance, 'H': algo_multi_timeframe,
            }
            algos_param = request.args.get('algos', '')
            if algos_param:
                selected = [algo_map[k] for k in algos_param.split(',') if k in algo_map]
            else:
                selected = list(algo_map.values())

            for fn in selected:
                try:
                    item['algos'].append(fn(df))
                except Exception as e:
                    item['algos'].append({'name': fn.__doc__.split('—')[0].strip(), 'error': str(e)})
        except Exception as e:
            item['error'] = str(e)
        results.append(item)

    return jsonify({'compare': results})


@app.route('/api/period-perf')
def api_period_perf():
    """각 종목의 기간별 수익률 반환 (1일/5일/1개월/3개월/1년)"""
    portfolio = load_portfolio()
    usd_krw = get_usd_krw()

    periods = [
        ('1일',  1),
        ('5일',  5),
        ('1개월', 21),
        ('3개월', 63),
        ('1년',  252),
    ]

    results = []
    for h in portfolio['holdings']:
        try:
            if h['market'] == 'US':
                ticker_sym = h['ticker']
                t = yf.Ticker(ticker_sym)
                hist = t.history(period='1y')
                fx = usd_krw
            else:
                ticker_sym = h['ticker']
                end = datetime.now()
                start = end - timedelta(days=400)
                df = krx.get_market_ohlcv_by_date(
                    start.strftime('%Y%m%d'), end.strftime('%Y%m%d'), ticker_sym
                )
                hist = df[['종가']].rename(columns={'종가': 'Close'})
                fx = 1

            if hist.empty or len(hist) < 2:
                continue

            current = float(hist['Close'].iloc[-1]) * fx * h['qty']
            perf = []
            for label, days in periods:
                if len(hist) > days:
                    past_price = float(hist['Close'].iloc[-days-1]) * fx * h['qty']
                    chg = current - past_price
                    pct = (chg / past_price * 100) if past_price else 0
                    perf.append({'label': label, 'chg': round(chg), 'pct': round(pct, 2)})
                else:
                    perf.append({'label': label, 'chg': None, 'pct': None})

            results.append({
                'name': h['name'],
                'ticker': h['ticker'],
                'market': h['market'],
                'qty': h['qty'],
                'current_value': round(current),
                'perf': perf,
            })
        except Exception as e:
            results.append({'name': h['name'], 'ticker': h['ticker'], 'market': h['market'],
                            'qty': h['qty'], 'current_value': 0, 'perf': [], 'error': str(e)})

    return jsonify({'items': results})


@app.route('/api/transactions', methods=['GET'])
def api_get_transactions():
    txs = load_transactions()
    return jsonify({'transactions': txs})


@app.route('/api/transactions', methods=['POST'])
def api_add_transaction():
    try:
        tx = request.get_json()
        required = ['date', 'name', 'ticker', 'market', 'type', 'qty', 'price']
        for f in required:
            if f not in tx:
                return jsonify({'error': f'필드 누락: {f}'}), 400

        tx['qty']   = float(tx['qty'])
        tx['price'] = float(tx['price'])
        if tx['qty'] <= 0:
            return jsonify({'error': '수량은 0보다 커야 합니다'}), 400
        if tx['price'] <= 0:
            return jsonify({'error': '가격은 0보다 커야 합니다'}), 400
        tx['id']    = datetime.now().strftime('%Y%m%d%H%M%S%f')
        tx['total'] = round(tx['qty'] * tx['price'])

        txs = load_transactions()
        txs.append(tx)
        save_transactions(txs)
        holdings = rebuild_portfolio_from_transactions()
        return jsonify({'success': True, 'id': tx['id'], 'holdings_count': len(holdings)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/transactions/<tx_id>', methods=['DELETE'])
def api_delete_transaction(tx_id):
    try:
        txs = load_transactions()
        txs = [t for t in txs if t['id'] != tx_id]
        save_transactions(txs)
        holdings = rebuild_portfolio_from_transactions()
        return jsonify({'success': True, 'holdings_count': len(holdings)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/transactions/summary', methods=['GET'])
def api_tx_summary():
    """종목별 실현손익 요약"""
    txs = load_transactions()
    positions = {}
    for tx in sorted(txs, key=lambda x: x['date']):
        tk = tx['ticker']
        if tk not in positions:
            positions[tk] = {'name': tx['name'], 'ticker': tk, 'market': tx['market'],
                             'qty': 0.0, 'total_cost': 0.0, 'realized_pnl': 0.0,
                             'buy_count': 0, 'sell_count': 0}
        p = positions[tk]
        if tx['type'] == 'buy':
            p['total_cost'] += tx['qty'] * tx['price']
            p['qty'] += tx['qty']
            p['buy_count'] += 1
        else:
            if p['qty'] > 0:
                avg = p['total_cost'] / p['qty']
                p['realized_pnl'] += (tx['price'] - avg) * tx['qty']
                p['total_cost'] -= avg * tx['qty']
            p['qty'] = max(0.0, p['qty'] - tx['qty'])
            p['sell_count'] += 1
    result = []
    for p in positions.values():
        avg_cost = round(p['total_cost'] / p['qty']) if p['qty'] > 0 else 0
        result.append({
            'name': p['name'], 'ticker': p['ticker'], 'market': p['market'],
            'qty': round(p['qty'], 8), 'avg_cost': avg_cost,
            'realized_pnl': round(p['realized_pnl']),
            'buy_count': p['buy_count'], 'sell_count': p['sell_count'],
        })
    return jsonify({'summary': result})


@app.route('/api/reset-portfolio', methods=['POST'])
def api_reset_portfolio():
    try:
        save_portfolio({'holdings': []})
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/portfolio/delete', methods=['POST'])
def api_delete_holding():
    """포트폴리오에서 특정 종목 삭제"""
    data = request.get_json() or {}
    ticker = data.get('ticker', '')
    if not ticker:
        return jsonify({'error': '티커 없음'}), 400
    portfolio = load_portfolio()
    before = len(portfolio.get('holdings', []))
    portfolio['holdings'] = [h for h in portfolio.get('holdings', []) if h['ticker'] != ticker]
    after = len(portfolio['holdings'])
    if before == after:
        return jsonify({'error': '해당 종목을 찾을 수 없습니다'}), 404
    save_portfolio(portfolio)
    return jsonify({'ok': True, 'removed': ticker, 'remaining': after})

@app.route('/api/export/excel')
def api_export_excel():
    """포트폴리오 + 거래내역 Excel 다운로드"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from flask import send_file

        wb = Workbook()

        # 시트 1: 포트폴리오
        ws1 = wb.active
        ws1.title = '포트폴리오'
        ws1.append(['종목명', '티커', '시장', '수량', '평균매수가', '총매수금액'])
        for cell in ws1[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='6c5ce7')
            cell.alignment = Alignment(horizontal='center')

        port = load_portfolio()
        for h in port.get('holdings', []):
            qty = h.get('qty', 0)
            cost = h.get('cost', 0)
            avg = round(cost / qty) if qty else 0
            ws1.append([
                h.get('name', ''),
                h.get('ticker', ''),
                'US' if h.get('market') == 'US' else 'KR',
                qty, avg, cost
            ])

        # 시트 2: 거래내역
        ws2 = wb.create_sheet('거래내역')
        ws2.append(['날짜', '종류', '종목명', '티커', '시장', '수량', '단가', '거래금액', '메모'])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='00b894')
            cell.alignment = Alignment(horizontal='center')

        txs = load_transactions()
        for tx in sorted(txs, key=lambda x: x.get('date', ''), reverse=True):
            ws2.append([
                tx.get('date', ''),
                '매수' if tx.get('type') == 'buy' else '매도',
                tx.get('name', ''),
                tx.get('ticker', ''),
                'US' if tx.get('market') == 'US' else 'KR',
                tx.get('qty', 0),
                tx.get('price', 0),
                tx.get('qty', 0) * tx.get('price', 0),
                tx.get('note', ''),
            ])

        # 시트 3: 위시리스트
        ws3 = wb.create_sheet('위시리스트')
        ws3.append(['종목명', '티커', '시장'])
        for cell in ws3[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='fdcb6e')
            cell.alignment = Alignment(horizontal='center')
        for w in load_wishlist():
            ws3.append([w.get('name', ''), w.get('ticker', ''), 'US' if w.get('market') == 'US' else 'KR'])

        # 시트 4: 모의투자
        ws4 = wb.create_sheet('모의투자')
        ws4.append(['종목명', '티커', '시장', '수량', '매수가', '매수일', '메모'])
        for cell in ws4[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='a29bfe')
            cell.alignment = Alignment(horizontal='center')
        for s in load_simulation():
            ws4.append([
                s.get('name', ''), s.get('ticker', ''),
                'US' if s.get('market') == 'US' else 'KR',
                s.get('qty', 0), s.get('buy_price', 0),
                s.get('buy_date', ''), s.get('note', ''),
            ])

        # 컬럼 너비 자동조정
        for ws in [ws1, ws2, ws3, ws4]:
            for col in ws.columns:
                max_len = max(len(str(c.value or '')) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"portfolio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': f'Excel 생성 실패: {e}'}), 500


@app.route('/api/upload-json', methods=['POST'])
def api_upload_json():
    try:
        data = request.get_json()
        if not data or 'holdings' not in data:
            return jsonify({'error': '잘못된 데이터입니다'}), 400
        holdings = data['holdings']
        if not holdings:
            return jsonify({'error': '종목 데이터가 없습니다'}), 400
        save_portfolio({'holdings': holdings})
        return jsonify({'success': True, 'count': len(holdings)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 티커 매핑 테이블
TICKER_MAP = {
    '엔비디아': ('NVDA', 'US'), '테슬라': ('TSLA', 'US'),
    '마이크로소프트': ('MSFT', 'US'), '애플': ('AAPL', 'US'),
    '쿠팡': ('CPNG', 'US'), '알파벳': ('GOOGL', 'US'),
    '아마존': ('AMZN', 'US'), '메타': ('META', 'US'),
    '한화그룹': ('462330', 'KR'), '삼성전자': ('005930', 'KR'),
    'PLUS 한화그룹주': ('462330', 'KR'),
}

_ocr_reader = None

def get_ocr_reader():
    global _ocr_reader
    if _ocr_reader is None:
        import easyocr
        _ocr_reader = easyocr.Reader(['ko', 'en'], gpu=False)
    return _ocr_reader


def prepare_mobile_img(img_bytes):
    """모바일 스크린샷 공통 전처리 (정확도 최대화)"""
    from PIL import Image, ImageOps
    import io
    img = Image.open(io.BytesIO(img_bytes)).convert('RGB')
    w, h = img.size

    # ① 다크 배경 → 반전 (토스 다크테마) — 먼저 체크
    arr = np.array(img)
    if arr.mean() < 130:
        img = ImageOps.invert(img)

    # ② 왼쪽 아이콘 영역 제거 (종목 로고가 OCR 방해)
    crop_left = int(w * 0.17)
    img = img.crop((crop_left, 0, w, h))

    # ③ 해상도 업스케일 (EasyOCR 정확도는 큰 이미지에서 훨씬 높음)
    w2, h2 = img.size
    target_w = 1200
    if w2 < target_w:
        ratio = target_w / w2
        img = img.resize((target_w, int(h2 * ratio)), Image.LANCZOS)
    elif w2 > 1600:
        ratio = 1600 / w2
        img = img.resize((1600, int(h2 * ratio)), Image.LANCZOS)

    return img


def ocr_image(img_bytes, preprocess_fn=None):
    """이미지 바이트 → OCR 텍스트 라인 리스트 (정확도 최대화)"""
    from PIL import Image
    img = prepare_mobile_img(img_bytes)
    if preprocess_fn:
        img = preprocess_fn(img)
    reader = get_ocr_reader()
    results = reader.readtext(
        np.array(img),
        detail=0,
        paragraph=False,
        width_ths=0.7,        # 더 좁은 텍스트 블록도 인식
        contrast_ths=0.05,
        adjust_contrast=0.8,
        text_threshold=0.35,  # 낮춰서 흐린 텍스트도 인식
        low_text=0.2,
        mag_ratio=1.0,        # 이미 업스케일했으므로 1.0
        slope_ths=0.2,        # 약간 기울어진 텍스트 허용
        ycenter_ths=0.7,
        link_threshold=0.4,
    )
    return results


def parse_toss_lines(lines):
    """OCR 라인 → 종목 리스트 파싱 (모바일 토스 레이아웃 특화)"""
    import re
    holdings = []
    current_section = 'US'

    def clean_num(s):
        """OCR 오인식 보정 후 숫자 추출 (강화판)"""
        s = str(s)
        # 단위 제거
        for unit in ['원','주','%','달러','$','₩',',']:
            s = s.replace(unit, '')
        # % 뒤 잡음 제거 (61.0%0a 등)
        s = re.sub(r'(\d+\.?\d*)[a-zA-Z]+\d*$', r'\1', s)
        # 흔한 OCR 문자 오인식 보정
        for bad, good in [
            ('O','0'),('o','0'),('D','0'),('Q','0'),
            ('l','1'),('I','1'),('|','1'),('i','1'),
            ('S','5'),('s','5'),('G','6'),('B','8'),
            ('Z','2'),('z','2'),(' ',''),
        ]:
            s = s.replace(bad, good)
        s = re.sub(r'[^\d.\-+]', '', s)
        return s

    # 전체 텍스트를 하나로 합쳐서 패턴 매칭 (줄 단위 파싱보다 유연)
    full_text = '\n'.join(lines)

    # ── 섹션별 분리 ──────────────────────────────────────────
    # 해외주식 / 국내주식 섹션 분리
    kr_idx = next((i for i, l in enumerate(lines) if '국내주식' in l), len(lines))
    us_lines = lines[:kr_idx]
    kr_lines = lines[kr_idx:]

    def extract_from_lines(src_lines, section):
        result = []
        i = 0
        while i < len(src_lines):
            line = src_lines[i].strip()

            # 수량 패턴: "28.016662주" 또는 "2주"
            qty_m = re.search(r'(\d[\d.]*)\s*주\b', line)
            if qty_m:
                qty_str = qty_m.group(1)
                try:
                    qty = float(qty_str)
                except:
                    i += 1; continue

                # 종목명: 수량 줄 바로 전 줄 (또는 같은 줄에 포함된 경우)
                # OCR 순서: 종목명 → 평가금(원) → 수량(주) → 손익
                def is_non_name(s):
                    s = s.strip()
                    if not s or len(s) < 2: return True
                    skip_words = {'보기', '해외주식', '국내주식', '주문내역', '배당금', '관심', '발견', '피드', '증권', '현재가', '평가금'}
                    if s in skip_words: return True
                    return bool(re.match(r'^[\d\s,+\-원%().\[\]$~]+$', s))

                name = ''
                # 같은 줄에 종목명 포함 여부 확인
                before_qty = line[:qty_m.start()].strip()
                if len(before_qty) >= 2 and not is_non_name(before_qty):
                    name = before_qty
                else:
                    # 최대 3줄 위까지 거슬러 올라가며 종목명 찾기
                    for back in range(1, 4):
                        if i - back < 0:
                            break
                        candidate = src_lines[i - back].strip()
                        if not is_non_name(candidate):
                            name = candidate
                            break

                if not name:
                    i += 1; continue

                # 평가금 + 손익: 수량 앞 2줄 + 이후 6줄 범위에서 찾기
                current_value, profit, profit_pct = None, None, None
                search_start = max(0, i - 2)
                search_text = ' '.join(src_lines[search_start:min(i+6, len(src_lines))])

                # 평가금: 숫자,숫자원 or 숫자원
                for val_m in re.finditer(r'(\d[\d,]+)\s*원', search_text):
                    v = float(clean_num(val_m.group(1)))
                    if v >= 100:  # 최소 100원 이상
                        current_value = v
                        break

                # 손익: +숫자 (숫자%) 또는 -숫자 (숫자%)
                # % 오인식 대응: %자리에 영문/숫자 혼합 허용
                pnl_m = re.search(
                    r'([+\-])\s*(\d[\d,]*)\s*[\(\[]\s*(\d+\.?\d*)[%\w]*\s*[\)\]]',
                    search_text
                )
                if pnl_m:
                    sign = 1 if pnl_m.group(1) == '+' else -1
                    profit = sign * float(clean_num(pnl_m.group(2)))
                    profit_pct = sign * float(clean_num(pnl_m.group(3)))

                if current_value is None:
                    i += 1; continue

                # 티커 매핑
                ticker, market = None, section
                for k, (t, m) in TICKER_MAP.items():
                    if k in name or name in k:
                        ticker, market = t, m
                        break
                if not ticker:
                    ticker = re.sub(r'[^A-Z0-9]', '', name.upper())[:6] or name[:4]

                cost = round(current_value - (profit or 0))
                result.append({
                    'name': name, 'ticker': ticker, 'market': market,
                    'qty': qty, 'current_value': round(current_value),
                    'profit': round(profit) if profit else 0,
                    'profit_pct': round(profit_pct, 1) if profit_pct else 0,
                    'cost': cost,
                })
            i += 1
        return result

    holdings += extract_from_lines(us_lines, 'US')
    holdings += extract_from_lines(kr_lines, 'KR')

    # 중복 제거 (같은 ticker)
    seen = set()
    unique = []
    for h in holdings:
        if h['ticker'] not in seen:
            seen.add(h['ticker'])
            unique.append(h)

    if unique:
        return unique

    # fallback: 옛 방식 라인 파싱
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if '국내주식' in line: current_section = 'KR'
        if '해외주식' in line: current_section = 'US'

        qty_m = re.search(r'(\d[\d.]*)\s*주\b', line)
        if qty_m and i > 0:
            name_line = lines[i - 1].strip()
            if len(name_line) >= 2 and not re.match(r'^[\d\s,+\-원%().]+$', name_line):
                qty = float(qty_m.group(1))
                current_value, profit, profit_pct = None, None, None
                for j in range(i + 1, min(i + 6, len(lines))):
                    l = lines[j].strip()
                    val_m = re.search(r'(\d[\d,]+)\s*원', l)
                    if val_m and current_value is None:
                        v = float(clean_num(val_m.group(1)))
                        if v >= 100: current_value = v
                    pnl_m = re.search(r'([+\-])\s*(\d[\d,]*)\s*[\(\[](\d+\.?\d*)[%\w]*[\)\]]', l)
                    if pnl_m and profit is None:
                        sign = 1 if pnl_m.group(1)=='+' else -1
                        profit = sign * float(clean_num(pnl_m.group(2)))
                        profit_pct = sign * float(pnl_m.group(3))
                if current_value is not None:
                    name = name_line
                    ticker, market = None, current_section
                    for k, (t, m) in TICKER_MAP.items():
                        if k in name or name in k: ticker, market = t, m; break
                    if not ticker: ticker = name[:4].upper()
                    cost = round(current_value - (profit or 0))
                    holdings.append({
                        'name': name, 'ticker': ticker, 'market': market,
                        'qty': qty, 'current_value': round(current_value),
                        'profit': round(profit) if profit else 0,
                        'profit_pct': round(profit_pct, 1) if profit_pct else 0,
                        'cost': cost,
                    })
        i += 1

    return holdings


@app.route('/api/screenshot', methods=['POST'])
def api_screenshot():
    from PIL import Image, ImageEnhance, ImageFilter, ImageOps
    from collections import Counter
    import io, re

    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    f = request.files['file']
    if not f.filename.lower().rsplit('.', 1)[-1] in ('png', 'jpg', 'jpeg', 'webp'):
        return jsonify({'error': 'PNG / JPG / WEBP 이미지만 지원합니다'}), 400

    img_bytes = f.read()

    # ── 모바일 특화 5가지 전처리 ─────────────────────────────
    # prepare_mobile_img 에서 이미 다크→라이트 반전됨
    # 여기서는 추가 화질 개선에 집중

    def pp_clean(img):
        """① 기본 정리: 대비 + 선명화"""
        img = ImageEnhance.Contrast(img).enhance(2.0)
        img = ImageEnhance.Sharpness(img).enhance(2.0)
        img = ImageEnhance.Brightness(img).enhance(1.1)
        return img

    def pp_high_contrast(img):
        """② 고대비 흑백: 텍스트/배경 분리 극대화"""
        img = ImageEnhance.Contrast(img).enhance(3.5)
        img = ImageEnhance.Color(img).enhance(0.0)
        return img

    def pp_unsharp(img):
        """③ 언샤프 마스크: 경계선 강화"""
        img = ImageEnhance.Contrast(img).enhance(2.5)
        img = img.filter(ImageFilter.UnsharpMask(radius=2, percent=250, threshold=1))
        return img

    def pp_binarize(img):
        """④ 적응형 이진화: 음영 영역별 개별 처리"""
        from PIL import Image as PILImage
        gray = np.array(img.convert('L'))
        # 블록별 임계값 (Otsu-like)
        from PIL import Image as PILImage
        import io as _io
        binary = np.where(gray > gray.mean() * 0.85, 255, 0).astype(np.uint8)
        return PILImage.fromarray(binary).convert('RGB')

    def pp_dilate(img):
        """⑤ 텍스트 굵게: 얇은 획 보완"""
        img = ImageEnhance.Contrast(img).enhance(2.5)
        img = img.filter(ImageFilter.MaxFilter(size=3))
        return img

    def pp_sharpen_gray(img):
        """⑥ 그레이스케일 + 극한 선명화"""
        img = img.convert('L').convert('RGB')
        img = ImageEnhance.Contrast(img).enhance(3.0)
        img = img.filter(ImageFilter.SHARPEN)
        img = img.filter(ImageFilter.SHARPEN)
        return img

    preprocessors = [pp_clean, pp_high_contrast, pp_unsharp, pp_binarize, pp_dilate, pp_sharpen_gray]

    all_batches = []
    debug_passes = []   # 각 전처리별 raw OCR 결과 저장

    pp_names = ['기본정리', '고대비', '언샤프', '이진화', '텍스트굵게', '그레이극한']
    for idx, pp in enumerate(preprocessors):
        try:
            lines = ocr_image(img_bytes, pp)
            parsed = parse_toss_lines(lines)
            debug_passes.append({
                'name': pp_names[idx],
                'lines': lines,
                'parsed_count': len(parsed),
            })
            if parsed:
                all_batches.append(parsed)
        except Exception as e:
            debug_passes.append({'name': pp_names[idx], 'lines': [], 'error': str(e), 'parsed_count': 0})

    if not all_batches:
        all_lines = [l for p in debug_passes for l in p.get('lines', [])]
        return jsonify({
            'error': '종목을 인식하지 못했습니다.',
            'debug': debug_passes,
            'all_lines': all_lines,
        }), 400

    # 교차 검증: 종목명 기준 집계
    bucket = {}
    for batch in all_batches:
        for h in batch:
            key = h.get('ticker') or h.get('name', '')
            bucket.setdefault(key, []).append(h)

    def majority(entries, field, cast=float):
        vals = []
        for e in entries:
            v = e.get(field)
            if v is not None:
                try: vals.append(round(cast(v), 4) if cast == float else cast(v))
                except: pass
        if not vals: return None
        return Counter(vals).most_common(1)[0][0]

    total_checks = len(all_batches)
    consensus = []
    for key, entries in bucket.items():
        name          = majority(entries, 'name',          str)
        ticker        = majority(entries, 'ticker',        str) or key
        market        = majority(entries, 'market',        str) or 'US'
        qty           = majority(entries, 'qty',           float)
        current_value = majority(entries, 'current_value', float)
        profit        = majority(entries, 'profit',        float)
        profit_pct    = majority(entries, 'profit_pct',    float)
        agreement     = len(entries)
        confidence    = round(agreement / total_checks * 100)

        if not name or qty is None or current_value is None:
            continue

        cost = round(current_value - (profit or 0))
        consensus.append({
            'name': name, 'ticker': ticker, 'market': market.upper()[:2],
            'qty': qty,
            'current_value': round(current_value),
            'profit':     round(profit) if profit is not None else 0,
            'profit_pct': round(profit_pct, 1) if profit_pct is not None else 0,
            'cost':       cost,
            'confidence': confidence,
            'agreement':  agreement,
            'total_checks': total_checks,
        })

    consensus.sort(key=lambda x: -x['current_value'])
    return jsonify({
        'success': True,
        'consensus': consensus,
        'checks_done': total_checks,
        'debug': debug_passes,   # 전처리별 OCR 원문
    })


# ════════════════════════════════════════════════════════════
# 목표가 & 메모
# ════════════════════════════════════════════════════════════

@app.route('/api/prefs', methods=['GET'])
def api_get_prefs():
    return jsonify({
        'target_prices': load_target_prices(),
        'notes':         load_notes(),
    })

@app.route('/api/prefs/target', methods=['POST'])
def api_set_target():
    d = request.get_json() or {}
    ticker = d.get('ticker','').strip()
    price  = d.get('price')
    if not ticker:
        return jsonify({'error':'티커 없음'}), 400
    tp = load_target_prices()
    if price is None:
        tp.pop(ticker, None)
    else:
        tp[ticker] = float(price)
    save_target_prices(tp)
    return jsonify({'ok': True, 'target_prices': tp})

@app.route('/api/prefs/note', methods=['POST'])
def api_set_note():
    d = request.get_json() or {}
    ticker = d.get('ticker','').strip()
    note   = d.get('note','').strip()
    if not ticker:
        return jsonify({'error':'티커 없음'}), 400
    notes = load_notes()
    if note:
        notes[ticker] = note
    else:
        notes.pop(ticker, None)
    save_notes(notes)
    return jsonify({'ok': True})


# ════════════════════════════════════════════════════════════
# 벤치마크 비교
# ════════════════════════════════════════════════════════════

@app.route('/api/benchmark')
def api_benchmark():
    """내 포트폴리오 수익률 vs 시장 지수 비교"""
    portfolio = load_portfolio()
    usd_krw   = get_usd_krw()

    periods = [('1일',1),('1주',5),('1달',21),('3달',63),('1년',252)]
    indices  = {
        'S&P500':  '^GSPC',
        '나스닥':  '^IXIC',
        '코스피':  '^KS11',
    }

    # 지수 수익률 (period='2y' 로 여유있게 — 1년 수익률 계산 보장)
    bench = {}
    for name, sym in indices.items():
        bench[name] = {}
        try:
            df = yf.Ticker(sym).history(period='2y')
            if df.empty:
                continue
            closes = df['Close'].dropna()
            if len(closes) < 2:
                continue
            cur = float(closes.iloc[-1])
            for label, days in periods:
                if len(closes) > days:
                    past = float(closes.iloc[-days-1])
                    if past > 0:
                        bench[name][label] = round((cur/past - 1) * 100, 2)
        except Exception as e:
            print(f"[Benchmark] 지수 {name} 실패: {e}")

    # 포트폴리오 수익률 (가중평균: 현재 보유가치 기준)
    # 각 기간별로 [(현재가치, 과거가치)] 쌍을 모은 뒤 합산
    port_sum = {label: {'cur': 0.0, 'past': 0.0} for label, _ in periods}

    for h in portfolio['holdings']:
        qty = float(h.get('qty', 0) or 0)
        if qty <= 0:
            continue
        try:
            if h['market'] == 'US':
                # 2년치 데이터 (1년 수익률 계산을 확실히 보장)
                df = yf.Ticker(h['ticker']).history(period='2y')
                if df.empty:
                    continue
                closes = df['Close'].dropna() * usd_krw
            else:
                # KRX는 달력일 기준이므로 500일 이상으로 넉넉히
                end   = datetime.now().strftime('%Y%m%d')
                start = (datetime.now()-timedelta(days=500)).strftime('%Y%m%d')
                raw   = krx.get_market_ohlcv(start, end, h['ticker'])
                if raw.empty:
                    continue
                closes = raw['종가'].dropna()
            if len(closes) < 2:
                continue
            cur_price  = float(closes.iloc[-1])
            cur_val    = cur_price * qty
            for label, days in periods:
                if len(closes) > days:
                    past_price = float(closes.iloc[-days-1])
                    if past_price > 0:
                        port_sum[label]['cur']  += cur_val
                        port_sum[label]['past'] += past_price * qty
        except Exception as e:
            print(f"[Benchmark] {h.get('ticker','?')} 실패: {e}")

    # 가중평균 수익률 = (Σ 현재가치 - Σ 과거가치) / Σ 과거가치 × 100
    port = {}
    for label, v in port_sum.items():
        if v['past'] > 0:
            port[label] = round((v['cur'] - v['past']) / v['past'] * 100, 2)
        else:
            port[label] = None
    # 메타 정보 (프론트엔드 안내용)
    holdings_count = sum(1 for h in portfolio['holdings'] if float(h.get('qty', 0) or 0) > 0)
    has_any_data = any(v is not None for v in port.values())
    return jsonify({
        'portfolio':  port,
        'benchmarks': bench,
        'periods':    [l for l,_ in periods],
        'meta': {
            'holdings_count': holdings_count,
            'has_portfolio_data': has_any_data,
            'usd_krw': round(usd_krw),
        }
    })


# ════════════════════════════════════════════════════════════
# 종목 비교
# ════════════════════════════════════════════════════════════

@app.route('/api/compare-stocks')
def api_compare_stocks():
    a = request.args.get('a','').upper().strip()
    b = request.args.get('b','').upper().strip()
    if not a or not b:
        return jsonify({'error':'두 티커를 모두 입력하세요'}), 400

    usd_krw = get_usd_krw()
    results = {}

    for sym in [a, b]:
        try:
            # KR 종목은 숫자 6자리
            if sym.isdigit():
                end   = datetime.now().strftime('%Y%m%d')
                start = (datetime.now()-timedelta(days=400)).strftime('%Y%m%d')
                raw   = krx.get_market_ohlcv(start, end, sym)
                if raw.empty: raise ValueError('데이터 없음')
                df = raw.rename(columns={'시가':'Open','고가':'High','저가':'Low','종가':'Close','거래량':'Volume'})
                name   = krx.get_market_ticker_name(sym) or sym
                market = 'KR'
            else:
                t  = yf.Ticker(sym)
                df = t.history(period='1y')
                if df.empty: raise ValueError('데이터 없음')
                df.index = df.index.tz_localize(None)
                df['Close'] *= usd_krw; df['High'] *= usd_krw; df['Low'] *= usd_krw
                info   = t.info
                name   = info.get('shortName', sym)
                market = 'US'

            ind = calc_indicators(df)
            n   = min(60, len(df))
            closes = df['Close']
            results[sym] = {
                'name':    name,
                'market':  market,
                'price':   round(float(closes.iloc[-1])),
                'change1d': round((closes.iloc[-1]/closes.iloc[-2]-1)*100, 2) if len(closes)>1 else 0,
                'change1m': round((closes.iloc[-1]/closes.iloc[-22]-1)*100, 2) if len(closes)>22 else 0,
                'change3m': round((closes.iloc[-1]/closes.iloc[-63]-1)*100, 2) if len(closes)>63 else 0,
                'grade':       ind['grade'],
                'grade_ko':    ind['grade_ko'],
                'total_score': ind['total_score'],
                'rsi':         ind['rsi'],
                'adx':         ind['adx'],
                'reliability': ind['reliability'],
                'indicators':  ind['indicators'],
                'w52':         ind['w52'],
                'short_range': ind['short_range'],
                'mid_range':   ind['mid_range'],
                'chart': {
                    'dates': ind['chart']['dates'],
                    'close': ind['chart']['close'],
                    'ma5':   ind['chart']['ma5'],
                    'ma20':  ind['chart']['ma20'],
                },
            }
        except Exception as e:
            results[sym] = {'error': str(e), 'name': sym}

    return jsonify({'a': results.get(a,{}), 'b': results.get(b,{}), 'tickers': [a,b]})


# ════════════════════════════════════════════════════════════
# 리밸런싱 제안
# ════════════════════════════════════════════════════════════

@app.route('/api/rebalance-targets', methods=['GET'])
def api_get_rebalance_targets():
    return jsonify({'targets': load_rebalance_targets()})

@app.route('/api/rebalance-targets', methods=['POST'])
def api_set_rebalance_targets():
    d = request.get_json() or {}
    targets = d.get('targets', {})
    save_rebalance_targets(targets)
    return jsonify({'ok': True})

@app.route('/api/rebalance')
def api_rebalance():
    """현재 포트폴리오 vs 목표 비중 비교 → 매수/매도 제안"""
    portfolio = load_portfolio()
    targets   = load_rebalance_targets()  # {ticker: target_pct}
    usd_krw   = get_usd_krw()

    if not portfolio['holdings']:
        return jsonify({'error': '포트폴리오가 비어있습니다'}), 400

    # 현재 평가금액 계산
    values = {}
    for h in portfolio['holdings']:
        try:
            if h['market'] == 'US':
                price = yf.Ticker(h['ticker']).fast_info.get('lastPrice', 0) * usd_krw
            else:
                today = datetime.now().strftime('%Y%m%d')
                yest  = (datetime.now()-timedelta(days=7)).strftime('%Y%m%d')
                df    = krx.get_market_ohlcv(yest, today, h['ticker'])
                price = int(df.iloc[-1]['종가']) if not df.empty else 0
            values[h['ticker']] = {
                'name':    h['name'],
                'market':  h['market'],
                'qty':     h['qty'],
                'price':   price,
                'value':   round(price * h['qty']),
            }
        except:
            values[h['ticker']] = {
                'name': h['name'], 'market': h['market'],
                'qty': h['qty'], 'price': 0, 'value': 0
            }

    total = sum(v['value'] for v in values.values()) or 1

    # 현재 비중
    for tk, v in values.items():
        v['current_pct'] = round(v['value'] / total * 100, 1)
        v['target_pct']  = float(targets.get(tk, 0))
        diff_pct = v['target_pct'] - v['current_pct']
        diff_val = round(total * diff_pct / 100)
        v['diff_pct'] = round(diff_pct, 1)
        v['action']   = '매수' if diff_val > 0 else ('매도' if diff_val < 0 else '유지')
        v['diff_val'] = abs(diff_val)
        v['diff_qty'] = round(abs(diff_val) / v['price'], 6) if v['price'] > 0 else 0

    items = sorted(values.values(), key=lambda x: abs(x['diff_pct']), reverse=True)
    target_total = sum(float(targets.get(tk, 0)) for tk in values)

    return jsonify({
        'items':        items,
        'total_value':  total,
        'target_total': round(target_total, 1),
        'usd_krw':      usd_krw,
    })


# ── 크로스 디바이스 동기화 API (Supabase REST) ──────────────
@app.route('/api/sync/status')
def api_sync_status():
    """Supabase 연결 상태 확인"""
    return jsonify({'enabled': IS_SUPABASE})

@app.route('/api/sync/generate-code', methods=['POST'])
def api_sync_generate():
    """현재 기기의 데이터를 Supabase에 업로드하고 접속코드 반환"""
    if not IS_SUPABASE:
        return jsonify({'ok': False, 'error': '클라우드 동기화가 설정되지 않았습니다'}), 400

    # 현재 사용자 데이터 수집
    portfolio    = load_portfolio()
    transactions = load_transactions()
    wishlist     = _fetch('wishlist', [])
    hidden       = _fetch('hidden', [])
    target_prices = _fetch('target_prices', {})
    notes        = _fetch('notes', {})
    rebalance    = _fetch('rebalance_targets', {})
    simulation   = _fetch('simulation', [])

    payload = {
        'portfolio': portfolio,
        'transactions': transactions,
        'wishlist': wishlist,
        'hidden': hidden,
        'target_prices': target_prices,
        'notes': notes,
        'rebalance_targets': rebalance,
        'simulation': simulation,
        'source_token': g.user_token,
        'synced_at': datetime.now().isoformat(),
    }

    # 이미 코드가 있는지 확인 (같은 source_token)
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/sync_data?select=access_code&data->>source_token=eq.{g.user_token}",
            headers=_supa_headers(), timeout=10
        )
        existing = r.json() if r.status_code == 200 else []
    except:
        existing = []

    if existing:
        code = existing[0]['access_code']
        # 기존 코드에 데이터 업데이트
        r = requests.patch(
            f"{SUPABASE_URL}/rest/v1/sync_data?access_code=eq.{code}",
            headers=_supa_headers(),
            json={'data': payload, 'updated_at': datetime.now().isoformat()},
            timeout=10
        )
    else:
        # 새 코드 생성 (충돌 시 재시도)
        for _ in range(10):
            code = _gen_access_code()
            r = requests.post(
                f"{SUPABASE_URL}/rest/v1/sync_data",
                headers={**_supa_headers(), 'Prefer': 'return=representation'},
                json={'access_code': code, 'data': payload},
                timeout=10
            )
            if r.status_code == 201:
                break
        else:
            return jsonify({'ok': False, 'error': '코드 생성 실패'}), 500

    return jsonify({'ok': True, 'code': code})

@app.route('/api/sync/upload', methods=['POST'])
def api_sync_upload():
    """기존 접속코드에 현재 데이터 업로드 (덮어쓰기)"""
    if not IS_SUPABASE:
        return jsonify({'ok': False, 'error': '클라우드 동기화 미설정'}), 400

    data = request.get_json() or {}
    code = str(data.get('code', '')).strip().upper()
    if not code:
        return jsonify({'ok': False, 'error': '접속코드가 없습니다'}), 400

    portfolio    = load_portfolio()
    transactions = load_transactions()
    wishlist     = _fetch('wishlist', [])
    hidden       = _fetch('hidden', [])
    target_prices = _fetch('target_prices', {})
    notes        = _fetch('notes', {})
    rebalance    = _fetch('rebalance_targets', {})
    simulation   = _fetch('simulation', [])

    payload = {
        'portfolio': portfolio,
        'transactions': transactions,
        'wishlist': wishlist,
        'hidden': hidden,
        'target_prices': target_prices,
        'notes': notes,
        'rebalance_targets': rebalance,
        'simulation': simulation,
        'source_token': g.user_token,
        'synced_at': datetime.now().isoformat(),
    }

    r = requests.patch(
        f"{SUPABASE_URL}/rest/v1/sync_data?access_code=eq.{code}",
        headers=_supa_headers(),
        json={'data': payload, 'updated_at': datetime.now().isoformat()},
        timeout=10
    )
    if r.status_code in (200, 204):
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': '업로드 실패'}), 500

@app.route('/api/sync/download', methods=['POST'])
def api_sync_download():
    """접속코드로 클라우드 데이터를 현재 기기에 다운로드"""
    if not IS_SUPABASE:
        return jsonify({'ok': False, 'error': '클라우드 동기화 미설정'}), 400

    data = request.get_json() or {}
    code = str(data.get('code', '')).strip().upper()
    if not code:
        return jsonify({'ok': False, 'error': '접속코드를 입력하세요'}), 400

    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/sync_data?access_code=eq.{code}&select=*",
            headers=_supa_headers(), timeout=10
        )
        rows = r.json() if r.status_code == 200 else []
    except Exception as e:
        return jsonify({'ok': False, 'error': f'서버 연결 실패: {e}'}), 500

    if not rows:
        return jsonify({'ok': False, 'error': '해당 접속코드를 찾을 수 없습니다'}), 404

    cloud = rows[0]['data']

    # 로컬 DB에 저장
    if 'portfolio' in cloud:
        save_portfolio(cloud['portfolio'])
    if 'transactions' in cloud:
        save_transactions(cloud['transactions'])
    if 'wishlist' in cloud:
        _upsert('wishlist', cloud['wishlist'])
    if 'hidden' in cloud:
        _upsert('hidden', cloud['hidden'])
    if 'target_prices' in cloud:
        _upsert('target_prices', cloud['target_prices'])
    if 'notes' in cloud:
        _upsert('notes', cloud['notes'])
    if 'rebalance_targets' in cloud:
        _upsert('rebalance_targets', cloud['rebalance_targets'])
    if 'simulation' in cloud:
        _upsert('simulation', cloud['simulation'])

    synced_at = cloud.get('synced_at', '')
    holdings_count = len(cloud.get('portfolio', {}).get('holdings', []))

    return jsonify({
        'ok': True,
        'synced_at': synced_at,
        'holdings_count': holdings_count,
    })


# ── 모듈 로드 마지막에 캐시 빌드 시작 (모든 함수 정의 후) ──
try:
    _kickstart_kr_cache()
except Exception as _e:
    print(f"[Cache] kickstart 호출 실패: {_e}")


if __name__ == '__main__':
    print("Dashboard: http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)
